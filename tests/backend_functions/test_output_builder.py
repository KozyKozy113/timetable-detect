"""output_builder.py のリファクタリング対象関数のテスト"""

import os

import pandas as pd

from backend_functions import output_builder as _output


# ---------------------------------------------------------------------------
# find_or_create_stage_id
# ---------------------------------------------------------------------------

def test_find_or_create_stage_id_existing_returns_existing_id():
    stage_master = {
        0: {"ステージ名": "メインステージ", "特典会フラグ": False, "表示順": 0, "非活性化フラグ": False},
        1: {"ステージ名": "サブステージ", "特典会フラグ": False, "表示順": 1, "非活性化フラグ": False},
    }

    stage_id, next_id = _output.find_or_create_stage_id(
        stage_master, "サブステージ", False, next_id=2,
    )

    assert stage_id == 1
    assert next_id == 2  # next_idは変化しない
    # stage_masterは変化しない
    assert stage_master[1]["ステージ名"] == "サブステージ"


def test_find_or_create_stage_id_new_appends_and_increments():
    stage_master = {
        0: {"ステージ名": "メインステージ", "特典会フラグ": False, "表示順": 0, "非活性化フラグ": False},
    }

    stage_id, next_id = _output.find_or_create_stage_id(
        stage_master, "新ブース", True, next_id=1,
    )

    assert stage_id == 1
    assert next_id == 2
    assert stage_master[1]["ステージ名"] == "新ブース"
    assert stage_master[1]["特典会フラグ"] is True
    # 新規ステージの表示順は既存最大値 + 1
    assert stage_master[1]["表示順"] == 1
    assert stage_master[1]["非活性化フラグ"] is False


def test_find_or_create_stage_id_from_empty_master():
    stage_master = {}

    stage_id, next_id = _output.find_or_create_stage_id(
        stage_master, "初登場ステージ", False, next_id=0,
    )

    assert stage_id == 0
    assert next_id == 1
    assert stage_master[0]["ステージ名"] == "初登場ステージ"
    assert stage_master[0]["特典会フラグ"] is False
    assert stage_master[0]["表示順"] == 0
    assert stage_master[0]["非活性化フラグ"] is False


def test_find_or_create_stage_id_existing_stage_id_hits_master():
    """existing_stage_id 指定時はステージ名一致を経由せずIDで引き当て、
    ステージ名が異なれば**マスタを更新**する。"""
    stage_master = {
        0: {"ステージ名": "旧名", "特典会フラグ": False, "表示順": 0, "非活性化フラグ": False},
    }

    stage_id, next_id = _output.find_or_create_stage_id(
        stage_master, "新名", False, next_id=1, existing_stage_id=0,
    )

    assert stage_id == 0
    assert next_id == 1
    # マスタ側のステージ名が新しい名前に更新される(編集モードの反映)
    assert stage_master[0]["ステージ名"] == "新名"


def test_find_or_create_stage_id_existing_stage_id_handles_string_keys():
    """CSV 経由で読み込まれた stage_master はキーが文字列になっているケース。"""
    stage_master = {
        "0": {"ステージ名": "メインステージ", "特典会フラグ": False, "表示順": 0, "非活性化フラグ": False},
    }

    stage_id, next_id = _output.find_or_create_stage_id(
        stage_master, "メインステージ", False, next_id=1, existing_stage_id=0,
    )

    assert stage_id == 0
    assert next_id == 1
    # 文字列キーでも引き当て成功
    assert stage_master["0"]["ステージ名"] == "メインステージ"


# ---------------------------------------------------------------------------
# load_existing_masters
# ---------------------------------------------------------------------------

def test_load_existing_masters_pre_confirm(project_pre_confirm):
    pj_path, _ = project_pre_confirm
    output_path = os.path.join(pj_path, "event_1")

    stage_master, next_stage_id, idolname_df, next_artist_id = \
        _output.load_existing_masters(output_path)

    assert stage_master == {}
    assert next_stage_id == 0
    assert list(idolname_df.columns) == ["グループ名_採用"]
    assert idolname_df.index.name == "グループID"
    assert len(idolname_df) == 0
    assert next_artist_id == 0


def test_load_existing_masters_post_confirm(project_post_confirm):
    pj_path, _ = project_post_confirm
    output_path = os.path.join(pj_path, "event_1")

    stage_master, next_stage_id, idolname_df, next_artist_id = \
        _output.load_existing_masters(output_path)

    # master_stage.csv に "0,メインステージ,False" と "1,特典会ブース,True"
    # json.loads(df.T.to_json()) によりキーは文字列になる(既存挙動)
    # 後方互換: 表示順 / 非活性化フラグ が無い旧CSVには自動補完される
    assert stage_master["0"]["ステージ名"] == "メインステージ"
    assert stage_master["0"]["特典会フラグ"] is False
    assert stage_master["0"]["表示順"] == 0
    assert stage_master["0"]["非活性化フラグ"] is False
    assert stage_master["1"]["ステージ名"] == "特典会ブース"
    assert stage_master["1"]["特典会フラグ"] is True
    assert stage_master["1"]["表示順"] == 1
    assert stage_master["1"]["非活性化フラグ"] is False
    assert next_stage_id == 2

    # master_idolname.csv は "グループ名" カラムを "グループ名_採用" にリネーム
    assert list(idolname_df.columns) == ["グループ名_採用"]
    assert idolname_df.index.name == "グループID"
    assert idolname_df.loc[0, "グループ名_採用"] == "アルファ"
    assert idolname_df.loc[1, "グループ名_採用"] == "ベータ"
    assert next_artist_id == 2


# ---------------------------------------------------------------------------
# build_event_output
# ---------------------------------------------------------------------------

EXPECTED_LIVE_COLUMNS = [
    "ライブ_from", "ライブ_to", "ライブ_長さ(分)", "グループID", "ステージID",
    "グループ名_raw", "グループ名", "ステージ名", "備考",
    "コラボタイトル",
]


def test_build_event_output_pre_confirm(project_pre_confirm):
    pj_path, project_info_json = project_pre_confirm

    result = _output.build_event_output(pj_path, "event_1", 0, project_info_json)

    assert result is not None
    df_stage = result["stage"]
    df_idolname = result["idolname"]
    df_live = result["live"]

    # ステージマスタ: 0=メインステージ, 1=特典会ブース
    assert df_stage.index.name == "ステージID"
    assert sorted(df_stage.index.tolist()) == [0, 1]
    assert df_stage.loc[0, "ステージ名"] == "メインステージ"
    assert not df_stage.loc[0, "特典会フラグ"]
    assert df_stage.loc[1, "ステージ名"] == "特典会ブース"
    assert bool(df_stage.loc[1, "特典会フラグ"]) is True

    # アーティストマスタ: ソート順で 0=アルファ, 1=ガンマ, 2=ベータ
    assert df_idolname.index.name == "グループID"
    assert len(df_idolname) == 3
    assert sorted(df_idolname["グループ名_採用"].tolist()) == ["アルファ", "ガンマ", "ベータ"]

    # 出番データ: 5行(ライブ3 + 特典会2)
    assert list(df_live.columns) == EXPECTED_LIVE_COLUMNS
    assert df_live.index.name == "出番ID"
    assert len(df_live) == 5
    assert sorted(df_live.index.tolist()) == [0, 1, 2, 3, 4]


def test_build_event_output_post_confirm(project_post_confirm):
    pj_path, project_info_json = project_post_confirm

    result = _output.build_event_output(pj_path, "event_1", 0, project_info_json)

    assert result is not None
    df_idolname = result["idolname"]
    df_live = result["live"]

    # アーティストマスタ: 既存 0=アルファ, 1=ベータ + 新規 2=ガンマ
    assert df_idolname.loc[0, "グループ名_採用"] == "アルファ"
    assert df_idolname.loc[1, "グループ名_採用"] == "ベータ"
    assert df_idolname.loc[2, "グループ名_採用"] == "ガンマ"

    # 出番ID: 既存(0,1,2,3)維持 + ガンマに新規ID 4
    assert sorted(df_live.index.tolist()) == [0, 1, 2, 3, 4]
    gamma_row = df_live[df_live["グループ名"] == "ガンマ"]
    assert len(gamma_row) == 1
    assert gamma_row.index[0] == 4

    # アルファの既存IDが維持されているか確認
    alpha_live = df_live[(df_live["グループ名"] == "アルファ") & (df_live["ステージ名"] == "メインステージ")]
    assert len(alpha_live) == 1
    assert alpha_live.index[0] == 0


def test_build_event_output_tokutenkai_heiki(project_tokutenkai_heiki):
    pj_path, project_info_json = project_tokutenkai_heiki

    result = _output.build_event_output(pj_path, "event_1", 0, project_info_json)

    assert result is not None
    df_stage = result["stage"]
    df_live = result["live"]

    # ステージマスタ: メインステージ(ライブ) + ブースX + ブースY = 3行
    stage_names = df_stage["ステージ名"].tolist()
    assert "メインステージ" in stage_names
    assert "ブースX" in stage_names
    assert "ブースY" in stage_names

    # メインステージは特典会フラグFalse、ブースはTrue
    main_row = df_stage[df_stage["ステージ名"] == "メインステージ"].iloc[0]
    assert not main_row["特典会フラグ"]
    booth_rows = df_stage[df_stage["ステージ名"].isin(["ブースX", "ブースY"])]
    assert booth_rows["特典会フラグ"].all()

    # 出番データ: ライブ3 + 特典会3 = 6行
    assert len(df_live) == 6
    # ライブ行: ステージ名がメインステージ
    live_rows = df_live[df_live["ステージ名"] == "メインステージ"]
    assert len(live_rows) == 3
    # 特典会行: ステージ名がブースX/ブースY
    tokutenkai_rows = df_live[df_live["ステージ名"].isin(["ブースX", "ブースY"])]
    assert len(tokutenkai_rows) == 3


def test_build_event_output_no_data(tmp_path):
    """JSON ファイルが1つも存在しないイベントは None を返す"""
    project_info_json = {
        "project_name": "empty",
        "event_num": 1,
        "event_detail": [
            {
                "event_no": 0,
                "event_name": "event_1",
                "timetables": [
                    {
                        "image_no": 0,
                        "dir_name": "ライブ",
                        "display_name": "ライブ",
                        "kind": "live",
                        "format": "通常",
                        "stage_num": 1,
                        "stage_list": [
                            {"stage_no": 0, "stage_name": "main", "kind": "live"}
                        ],
                    }
                ],
            }
        ],
    }
    os.makedirs(tmp_path / "event_1" / "ライブ", exist_ok=True)

    result = _output.build_event_output(
        str(tmp_path), "event_1", 0, project_info_json,
    )

    assert result is None


# ---------------------------------------------------------------------------
# build_all_event_outputs
# ---------------------------------------------------------------------------

def test_build_all_event_outputs_returns_data_per_event(project_pre_confirm):
    pj_path, project_info_json = project_pre_confirm

    result = _output.build_all_event_outputs(pj_path, project_info_json)

    assert "event_1" in result
    assert "stage" in result["event_1"]
    assert "idolname" in result["event_1"]
    assert "live" in result["event_1"]


def test_build_all_event_outputs_preserves_empty_for_no_data(tmp_path):
    """データなしイベントは空dictとして保持(キーは残す)"""
    def _live_entry():
        return {
            "image_no": 0,
            "dir_name": "ライブ",
            "display_name": "ライブ",
            "kind": "live",
            "format": "通常",
            "stage_num": 1,
            "stage_list": [{"stage_no": 0, "stage_name": "main", "kind": "live"}],
        }

    project_info_json = {
        "project_name": "two_events",
        "event_num": 2,
        "event_detail": [
            {
                "event_no": 0,
                "event_name": "event_1",
                "timetables": [_live_entry()],
            },
            {
                "event_no": 1,
                "event_name": "event_2",
                "timetables": [_live_entry()],
            },
        ],
    }
    os.makedirs(tmp_path / "event_1" / "ライブ", exist_ok=True)
    os.makedirs(tmp_path / "event_2" / "ライブ", exist_ok=True)

    result = _output.build_all_event_outputs(str(tmp_path), project_info_json)

    assert set(result.keys()) == {"event_1", "event_2"}
    assert result["event_1"] == {}
    assert result["event_2"] == {}


# ---------------------------------------------------------------------------
# 空dictイベント混在時の後段関数の挙動
# ---------------------------------------------------------------------------

def _make_two_event_project(tmp_path):
    """event_1=データあり, event_2=空dict のテストデータを作る"""
    def _live_entry():
        return {
            "image_no": 0,
            "dir_name": "ライブ",
            "display_name": "ライブ",
            "kind": "live",
            "format": "通常",
            "stage_num": 1,
            "stage_list": [{"stage_no": 0, "stage_name": "メイン", "kind": "live"}],
        }

    project_info_json = {
        "project_name": "p",
        "event_num": 2,
        "event_detail": [
            {
                "event_no": 0, "event_name": "event_1",
                "timetables": [_live_entry()],
            },
            {
                "event_no": 1, "event_name": "event_2",
                "timetables": [_live_entry()],
            },
        ],
    }
    os.makedirs(tmp_path / "event_1", exist_ok=True)
    os.makedirs(tmp_path / "event_2", exist_ok=True)
    output_df = {
        "event_1": {
            "stage": pd.DataFrame(
                {"ステージ名": ["メイン"], "特典会フラグ": [False]},
                index=pd.Index([0], name="ステージID"),
            ),
            "idolname": pd.DataFrame(
                {"グループ名_採用": ["ZZグループ"]},
                index=pd.Index([0], name="グループID"),
            ),
            "live": pd.DataFrame(
                {
                    "ライブ_from": ["11:00"], "ライブ_長さ(分)": [30],
                    "グループID": [0], "ステージID": [0],
                    "グループ名_raw": ["ZZグループ"], "グループ名": ["ZZグループ"],
                    "ステージ名": ["メイン"], "備考": [""],
                },
                index=pd.Index([0], name="出番ID"),
            ),
        },
        "event_2": {},
    }
    return project_info_json, output_df


def test_determine_id_master_skips_empty_event(tmp_path):
    project_info_json, output_df = _make_two_event_project(tmp_path)

    _output.determine_id_master(output_df, str(tmp_path), project_info_json)

    # event_1 はCSV出力される
    assert (tmp_path / "event_1" / "master_stage.csv").exists()
    assert (tmp_path / "event_1" / "master_idolname.csv").exists()
    # event_2 はスキップされてCSVは作られない(KeyErrorも起きない)
    assert not (tmp_path / "event_2" / "master_stage.csv").exists()


def test_export_excel_skips_empty_event(tmp_path):
    project_info_json, output_df = _make_two_event_project(tmp_path)
    event_list = ["event_1", "event_2"]

    output_path = _output.export_excel(output_df, str(tmp_path), event_list)

    # 出力ファイルが作られ、event_1 のシートだけが含まれる
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    assert "event_1" in wb.sheetnames
    assert "event_2" not in wb.sheetnames


# ---------------------------------------------------------------------------
# Excel 出力: stage 詳細列 / コラボ統合
# ---------------------------------------------------------------------------

def _sheet_cells(ws) -> list[list]:
    return [
        [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, ws.max_row + 1)
    ]


def _make_collab_output_df():
    """出番ID=0 がコラボ (同一ステージ・同一時刻の2グループ) のサンプル。"""
    stage = pd.DataFrame(
        {
            "ステージ名": ["メイン", "サブ"],
            "特典会フラグ": [False, False],
            "表示順": [1, 0],
            "非活性化フラグ": [False, False],
            "ステージ名_短縮": ["M", "S"],
            "カラー名": ["stage-red", "stage-blue"],
        },
        index=pd.Index([0, 1], name="ステージID"),
    )
    idol = pd.DataFrame(
        {"グループ名_採用": ["A", "B", "C"]},
        index=pd.Index([0, 1, 2], name="グループID"),
    )
    live = pd.DataFrame(
        {
            "ライブ_from": ["11:00", "11:00", "12:00"],
            "ライブ_to": ["11:30", "11:30", "12:30"],
            "ライブ_長さ(分)": [30, 30, 30],
            "グループID": [0, 1, 2],
            "ステージID": [0, 0, 1],
            "グループ名_raw": ["A", "B", "C"],
            "グループ名": ["A", "B", "C"],
            "ステージ名": ["メイン", "メイン", "サブ"],
            "備考": ["x", "x", "y"],
            "コラボタイトル": ["コラボ!", "コラボ!", ""],
        },
        index=pd.Index([0, 0, 1], name="出番ID"),
    )
    return {"ev": {"stage": stage, "idolname": idol, "live": live}}


def test_export_excel_stage_includes_detail_columns(tmp_path):
    """stage シートに ステージ名_短縮 / カラー名 / 表示順 が欠落せず出力される。"""
    from openpyxl import load_workbook

    output_df = _make_collab_output_df()
    output_path = _output.export_excel(output_df, str(tmp_path), ["ev"])
    wb = load_workbook(output_path)
    grid = _sheet_cells(wb["ev"])
    headers = grid[0]

    for label in ("ステージID", "ステージ名", "ステージ名_短縮", "カラー名", "表示順"):
        assert label in headers, f"{label} がヘッダに存在しない"

    # 表示順=0 のサブが先頭、表示順=1 のメインが次 (表示順ソート)
    sid_col = headers.index("ステージID")
    short_col = headers.index("ステージ名_短縮")
    color_col = headers.index("カラー名")
    order_col = headers.index("表示順")
    first_stage = grid[1]
    assert first_stage[sid_col] == 1
    assert first_stage[short_col] == "S"
    assert first_stage[color_col] == "stage-blue"
    assert first_stage[order_col] == 0


def test_export_excel_merges_collab_rows(tmp_path):
    """コラボ出番が1行に統合され、グループIDがカンマ区切りになる。"""
    from openpyxl import load_workbook

    output_df = _make_collab_output_df()
    output_path = _output.export_excel(output_df, str(tmp_path), ["ev"])
    wb = load_workbook(output_path)
    grid = _sheet_cells(wb["ev"])
    headers = grid[0]

    turn_col = headers.index("出番ID")
    # live ブロックの グループID 列 (idolname 側の グループID と区別するため出番ID以降)
    gid_col = headers.index("グループID", turn_col)
    gname_col = headers.index("グループ名", turn_col)

    # 出番ID 行 (None でない出番ID を持つ行) を収集
    live_rows = [row for row in grid[1:] if row[turn_col] is not None]
    by_turn = {row[turn_col]: row for row in live_rows}

    # 出番ID=0 はコラボ → 1行に統合、グループID/グループ名がカンマ区切り
    assert 0 in by_turn
    assert by_turn[0][gid_col] == "0,1"
    assert by_turn[0][gname_col] == "A,B"
    # 出番ID=0 は1行だけ (重複行が無い)
    assert sum(1 for row in live_rows if row[turn_col] == 0) == 1

    # 単独出番 (出番ID=1) は グループID が数値のまま
    assert by_turn[1][gid_col] == 2


def test_export_excel_includes_stella_metadata(tmp_path):
    """metadata_by_event/project_meta を渡すと Stella メタデータブロックが出力される。"""
    from openpyxl import load_workbook

    output_df = _make_collab_output_df()
    metadata_by_event = {
        "ev": {
            "date": "20260504", "liveId": 547, "bundleId": "547",
            "jsonVersion": 13, "openTime": "12", "closeTime": "23",
            "notificationVersion": "2", "notification": "お知らせ本文",
        }
    }
    project_meta = {"liveName": "テストライブ", "genre": 2, "release": 0, "pref": 13}
    output_path = _output.export_excel(
        output_df, str(tmp_path), ["ev"],
        metadata_by_event=metadata_by_event, project_meta=project_meta,
    )
    wb = load_workbook(output_path)
    grid = _sheet_cells(wb["ev"])

    # 縦持ち (項目, 値) を dict 化して検証
    kv = {row[0]: row[1] for row in grid[1:] if row[0] is not None}
    assert kv["ライブ名"] == "テストライブ"
    assert kv["liveId"] == 547
    assert kv["開始時"] == "12"
    assert kv["終了時"] == "23"
    assert kv["お知らせメッセージ"] == "お知らせ本文"

    # メタデータを渡さない場合はメタデータ列が出ない (後方互換)
    output_path2 = _output.export_excel(output_df, str(tmp_path), ["ev"])
    grid2 = _sheet_cells(load_workbook(output_path2)["ev"])
    assert "Stellaメタデータ" not in grid2[0]


# ---------------------------------------------------------------------------
# kind ベース 特典会フラグ導出の回帰テスト (Phase 6)
# ---------------------------------------------------------------------------

import json as _json


def _write_stage_json(stage_dir, group_name: str, time_from: str = "11:00") -> None:
    os.makedirs(stage_dir, exist_ok=True)
    payload = {
        "タイムテーブル": [
            {
                "グループ名": group_name,
                "グループ名_採用": group_name,
                "ライブステージ": {"from": time_from, "to": "11:30"},
                "備考": "",
            }
        ]
    }
    with open(os.path.join(stage_dir, "stage_0.json"), "w", encoding="utf-8") as f:
        _json.dump(payload, f, ensure_ascii=False)


def _make_single_image_project(tmp_path, dir_name: str, kind: str,
                               include_format: bool = True) -> dict:
    """1 イベント / 1 画像 / 1 ステージのプロジェクトを作る"""
    entry = {
        "image_no": 0,
        "dir_name": dir_name,
        "display_name": dir_name,
        "kind": kind,
        "stage_num": 1,
        "stage_list": [{"stage_no": 0, "stage_name": "ステージA", "kind": kind}],
    }
    if include_format:
        entry["format"] = "通常"
    project_info_json = {
        "project_name": "p",
        "event_num": 1,
        "event_detail": [
            {
                "event_no": 0, "event_name": "event_1",
                "timetables": [entry],
            }
        ],
    }
    _write_stage_json(os.path.join(tmp_path, "event_1", dir_name), "アルファ")
    return project_info_json


def test_custom_dir_name_with_kind_live_is_not_tokutenkai(tmp_path):
    """カスタム dir_name + kind=live → 特典会フラグ=False。

    旧仕様 (event_type == "特典会" 文字列一致) でも False だったので結論は変わらないが、
    新仕様が kind ベースで判定していることを担保する。
    """
    pij = _make_single_image_project(tmp_path, dir_name="縁日", kind="live")

    result = _output.build_event_output(str(tmp_path), "event_1", 0, pij)

    assert result is not None
    df_stage = result["stage"]
    row = df_stage[df_stage["ステージ名"] == "ステージA"].iloc[0]
    assert not row["特典会フラグ"]


def test_custom_dir_name_with_kind_tokutenkai_is_tokutenkai(tmp_path):
    """カスタム dir_name + kind=tokutenkai → 特典会フラグ=True。

    旧仕様では event_type=="特典会" の完全一致でしか特典会判定できず、
    "縁日" のようなカスタム名は常にライブ扱いだった。Plan §現状の問題点 1 のバグ修正。
    """
    pij = _make_single_image_project(tmp_path, dir_name="縁日", kind="tokutenkai")

    result = _output.build_event_output(str(tmp_path), "event_1", 0, pij)

    assert result is not None
    df_stage = result["stage"]
    row = df_stage[df_stage["ステージ名"] == "ステージA"].iloc[0]
    assert bool(row["特典会フラグ"]) is True


def test_custom_dir_name_with_kind_heiki_no_format_field(tmp_path):
    """カスタム dir_name + kind=live_tokutenkai_heiki は format フィールドを持たない設計だが、
    build_event_output が動作することを担保する (kind ベースで分岐するため)。
    """
    pij = _make_single_image_project(
        tmp_path, dir_name="特殊併記", kind="live_tokutenkai_heiki",
        include_format=False,
    )
    # 併記タイテ JSON (実フォーマットに合わせる: 各行に「ライブステージ」と「特典会」配列)
    stage_dir = os.path.join(tmp_path, "event_1", "特殊併記")
    payload = {
        "タイムテーブル": [
            {
                "グループ名": "アルファ",
                "グループ名_採用": "アルファ",
                "ライブステージ": {"from": "11:00", "to": "11:30"},
                "特典会": [
                    {"from": "11:40", "to": "12:10", "ブース": "ブースZ"},
                ],
                "備考": "",
            }
        ]
    }
    with open(os.path.join(stage_dir, "stage_0.json"), "w", encoding="utf-8") as f:
        _json.dump(payload, f, ensure_ascii=False)

    result = _output.build_event_output(str(tmp_path), "event_1", 0, pij)
    assert result is not None
    df_stage = result["stage"]
    # 親ステージはライブ枠 → 特典会フラグ False
    main_stage = df_stage[df_stage["ステージ名"] == "ステージA"].iloc[0]
    assert not main_stage["特典会フラグ"]
    # ブースZ は特典会扱い → 特典会フラグ True
    booth = df_stage[df_stage["ステージ名"] == "ブースZ"].iloc[0]
    assert bool(booth["特典会フラグ"]) is True


def _make_heiki_project_with_booths(tmp_path, booth_appearance, stage_id=None):
    """併記 (live_tokutenkai_heiki) 1イベント/1画像を作り、ブースを指定順に登場させる。

    booth_appearance: 各グループの特典会ブース名のリスト (登場順 = 採番の出現順)。
    stage_id: stage_list[0].stage_id に設定する値。None なら未採番、値ありで採番済み。
    """
    stage_entry = {"stage_no": 0, "stage_name": "ステージA", "kind": "live_tokutenkai_heiki"}
    if stage_id is not None:
        stage_entry["stage_id"] = stage_id
    entry = {
        "image_no": 0,
        "dir_name": "ライブ特典会",
        "display_name": "ライブ特典会",
        "kind": "live_tokutenkai_heiki",
        "stage_num": 1,
        "stage_list": [stage_entry],
    }
    pij = {
        "project_name": "p",
        "event_num": 1,
        "event_detail": [
            {"event_no": 0, "event_name": "event_1", "timetables": [entry]},
        ],
    }
    rows = []
    base_h = 11
    for i, booth in enumerate(booth_appearance):
        rows.append({
            "グループ名": f"G{i}",
            "グループ名_採用": f"G{i}",
            "ライブステージ": {"from": f"{base_h + i}:00", "to": f"{base_h + i}:30"},
            "特典会": [{"from": f"{base_h + i}:40", "to": f"{base_h + i}:50", "ブース": booth}],
            "備考": "",
        })
    stage_dir = os.path.join(tmp_path, "event_1", "ライブ特典会")
    os.makedirs(stage_dir, exist_ok=True)
    with open(os.path.join(stage_dir, "stage_0.json"), "w", encoding="utf-8") as f:
        _json.dump({"タイムテーブル": rows}, f, ensure_ascii=False)
    return pij


def test_heiki_unnumbered_booth_ids_sorted_by_name(tmp_path):
    """未採番の併記種別では、特典会ブースIDをブース名順で採番する。

    ブースが Y→X の出現順でも、未採番ならソートされ ブースX < ブースY のID/表示順になる。
    """
    pij = _make_heiki_project_with_booths(tmp_path, ["ブースY", "ブースX"], stage_id=None)

    result = _output.build_event_output(str(tmp_path), "event_1", 0, pij)

    assert result is not None
    df_stage = result["stage"]
    id_x = df_stage[df_stage["ステージ名"] == "ブースX"].index[0]
    id_y = df_stage[df_stage["ステージ名"] == "ブースY"].index[0]
    assert id_x < id_y
    order_x = df_stage.loc[id_x, "表示順"]
    order_y = df_stage.loc[id_y, "表示順"]
    assert order_x < order_y


def test_heiki_numbered_booth_ids_keep_appearance_order(tmp_path):
    """採番済みの併記種別では、新規ブースは出現順のまま採番する (ソートしない)。

    stage_list に stage_id 設定済み (= 種別として採番済み) のとき、
    JSON にブースIDを持たない新規ブースは Y→X の出現順どおり ブースY < ブースX になる。
    """
    pij = _make_heiki_project_with_booths(tmp_path, ["ブースY", "ブースX"], stage_id=0)

    result = _output.build_event_output(str(tmp_path), "event_1", 0, pij)

    assert result is not None
    df_stage = result["stage"]
    id_x = df_stage[df_stage["ステージ名"] == "ブースX"].index[0]
    id_y = df_stage[df_stage["ステージ名"] == "ブースY"].index[0]
    assert id_y < id_x


# ---------------------------------------------------------------------------
# build_duration_distribution / build_group_appearance_count
# ---------------------------------------------------------------------------

def _make_stage_df(rows):
    """rows = [(stage_id, stage_name, tokutenkai_flg), ...]"""
    df = pd.DataFrame(
        {
            "ステージ名": [r[1] for r in rows],
            "特典会フラグ": [r[2] for r in rows],
        },
        index=pd.Index([r[0] for r in rows], name="ステージID"),
    )
    return df


def _make_live_df(rows):
    """rows = [(出番ID, ライブ_長さ(分), グループID, グループ名, ステージID), ...]"""
    df = pd.DataFrame(
        {
            "ライブ_from": ["11:00"] * len(rows),
            "ライブ_長さ(分)": [r[1] for r in rows],
            "グループID": [r[2] for r in rows],
            "ステージID": [r[4] for r in rows],
            "グループ名_raw": [r[3] for r in rows],
            "グループ名": [r[3] for r in rows],
            "ステージ名": [""] * len(rows),
            "備考": [""] * len(rows),
        },
        index=pd.Index([r[0] for r in rows], name="出番ID"),
    )
    return df


def test_build_duration_distribution_splits_by_tokutenkai_flag():
    df_stage = _make_stage_df([
        (0, "メイン", False),
        (1, "ブースA", True),
    ])
    df_live = _make_live_df([
        (0, 20, 0, "A", 0),
        (1, 20, 1, "B", 0),
        (2, 30, 2, "C", 0),
        (3, 20, 0, "A", 1),
        (4, 25, 1, "B", 1),
        (5, 25, 1, "B", 1),
    ])

    result = _output.build_duration_distribution(df_live, df_stage)

    assert result.index.name == "長さ(分)"
    assert list(result.columns) == ["ライブステージ", "特典会ステージ"]
    assert result.loc[20, "ライブステージ"] == 2
    assert result.loc[20, "特典会ステージ"] == 1
    assert result.loc[25, "ライブステージ"] == 0
    assert result.loc[25, "特典会ステージ"] == 2
    assert result.loc[30, "ライブステージ"] == 1
    assert result.loc[30, "特典会ステージ"] == 0
    assert list(result.index) == [20, 25, 30]


def test_build_duration_distribution_empty_returns_empty_with_two_cols():
    df_stage = _make_stage_df([(0, "メイン", False)])
    df_live = _make_live_df([])

    result = _output.build_duration_distribution(df_live, df_stage)

    assert list(result.columns) == ["ライブステージ", "特典会ステージ"]
    assert len(result) == 0
    assert result.index.name == "長さ(分)"


def test_build_duration_distribution_only_live_still_has_tokutenkai_col():
    df_stage = _make_stage_df([(0, "メイン", False)])
    df_live = _make_live_df([
        (0, 30, 0, "A", 0),
        (1, 30, 1, "B", 0),
    ])

    result = _output.build_duration_distribution(df_live, df_stage)

    assert list(result.columns) == ["ライブステージ", "特典会ステージ"]
    assert result.loc[30, "ライブステージ"] == 2
    assert result.loc[30, "特典会ステージ"] == 0


def test_build_group_appearance_count_splits_by_tokutenkai_flag():
    df_stage = _make_stage_df([
        (0, "メイン", False),
        (1, "ブース", True),
    ])
    df_live = _make_live_df([
        (0, 30, 10, "アルファ", 0),
        (1, 30, 10, "アルファ", 0),
        (2, 25, 11, "ベータ", 0),
        (3, 20, 10, "アルファ", 1),
        (4, 20, 12, "ガンマ", 1),
    ])

    result = _output.build_group_appearance_count(df_live, df_stage)

    assert result.index.name == "グループID"
    assert list(result.columns) == [
        "グループ名", "ライブ出演回数", "特典会出演回数", "合計",
    ]
    assert result.loc[10, "グループ名"] == "アルファ"
    assert result.loc[10, "ライブ出演回数"] == 2
    assert result.loc[10, "特典会出演回数"] == 1
    assert result.loc[10, "合計"] == 3
    assert result.loc[11, "ライブ出演回数"] == 1
    assert result.loc[11, "特典会出演回数"] == 0
    assert result.loc[11, "合計"] == 1
    assert result.loc[12, "ライブ出演回数"] == 0
    assert result.loc[12, "特典会出演回数"] == 1
    assert result.loc[12, "合計"] == 1
    assert list(result.index) == [10, 11, 12]


def test_build_group_appearance_count_empty_returns_empty():
    df_stage = _make_stage_df([(0, "メイン", False)])
    df_live = _make_live_df([])

    result = _output.build_group_appearance_count(df_live, df_stage)

    assert list(result.columns) == [
        "グループ名", "ライブ出演回数", "特典会出演回数", "合計",
    ]
    assert len(result) == 0
    assert result.index.name == "グループID"


def test_build_event_output_includes_duration_and_group_count(project_pre_confirm):
    pj_path, project_info_json = project_pre_confirm

    result = _output.build_event_output(pj_path, "event_1", 0, project_info_json)

    assert "duration_distribution" in result
    assert "group_count" in result
    dist = result["duration_distribution"]
    grp = result["group_count"]
    assert list(dist.columns) == ["ライブステージ", "特典会ステージ"]
    assert list(grp.columns) == [
        "グループ名", "ライブ出演回数", "特典会出演回数", "合計",
    ]


# ---------------------------------------------------------------------------
# build_overlap_alerts / build_group_appearances
# ---------------------------------------------------------------------------

def _make_live_df_full(rows):
    """rows = [(出番ID, グループID, グループ名, ステージID, ステージ名,
              ライブ_from, ライブ_to, 長さ(分), 備考), ...]"""
    df = pd.DataFrame(
        {
            "ライブ_from": [r[5] for r in rows],
            "ライブ_to": [r[6] for r in rows],
            "ライブ_長さ(分)": [r[7] for r in rows],
            "グループID": [r[1] for r in rows],
            "ステージID": [r[3] for r in rows],
            "グループ名_raw": [r[2] for r in rows],
            "グループ名": [r[2] for r in rows],
            "ステージ名": [r[4] for r in rows],
            "備考": [r[8] for r in rows],
        },
        index=pd.Index([r[0] for r in rows], name="出番ID"),
    )
    return df


def test_build_overlap_alerts_detects_overlapping_pair():
    df_stage = _make_stage_df([
        (0, "メイン", False),
        (1, "サブ", False),
    ])
    df_live = _make_live_df_full([
        (0, 10, "アルファ", 0, "メイン", "11:00", "11:30", 30, ""),
        (1, 10, "アルファ", 1, "サブ", "11:15", "11:45", 30, ""),
        (2, 11, "ベータ", 0, "メイン", "12:00", "12:30", 30, ""),
    ])

    result = _output.build_overlap_alerts(df_live, df_stage)

    assert len(result) == 1
    row = result.iloc[0]
    assert row["グループID"] == 10
    assert row["グループ名"] == "アルファ"
    assert {row["出番ID_1"], row["出番ID_2"]} == {0, 1}
    assert row["重複(分)"] == 15


def test_build_overlap_alerts_no_overlap_returns_empty():
    df_stage = _make_stage_df([(0, "メイン", False)])
    df_live = _make_live_df_full([
        (0, 10, "アルファ", 0, "メイン", "11:00", "11:30", 30, ""),
        (1, 10, "アルファ", 0, "メイン", "12:00", "12:30", 30, ""),
    ])

    result = _output.build_overlap_alerts(df_live, df_stage)

    assert len(result) == 0
    assert list(result.columns) == [
        "グループID", "グループ名",
        "出番ID_1", "ステージID_1", "ステージ名_1", "開始_1", "終了_1",
        "出番ID_2", "ステージID_2", "ステージ名_2", "開始_2", "終了_2",
        "重複(分)",
    ]


def test_build_overlap_alerts_touching_intervals_not_overlap():
    """11:00-11:30 と 11:30-12:00 は重複扱いしない (境界一致は OK)。"""
    df_stage = _make_stage_df([(0, "メイン", False)])
    df_live = _make_live_df_full([
        (0, 10, "アルファ", 0, "メイン", "11:00", "11:30", 30, ""),
        (1, 10, "アルファ", 0, "メイン", "11:30", "12:00", 30, ""),
    ])

    result = _output.build_overlap_alerts(df_live, df_stage)

    assert len(result) == 0


def test_build_overlap_alerts_cross_stage_kind():
    """ライブステージと特典会ステージを跨いだ重複も検出する。"""
    df_stage = _make_stage_df([
        (0, "メイン", False),       # ライブ
        (1, "ブース", True),        # 特典会
    ])
    df_live = _make_live_df_full([
        (0, 10, "アルファ", 0, "メイン", "11:00", "11:30", 30, ""),
        (1, 10, "アルファ", 1, "ブース", "11:15", "11:45", 30, ""),
    ])

    result = _output.build_overlap_alerts(df_live, df_stage)

    assert len(result) == 1
    assert result.iloc[0]["重複(分)"] == 15


def test_build_overlap_alerts_three_way_overlap_yields_three_pairs():
    df_stage = _make_stage_df([(0, "メイン", False)])
    df_live = _make_live_df_full([
        (0, 10, "アルファ", 0, "メイン", "11:00", "12:00", 60, ""),
        (1, 10, "アルファ", 0, "メイン", "11:15", "12:15", 60, ""),
        (2, 10, "アルファ", 0, "メイン", "11:30", "12:30", 60, ""),
    ])

    result = _output.build_overlap_alerts(df_live, df_stage)

    # 3つすべて互いに重なる → C(3,2) = 3 ペア
    assert len(result) == 3


def test_build_overlap_alerts_empty_returns_empty():
    df_stage = _make_stage_df([(0, "メイン", False)])
    df_live = _make_live_df_full([])

    result = _output.build_overlap_alerts(df_live, df_stage)
    assert len(result) == 0


def test_build_group_appearances_returns_all_groups():
    df_stage = _make_stage_df([
        (0, "メイン", False),
        (1, "ブース", True),
    ])
    df_live = _make_live_df_full([
        (0, 10, "アルファ", 0, "メイン", "11:00", "11:30", 30, ""),
        (1, 10, "アルファ", 1, "ブース", "12:00", "12:30", 30, "特典会"),
        (2, 11, "ベータ", 0, "メイン", "13:00", "13:25", 25, ""),
    ])

    result = _output.build_group_appearances(df_live, df_stage)

    assert result.index.name == "出番ID"
    assert list(result.columns) == [
        "グループID", "グループ名", "ステージID", "ステージ名",
        "ライブ_from", "ライブ_to", "ライブ_長さ(分)", "備考",
    ]
    # 全3件保持される
    assert len(result) == 3
    # グループID=10 で絞り込めば 2 件
    alpha = result[result["グループID"] == 10]
    assert len(alpha) == 2
    assert list(alpha["ステージ名"]) == ["メイン", "ブース"]


def test_build_group_appearances_empty_returns_empty():
    df_stage = _make_stage_df([(0, "メイン", False)])
    df_live = _make_live_df_full([])

    result = _output.build_group_appearances(df_live, df_stage)
    assert len(result) == 0
    assert list(result.columns) == [
        "グループID", "グループ名", "ステージID", "ステージ名",
        "ライブ_from", "ライブ_to", "ライブ_長さ(分)", "備考",
    ]


def test_build_event_output_includes_overlap_and_appearances(project_pre_confirm):
    pj_path, project_info_json = project_pre_confirm

    result = _output.build_event_output(pj_path, "event_1", 0, project_info_json)

    assert "overlap_alerts" in result
    assert "group_appearances" in result
    assert list(result["group_appearances"].columns) == [
        "グループID", "グループ名", "ステージID", "ステージ名",
        "ライブ_from", "ライブ_to", "ライブ_長さ(分)", "備考",
    ]


def test_listup_new_idolname_skips_empty_event(tmp_path, monkeypatch):
    project_info_json, output_df = _make_two_event_project(tmp_path)
    event_list = ["event_1", "event_2"]

    # idolname.detect_new_data の依存をスタブ化
    from backend_functions import idolname as _idolname_mod
    monkeypatch.setattr(_idolname_mod, "detect_new_data", lambda names: list(names))

    df = _output.listup_new_idolname(output_df, event_list)

    # event_1 の "ZZグループ" だけ抽出される(event_2は空dictで無視)
    assert "ZZグループ" in df["グループ名"].tolist()


# ---------------------------------------------------------------------------
# detect_id_anomalies / detect_master_diff
# ---------------------------------------------------------------------------

def _make_live(rows):
    """rows: list of (turn_id, stage_id, live_from, group_id, group_name)."""
    df = pd.DataFrame(
        [
            {
                "グループID": gid,
                "ステージID": sid,
                "ライブ_from": lfrom,
                "グループ名": gname,
                "ステージ名": f"stage{sid}",
            }
            for (_t, sid, lfrom, gid, gname) in rows
        ],
        index=[t for (t, _s, _f, _g, _n) in rows],
    )
    df.index.name = "出番ID"
    return df


def _make_idolname(pairs):
    """pairs: list of (group_id, group_name)."""
    df = pd.DataFrame(
        {"グループ名_採用": [n for (_i, n) in pairs]},
        index=[i for (i, _n) in pairs],
    )
    df.index.name = "グループID"
    return df


def _make_stage(pairs):
    """pairs: list of (stage_id, stage_name)."""
    df = pd.DataFrame(
        {"ステージ名": [n for (_i, n) in pairs]},
        index=[i for (i, _n) in pairs],
    )
    df.index.name = "ステージID"
    return df


def test_detect_id_anomalies_clean_with_collab():
    # 出番ID=0 はコラボ (同一ステージ・同一from・別グループ2行) → 正当
    live = _make_live([
        (0, 0, "12:00", 10, "A"),
        (0, 0, "12:00", 11, "B"),
        (1, 0, "13:00", 12, "C"),
    ])
    idolname = _make_idolname([(10, "A"), (11, "B"), (12, "C")])
    stage = _make_stage([(0, "メイン")])

    result = _output.detect_id_anomalies(
        {"live": live, "idolname": idolname, "stage": stage},
    )
    assert result == []


def test_detect_id_anomalies_turn_id_across_timeslots():
    # 同一出番IDが別 ライブ_from に重複
    live = _make_live([
        (5, 0, "12:00", 10, "A"),
        (5, 0, "13:00", 11, "B"),
    ])
    result = _output.detect_id_anomalies({"live": live})
    assert any("出番ID=5" in m for m in result)


def test_detect_id_anomalies_turn_id_across_stages():
    # 同一出番IDが別ステージに重複
    live = _make_live([
        (7, 0, "12:00", 10, "A"),
        (7, 1, "12:00", 10, "A"),
    ])
    result = _output.detect_id_anomalies({"live": live})
    assert any("出番ID=7" in m for m in result)


def test_detect_id_anomalies_group_id_multiple_names():
    idolname = _make_idolname([(3, "A"), (3, "B")])
    result = _output.detect_id_anomalies({"idolname": idolname})
    assert any("グループID=3" in m for m in result)


def test_detect_id_anomalies_stage_id_multiple_names():
    stage = _make_stage([(2, "メイン"), (2, "サブ")])
    result = _output.detect_id_anomalies({"stage": stage})
    assert any("ステージID=2" in m for m in result)


def test_detect_master_diff_no_change(tmp_path):
    live = _make_live([
        (0, 0, "12:00", 10, "A"),
        (1, 0, "13:00", 11, "B"),
    ])
    live.to_csv(os.path.join(tmp_path, "turn_id_data.csv"))
    result = _output.detect_master_diff({"live": live}, str(tmp_path))
    assert result == []


def test_detect_master_diff_existing_turn_changed(tmp_path):
    old = _make_live([
        (0, 0, "12:00", 10, "A"),
        (1, 0, "13:00", 11, "B"),
    ])
    old.to_csv(os.path.join(tmp_path, "turn_id_data.csv"))
    # 出番ID=1 のグループ/時刻を変更
    new = _make_live([
        (0, 0, "12:00", 10, "A"),
        (1, 0, "14:00", 99, "Z"),
    ])
    result = _output.detect_master_diff({"live": new}, str(tmp_path))
    assert any("出番ID=1" in m for m in result)
    assert not any("出番ID=0" in m for m in result)


def test_detect_master_diff_new_turn_only(tmp_path):
    old = _make_live([(0, 0, "12:00", 10, "A")])
    old.to_csv(os.path.join(tmp_path, "turn_id_data.csv"))
    # 出番ID=1 を新規追加 (旧 master に無い) → 通知しない
    new = _make_live([
        (0, 0, "12:00", 10, "A"),
        (1, 0, "13:00", 11, "B"),
    ])
    result = _output.detect_master_diff({"live": new}, str(tmp_path))
    assert result == []


def test_detect_master_diff_missing_csv_returns_empty(tmp_path):
    live = _make_live([(0, 0, "12:00", 10, "A")])
    result = _output.detect_master_diff({"live": live}, str(tmp_path))
    assert result == []
