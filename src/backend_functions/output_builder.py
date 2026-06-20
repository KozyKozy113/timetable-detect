"""
出力データの構築・エクスポート。

ステージマスタ・アーティストマスタ・出番データの組み立て、
ID確定、Excelエクスポート、グループ名マスタ更新を行う。
Streamlitに依存しない。
"""

from __future__ import annotations

import json
import os
from datetime import datetime
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook

from backend_functions import timetabledata, idolname, s3access
from backend_functions import project_repository as repo
from backend_functions import stage_color


# ---------------------------------------------------------------------------
# カラープリセット (Stella stageList.colorName)
# ---------------------------------------------------------------------------

def _default_color_name(stage_order: int, is_tokutenkai: bool, kind_hint: str | None = None) -> str:
    """ステージ種別に応じてデフォルト カラー名 を返す (Phase 2-2)。

    - 縁日ステージ (kind_hint="ennichi"): stage-blueGrey
    - 特典会ステージ: stage-grey
    - ライブステージ: プリセット色 (color_preset.json) を `stage_order` で循環
    """
    if kind_hint == "ennichi":
        return "stage-blueGrey"
    if is_tokutenkai:
        return "stage-grey"
    names = stage_color.get_preset_color_names()
    if not names:
        return "stage-grey"
    return names[int(stage_order) % len(names)]


# ---------------------------------------------------------------------------
# ステージマスタ・出力データ組み立て
# ---------------------------------------------------------------------------

def _stage_master_get(stage_master: dict, stage_id: int) -> dict | None:
    """stage_master の key 型ゆれ (int / str) を吸収して引き当てる。"""
    if stage_id in stage_master:
        return stage_master[stage_id]
    if str(stage_id) in stage_master:
        return stage_master[str(stage_id)]
    return None


def _stage_master_keys_as_int(stage_master: dict) -> list[int]:
    """stage_master の key を int に正規化して返す。"""
    return [int(k) for k in stage_master.keys()]


def find_or_create_stage_id(
    stage_master: dict,
    stage_name: str,
    is_tokutenkai: bool,
    next_id: int,
    existing_stage_id: int | None = None,
    color_name: str | None = None,
) -> tuple[int, int]:
    """既存マスタからステージIDを探し、なければnext_idで新規登録する。

    引き当て優先順位:
        1. existing_stage_id 指定があり、マスタにそのIDがあれば使用。
           ステージ名が異なれば**マスタ側を更新**(編集モードの反映)。
        2. existing_stage_id 指定がない/マスタ未登録のとき、ステージ名一致で検索。
        3. いずれにも当たらなければ next_id で新規登録。

    新規登録時は引数 stage_master が副作用で更新される(参照渡し)。
    新規登録時に既存ステージマスタの `表示順` 最大値 + 1 を表示順として採番する。
    `非活性化フラグ` は False で初期化する。
    新規登録時、color_name (LLM推定カラー) があればそれを `カラー名` に採用し、
    無ければ従来どおりデフォルトカラーを割り当てる。既存ステージのカラーは
    上書きしない(⑥での手動編集を尊重)。

    Returns:
        (assigned_id, updated_next_id):
            既存ステージ → (既存ID, next_id)
            新規ステージ → (next_id, next_id + 1)
    """
    if existing_stage_id is not None:
        existing_entry = _stage_master_get(stage_master, existing_stage_id)
        if existing_entry is not None:
            if existing_entry.get("ステージ名") != stage_name:
                existing_entry["ステージ名"] = stage_name
            return int(existing_stage_id), next_id

    for k, v in stage_master.items():
        if v["ステージ名"] == stage_name:
            return int(k), next_id

    # 新規登録: 表示順 = 既存最大値 + 1
    existing_orders = [
        v.get("表示順")
        for v in stage_master.values()
        if v.get("表示順") is not None
    ]
    next_order = max(existing_orders) + 1 if existing_orders else 0
    resolved_color = (
        color_name if (isinstance(color_name, str) and color_name)
        else _default_color_name(next_order, is_tokutenkai)
    )
    stage_master[next_id] = {
        "ステージ名": stage_name,
        "特典会フラグ": is_tokutenkai,
        "表示順": next_order,
        "非活性化フラグ": False,
        # Phase 2: Stella stageList 拡張フィールド
        "ステージ名_短縮": stage_name,
        "カラー名": resolved_color,
    }
    return next_id, next_id + 1


def load_existing_masters(
    output_path: str,
) -> tuple[dict, int, pd.DataFrame, int]:
    """master_stage.csv / master_idolname.csv が存在すれば読み込み、なければ空の初期値を返す。

    「IDマスタ確定」前後の挙動差をここで吸収する。

    Returns:
        (stage_master, next_stage_id, idolname_master_df, next_artist_id)
    """
    stage_csv = os.path.join(output_path, "master_stage.csv")
    if os.path.exists(stage_csv):
        stage_master_df = pd.read_csv(stage_csv, index_col=0)
        # 後方互換: 表示順 / 非活性化フラグ カラムが無ければ補完
        if "表示順" not in stage_master_df.columns:
            stage_master_df["表示順"] = range(len(stage_master_df))
        if "非活性化フラグ" not in stage_master_df.columns:
            stage_master_df["非活性化フラグ"] = False
        stage_master_df["非活性化フラグ"] = (
            stage_master_df["非活性化フラグ"].fillna(False).astype(bool)
        )
        # Phase 2: ステージ名_短縮 / カラー名 の補完
        if "ステージ名_短縮" not in stage_master_df.columns:
            stage_master_df["ステージ名_短縮"] = stage_master_df["ステージ名"]
        stage_master_df["ステージ名_短縮"] = (
            stage_master_df["ステージ名_短縮"]
            .fillna(stage_master_df["ステージ名"]).astype(str)
        )
        if "カラー名" not in stage_master_df.columns:
            stage_master_df["カラー名"] = None
        # 欠損行を デフォルト割当ルール で埋める
        for sid in stage_master_df.index:
            cur = stage_master_df.at[sid, "カラー名"]
            if cur is None or (isinstance(cur, float) and pd.isna(cur)) or cur == "":
                order = int(stage_master_df.at[sid, "表示順"])
                is_tk = bool(stage_master_df.at[sid, "特典会フラグ"])
                stage_master_df.at[sid, "カラー名"] = _default_color_name(order, is_tk)
        stage_master = json.loads(stage_master_df.T.to_json())
        next_stage_id = int(max(stage_master_df.index)) + 1
    else:
        stage_master = {}
        next_stage_id = 0

    idolname_csv = os.path.join(output_path, "master_idolname.csv")
    if os.path.exists(idolname_csv):
        idolname_master_df = pd.read_csv(idolname_csv, index_col=0).rename(
            columns={"グループ名": "グループ名_採用"},
        )
        next_artist_id = int(max(idolname_master_df.index)) + 1
    else:
        idolname_master_df = pd.DataFrame(
            columns=["グループID", "グループ名_採用"],
        ).set_index("グループID")
        next_artist_id = 0

    return stage_master, next_stage_id, idolname_master_df, next_artist_id


_LIVE_OUTPUT_COLUMNS = [
    "ライブ_from", "ライブ_to", "ライブ_長さ(分)", "グループID", "ステージID",
    "グループ名_raw", "グループ名", "ステージ名", "備考",
    "コラボタイトル",
]
# Excel 出力時に省略する列 (UI 表示と Stella用算出のために output_df には含めるが、
# 既存 Excel 出力フォーマットを維持するため書き出さない)
_LIVE_EXCEL_DROP_COLUMNS = ["ライブ_to"]

# Excel の stage シートに出力する列と順序 (Stella JSON の stageList と同等の情報)。
# 非活性化ステージは Excel から除外されるため `非活性化フラグ` 列は出さない。
_STAGE_EXCEL_COLUMNS = ["ステージ名", "ステージ名_短縮", "カラー名", "表示順", "特典会フラグ"]

# コラボ出番 (同一 出番ID の複数行) を 1 行に統合する際、カンマ区切りで連結する列。
_LIVE_COLLAB_JOIN_COLUMNS = ["グループID", "グループ名", "グループ名_raw"]

_DURATION_COL_LIVE = "ライブステージ"
_DURATION_COL_TOKUTENKAI = "特典会ステージ"
_DURATION_INDEX_NAME = "長さ(分)"

_GROUP_COUNT_COL_NAME = "グループ名"
_GROUP_COUNT_COL_LIVE = "ライブ出演回数"
_GROUP_COUNT_COL_TOKUTENKAI = "特典会出演回数"
_GROUP_COUNT_COL_TOTAL = "合計"

_OVERLAP_COLUMNS = [
    "グループID", "グループ名",
    "出番ID_1", "ステージID_1", "ステージ名_1", "開始_1", "終了_1",
    "出番ID_2", "ステージID_2", "ステージ名_2", "開始_2", "終了_2",
    "重複(分)",
]
_APPEARANCE_COLUMNS = [
    "グループID", "グループ名", "ステージID", "ステージ名",
    "ライブ_from", "ライブ_to", "ライブ_長さ(分)", "備考",
]


def _parse_hhmm(value) -> "datetime | None":
    """'HH:MM' 文字列を datetime に変換。失敗時 None。"""
    if value is None or (isinstance(value, float) and pd.isna(value)) or value == "":
        return None
    try:
        return datetime.strptime(str(value), "%H:%M")
    except (ValueError, TypeError):
        return None


def build_duration_distribution(
    df_live: pd.DataFrame,
    df_stage: pd.DataFrame,
) -> pd.DataFrame:
    """出演枠の長さ(分)ごとの出現回数を、ライブ／特典会ステージ別にカウントする。

    Returns:
        index : 長さ(分) 昇順
        cols  : [ライブステージ, 特典会ステージ]  (int, 0埋め)
    """
    empty = pd.DataFrame(
        {_DURATION_COL_LIVE: pd.Series(dtype="int64"),
         _DURATION_COL_TOKUTENKAI: pd.Series(dtype="int64")},
    )
    empty.index.name = _DURATION_INDEX_NAME

    if df_live is None or len(df_live) == 0:
        return empty

    flag_map = df_stage["特典会フラグ"].to_dict()
    df = df_live[["ステージID", "ライブ_長さ(分)"]].copy()
    df = df[df["ライブ_長さ(分)"].notna() & (df["ライブ_長さ(分)"] != "")]
    if len(df) == 0:
        return empty
    df["特典会フラグ"] = df["ステージID"].map(flag_map).fillna(False).astype(bool)

    grouped = (
        df.groupby(["ライブ_長さ(分)", "特典会フラグ"])
        .size()
        .unstack("特典会フラグ", fill_value=0)
    )
    if False not in grouped.columns:
        grouped[False] = 0
    if True not in grouped.columns:
        grouped[True] = 0
    result = grouped.rename(
        columns={False: _DURATION_COL_LIVE, True: _DURATION_COL_TOKUTENKAI},
    )[[_DURATION_COL_LIVE, _DURATION_COL_TOKUTENKAI]]
    result = result.sort_index()
    result.index.name = _DURATION_INDEX_NAME
    return result.astype("int64")


def build_group_appearance_count(
    df_live: pd.DataFrame,
    df_stage: pd.DataFrame,
) -> pd.DataFrame:
    """グループごとの出演回数を、ライブ／特典会ステージ別にカウントする。

    Returns:
        index : グループID 昇順
        cols  : [グループ名, ライブ出演回数, 特典会出演回数, 合計]
    """
    empty = pd.DataFrame(
        {_GROUP_COUNT_COL_NAME: pd.Series(dtype="object"),
         _GROUP_COUNT_COL_LIVE: pd.Series(dtype="int64"),
         _GROUP_COUNT_COL_TOKUTENKAI: pd.Series(dtype="int64"),
         _GROUP_COUNT_COL_TOTAL: pd.Series(dtype="int64")},
    )
    empty.index.name = "グループID"

    if df_live is None or len(df_live) == 0:
        return empty

    flag_map = df_stage["特典会フラグ"].to_dict()
    df = df_live[["グループID", "グループ名", "ステージID"]].copy()
    df = df[df["グループID"].notna()]
    if len(df) == 0:
        return empty
    df["特典会フラグ"] = df["ステージID"].map(flag_map).fillna(False).astype(bool)

    grouped = (
        df.groupby(["グループID", "グループ名", "特典会フラグ"])
        .size()
        .unstack("特典会フラグ", fill_value=0)
    )
    if False not in grouped.columns:
        grouped[False] = 0
    if True not in grouped.columns:
        grouped[True] = 0
    grouped = grouped.rename(
        columns={False: _GROUP_COUNT_COL_LIVE, True: _GROUP_COUNT_COL_TOKUTENKAI},
    )
    grouped[_GROUP_COUNT_COL_TOTAL] = (
        grouped[_GROUP_COUNT_COL_LIVE] + grouped[_GROUP_COUNT_COL_TOKUTENKAI]
    )

    result = grouped.reset_index(level="グループ名")
    result = result[[
        _GROUP_COUNT_COL_NAME,
        _GROUP_COUNT_COL_LIVE,
        _GROUP_COUNT_COL_TOKUTENKAI,
        _GROUP_COUNT_COL_TOTAL,
    ]].sort_index()
    result.index = result.index.astype("int64")
    result.index.name = "グループID"
    for col in (_GROUP_COUNT_COL_LIVE, _GROUP_COUNT_COL_TOKUTENKAI, _GROUP_COUNT_COL_TOTAL):
        result[col] = result[col].astype("int64")
    return result


def build_overlap_alerts(
    df_live: pd.DataFrame,
    df_stage: pd.DataFrame,
) -> pd.DataFrame:
    """同一グループ内で出演時間が重なっているペアを検出する。

    ライブ／特典会ステージを跨いで検出する。

    Returns:
        cols : [グループID, グループ名,
                出番ID_1, ステージID_1, ステージ名_1, 開始_1, 終了_1,
                出番ID_2, ステージID_2, ステージ名_2, 開始_2, 終了_2,
                重複(分)]
        重複が無い場合は空 DataFrame を返す。
    """
    empty = pd.DataFrame(columns=_OVERLAP_COLUMNS)

    if df_live is None or len(df_live) == 0:
        return empty
    required = {"グループID", "グループ名", "ステージID", "ライブ_from", "ライブ_to"}
    if not required.issubset(df_live.columns):
        return empty

    df = df_live.copy()
    df["_from_dt"] = df["ライブ_from"].apply(_parse_hhmm)
    df["_to_dt"] = df["ライブ_to"].apply(_parse_hhmm)
    df = df[df["_from_dt"].notna() & df["_to_dt"].notna()]
    df = df[df["グループID"].notna()]
    if len(df) == 0:
        return empty

    rows = []
    for group_id, group_df in df.groupby("グループID"):
        appearances = group_df.sort_values("_from_dt").reset_index()
        n = len(appearances)
        if n < 2:
            continue
        group_name = appearances.iloc[0]["グループ名"]
        for i in range(n):
            for j in range(i + 1, n):
                a = appearances.iloc[i]
                b = appearances.iloc[j]
                overlap_start = max(a["_from_dt"], b["_from_dt"])
                overlap_end = min(a["_to_dt"], b["_to_dt"])
                if overlap_start < overlap_end:
                    overlap_min = int(
                        (overlap_end - overlap_start).total_seconds() // 60
                    )
                    rows.append({
                        "グループID": int(group_id),
                        "グループ名": group_name,
                        "出番ID_1": a["出番ID"] if "出番ID" in a.index else a["index"],
                        "ステージID_1": int(a["ステージID"]),
                        "ステージ名_1": a.get("ステージ名", ""),
                        "開始_1": a["ライブ_from"],
                        "終了_1": a["ライブ_to"],
                        "出番ID_2": b["出番ID"] if "出番ID" in b.index else b["index"],
                        "ステージID_2": int(b["ステージID"]),
                        "ステージ名_2": b.get("ステージ名", ""),
                        "開始_2": b["ライブ_from"],
                        "終了_2": b["ライブ_to"],
                        "重複(分)": overlap_min,
                    })

    if not rows:
        return empty
    return pd.DataFrame(rows, columns=_OVERLAP_COLUMNS).reset_index(drop=True)


def build_group_appearances(
    df_live: pd.DataFrame,
    df_stage: pd.DataFrame,
) -> pd.DataFrame:
    """全グループの出番一覧を返す。UI側でグループIDで絞り込んで使う。

    Returns:
        index : 出番ID
        cols  : [グループID, グループ名, ステージID, ステージ名,
                 ライブ_from, ライブ_to, ライブ_長さ(分), 備考]
    """
    empty = pd.DataFrame(columns=_APPEARANCE_COLUMNS)
    empty.index.name = "出番ID"

    if df_live is None or len(df_live) == 0:
        return empty
    required = {"グループID", "グループ名", "ステージID", "ステージ名",
                "ライブ_from", "ライブ_to", "ライブ_長さ(分)", "備考"}
    if not required.issubset(df_live.columns):
        return empty

    df = df_live[df_live["グループID"].notna()].copy()
    if len(df) == 0:
        return empty

    if "出番ID" in df.columns:
        df = df.set_index("出番ID")
    df = df[_APPEARANCE_COLUMNS]
    df.index.name = "出番ID"
    return df.sort_values(["グループID", "ライブ_from"])


def build_event_output(
    pj_path: str,
    event_name: str,
    event_no: int,
    project_info_json: dict,
    only_event_types: list[str] | None = None,
) -> dict[str, pd.DataFrame] | None:
    """1イベント分の出力データ(stage / idolname / live)を組み立てて返す。

    データが1件も存在しない場合は None を返す。

    `only_event_types` を指定すると、その種別(event_type)のみを処理対象とする。
    採番済み種別だけを再ビルドし、未採番種別を masters に含めないために使う
    (None の場合は全種別＝従来動作)。

    処理フロー:
        1. load_existing_masters() で既存マスタを読み込み(IDマスタ確定前後の差を吸収)
        2. event_type × stage を走査してステージマスタとライブデータを構築
        3. 特典会併記タイテの場合、特典会データをブース別にステージマスタへ統合
        4. アーティストマスタを構築
        5. 出番データを構築(既存出番IDの維持と新規ID採番)

    Returns:
        {"stage": df_stage, "idolname": df_idolname, "live": df_live} or None
    """
    output_path = os.path.join(pj_path, event_name)
    stage_master, stage_id, idolname_master_df, artist_id = load_existing_masters(
        output_path,
    )

    event_type_list = repo.get_event_type_list(project_info_json, event_no)
    if only_event_types is not None:
        _only = set(only_event_types)
        event_type_list = [et for et in event_type_list if et in _only]
    tokutenkai_timetable = []
    event_timetable_all = []
    # 特典会併記形式のブース名 -> 既存子ステージID (引き当てヒント)
    booth_existing_id_map_all: dict[str, int] = {}
    # 併記 (live_tokutenkai_heiki) 種別名の収集 (ブースID採番順の判定に使う)
    heiki_event_types: list[str] = []

    for event_type in event_type_list:
        tgt_event_type_info = repo.get_image_entry_by_dir_name(
            project_info_json, event_no, event_type,
        )
        stage_name_list = repo.get_stage_name_list(project_info_json, event_no, event_type)
        color_name_list = repo.get_stage_color_list(project_info_json, event_no, event_type)
        kind = tgt_event_type_info["kind"]
        tokutenkai_flg = kind == "tokutenkai"
        is_heiki = kind == "live_tokutenkai_heiki"
        if is_heiki:
            heiki_event_types.append(event_type)

        for stage_no in range(tgt_event_type_info["stage_num"]):
            json_path = os.path.join(output_path, event_type, f"stage_{stage_no}.json")
            if not os.path.exists(json_path):
                continue

            try:
                with open(json_path, encoding="utf-8") as f:
                    edit_tgt_json = json.load(f)

                # トップレベル ステージID (親=ライブステージID) を取得。
                # 未確定時は None。
                existing_top_stage_id = edit_tgt_json.get("ステージID")
                try:
                    existing_top_stage_id = (
                        int(existing_top_stage_id)
                        if existing_top_stage_id is not None else None
                    )
                except (ValueError, TypeError):
                    existing_top_stage_id = None

                # 特典会併記形式の場合、特典会[].ステージID (子=ブース別ID) を
                # ブース名 -> 子ID の辞書として収集
                booth_existing_id_map: dict[str, int] = {}
                if is_heiki:
                    for turn in edit_tgt_json.get("タイムテーブル", []):
                        for tk in turn.get("特典会", []) or []:
                            booth_name = tk.get("ブース")
                            tk_stage_id = tk.get("ステージID")
                            if booth_name and tk_stage_id is not None:
                                try:
                                    booth_existing_id_map[booth_name] = int(tk_stage_id)
                                except (ValueError, TypeError):
                                    pass

                if is_heiki:
                    df_edit_tgt = timetabledata.json_to_df(edit_tgt_json, tokutenkai=True)
                    df_edit_live, df_edit_tokutenkai = \
                        timetabledata.devide_df_live_tokutenkai(df_edit_tgt)
                    df_edit_tokutenkai = df_edit_tokutenkai[
                        df_edit_tokutenkai["ライブ_長さ(分)"].notnull()
                        & (df_edit_tokutenkai["ライブ_長さ(分)"] != "")
                    ]
                    tokutenkai_timetable.append(df_edit_tokutenkai)
                else:
                    df_edit_live = timetabledata.json_to_df(edit_tgt_json, tokutenkai=False)

                df_edit_live = df_edit_live[
                    df_edit_live["ライブ_長さ(分)"].notnull()
                    & (df_edit_live["ライブ_長さ(分)"] != "")
                ].copy()

                this_stage_id, stage_id = find_or_create_stage_id(
                    stage_master, stage_name_list[stage_no], tokutenkai_flg, stage_id,
                    existing_stage_id=existing_top_stage_id,
                    color_name=color_name_list[stage_no] if stage_no < len(color_name_list) else None,
                )
                if "ステージID" not in df_edit_live.columns:
                    df_edit_live["ステージID"] = None
                    df_edit_live["ステージ名"] = None
                df_edit_live.loc[:, "ステージID"] = this_stage_id
                df_edit_live.loc[:, "ステージ名"] = stage_name_list[stage_no]
                event_timetable_all.append(df_edit_live)
                # 子ブースIDの引き当てヒントを後続処理に引き渡す
                booth_existing_id_map_all.update(booth_existing_id_map)
            except KeyError:
                pass

    # 併記種別が「種別として未採番」なら、ブースIDをブース名順に採番する。
    # 採番済み (いずれかの併記種別の stage_id が確定済み) の場合は既存IDを
    # 乱さないよう従来どおり出現順を維持する。
    heiki_unnumbered = bool(heiki_event_types) and not any(
        repo.img_type_ids_assigned(project_info_json, event_no, et)
        for et in heiki_event_types
    )

    stage_master_tokutenkai = {}
    df_tokutenkai = None
    if len(tokutenkai_timetable) > 0:
        df_tokutenkai = pd.concat(tokutenkai_timetable).reset_index(drop=True)
        df_tokutenkai = df_tokutenkai.drop(columns=["ステージID"], errors="ignore")
        booth_name_list = df_tokutenkai["ステージ名"].drop_duplicates().tolist()
        if heiki_unnumbered:
            # ブース名順に採番 → ID・表示順がともにブース名順に揃う
            booth_name_list = sorted(booth_name_list)
        for booth_name in booth_name_list:
            existing_booth_id = booth_existing_id_map_all.get(booth_name)
            this_stage_id, stage_id = find_or_create_stage_id(
                stage_master, booth_name, True, stage_id,
                existing_stage_id=existing_booth_id,
            )
            stage_master_tokutenkai[this_stage_id] = {
                "ステージ名": booth_name, "特典会フラグ": True,
            }

    if len(event_timetable_all) == 0:
        return None

    df_stage = pd.DataFrame.from_dict(stage_master, orient="index")
    df_stage.index.name = "ステージID"
    df_stage.index = df_stage.index.astype(int)
    # 後方互換: 表示順 / 非活性化フラグ が無い (新規作成パス) 場合は補完
    if "表示順" not in df_stage.columns:
        df_stage["表示順"] = range(len(df_stage))
    else:
        # 欠損行 (find_or_create_stage_id が古い経路で作られた等) を index 順で埋める
        missing_mask = df_stage["表示順"].isna()
        if missing_mask.any():
            df_stage.loc[missing_mask, "表示順"] = list(range(len(df_stage)))[: int(missing_mask.sum())]
    if "非活性化フラグ" not in df_stage.columns:
        df_stage["非活性化フラグ"] = False
    df_stage["非活性化フラグ"] = df_stage["非活性化フラグ"].fillna(False).astype(bool)
    df_stage["表示順"] = df_stage["表示順"].astype(int)
    # Phase 2: ステージ名_短縮 / カラー名 補完
    if "ステージ名_短縮" not in df_stage.columns:
        df_stage["ステージ名_短縮"] = df_stage["ステージ名"]
    df_stage["ステージ名_短縮"] = (
        df_stage["ステージ名_短縮"].fillna(df_stage["ステージ名"]).astype(str)
    )
    if "カラー名" not in df_stage.columns:
        df_stage["カラー名"] = None
    for sid in df_stage.index:
        cur = df_stage.at[sid, "カラー名"]
        if cur is None or (isinstance(cur, float) and pd.isna(cur)) or cur == "":
            order = int(df_stage.at[sid, "表示順"])
            is_tk = bool(df_stage.at[sid, "特典会フラグ"])
            df_stage.at[sid, "カラー名"] = _default_color_name(order, is_tk)
    df_stage = df_stage.sort_values("表示順")

    df_live = pd.concat(event_timetable_all).reset_index(drop=True)
    if df_tokutenkai is not None:
        df_stage_tokutenkai = pd.DataFrame.from_dict(stage_master_tokutenkai, orient="index")
        df_stage_tokutenkai.index.name = "ステージID"
        df_tokutenkai = pd.merge(
            df_tokutenkai,
            df_stage_tokutenkai.reset_index().drop("特典会フラグ", axis=1),
            on="ステージ名", how="left",
        )
        df_live = pd.concat([df_live, df_tokutenkai]).reset_index(drop=True)

    # アーティストマスタ構築
    df_idolname = pd.DataFrame(
        df_live["グループ名_採用"].drop_duplicates().sort_values().reset_index(drop=True)
    )
    df_idolname = df_idolname[
        ~df_idolname["グループ名_採用"].isin(idolname_master_df["グループ名_採用"])
    ].reset_index(drop=True)
    df_idolname.index = df_idolname.index + artist_id
    df_idolname.index.name = "グループID"
    df_idolname = pd.concat([idolname_master_df, df_idolname])

    # 出番データ構築
    df_live = df_live.drop(columns=["グループID"], errors="ignore")
    df_live = pd.merge(
        df_live, df_idolname.reset_index(), on="グループ名_採用", how="left",
    ).rename(columns={"グループ名": "グループ名_raw", "グループ名_採用": "グループ名"})

    # Phase 1: コラボグループ統合
    # 方針: 出番ID 優先・コラボグループID 劣後。
    # - 既に 出番ID が振られた行はそのまま (コラボグループID は完全に無視)
    # - 出番ID が NULL かつ 同一 (ステージID, コラボグループID) を持つ行群には
    #   同一の新規 出番ID を採番
    # - 出番ID が NULL かつ コラボグループID も NULL の行は通常通り個別採番
    # - コラボタイトル は最終的な 出番ID 単位で正規化 (最初の非空値を全行に伝播)
    if "コラボタイトル" not in df_live.columns:
        df_live["コラボタイトル"] = None
    if "コラボグループID" not in df_live.columns:
        df_live["コラボグループID"] = None
    if "出番ID" not in df_live.columns:
        df_live["出番ID"] = None

    turn_id_series = pd.to_numeric(df_live["出番ID"], errors="coerce")
    cgid_series = pd.to_numeric(df_live["コラボグループID"], errors="coerce")

    existing_ids = turn_id_series.dropna()
    next_turn_id = int(existing_ids.max()) + 1 if len(existing_ids) > 0 else 0

    # 出番ID 未採番 かつ コラボグループID あり の行群を (ステージID, cgid) でまとめて採番
    unassigned_cgid_mask = turn_id_series.isna() & cgid_series.notna()
    if unassigned_cgid_mask.any():
        df_live.loc[unassigned_cgid_mask, "コラボグループID"] = (
            cgid_series[unassigned_cgid_mask].astype(int)
        )
        for (_sid, _cgid), group in df_live[unassigned_cgid_mask].groupby(
            ["ステージID", "コラボグループID"], dropna=False,
        ):
            df_live.loc[group.index, "出番ID"] = next_turn_id
            next_turn_id += 1

    # 残り (出番ID NULL かつ コラボグループID も NULL) は 1 行ずつ個別採番
    missing_mask = df_live["出番ID"].isna()
    new_ids = list(range(next_turn_id, next_turn_id + int(missing_mask.sum())))
    df_live.loc[missing_mask, "出番ID"] = new_ids
    df_live["出番ID"] = df_live["出番ID"].astype(int)

    # コラボタイトル正規化: 確定後の 出番ID 単位で最初の非空値を伝播
    # (cgid ベースではなく 出番ID ベース。出番ID が同じなら cgid が違っても同一タイトル)
    for _turn_id, group in df_live.groupby("出番ID"):
        if len(group) < 2:
            continue
        non_empty = [
            v for v in group["コラボタイトル"].tolist()
            if isinstance(v, str) and v != ""
        ]
        if non_empty:
            df_live.loc[group.index, "コラボタイトル"] = non_empty[0]

    df_live = df_live.set_index("出番ID")

    df_live_out = df_live[_LIVE_OUTPUT_COLUMNS]

    # `stage` / `live` は master 実体として「非活性化を含む全行」を保持する。
    # こうすることで determine_id_master / save_event_edits が CSV を上書きしても
    # 非活性化ステージ行が消失しない。
    # 集計 (duration_distribution / group_count / overlap_alerts / group_appearances)
    # は非活性化ステージとそれに紐づく出番を除外して算出する。
    # 表示・出力 (UI / Excel / Stella JSON) 側で必要に応じて `非活性化フラグ` で
    # フィルタすること。
    disabled_stage_ids = set(df_stage[df_stage["非活性化フラグ"]].index.tolist())
    if disabled_stage_ids:
        df_stage_visible = df_stage[~df_stage["非活性化フラグ"]]
        df_live_visible = df_live[~df_live["ステージID"].isin(disabled_stage_ids)]
        df_live_out_visible = df_live_out[~df_live_out["ステージID"].isin(disabled_stage_ids)]
    else:
        df_stage_visible = df_stage
        df_live_visible = df_live
        df_live_out_visible = df_live_out

    return {
        "stage": df_stage,
        "idolname": df_idolname,
        "live": df_live_out,
        "duration_distribution": build_duration_distribution(df_live_out_visible, df_stage_visible),
        "group_count": build_group_appearance_count(df_live_out_visible, df_stage_visible),
        "overlap_alerts": build_overlap_alerts(df_live_visible, df_stage_visible),
        "group_appearances": build_group_appearances(df_live_visible, df_stage_visible),
    }


def build_all_event_outputs(
    pj_path: str,
    project_info_json: dict,
) -> dict[str, dict[str, pd.DataFrame]]:
    """全イベントの出力データを組み立てて返す。

    データが存在しないイベントは空dictで保持する(キーは保つ)。
    後段の determine_id_master / export_excel / listup_new_idolname は
    空dictを検知してスキップする。
    """
    event_list = repo.get_event_name_list(project_info_json)
    result: dict[str, dict[str, pd.DataFrame]] = {}
    for event_name in event_list:
        event_no = repo.get_event_no_by_event_name(project_info_json, event_name)
        data = build_event_output(pj_path, event_name, event_no, project_info_json)
        result[event_name] = data if data is not None else {}
    return result


# ---------------------------------------------------------------------------
# ID確定
# ---------------------------------------------------------------------------

def determine_id_master(
    output_df: dict[str, dict[str, pd.DataFrame]],
    pj_path: str,
    project_info_json: dict,
    target_event_names: list[str] | None = None,
    only_event_types: list[str] | None = None,
) -> None:
    """ステージマスタ・グループマスタ・出番マスタのIDを確定させ、CSV/JSONに保存する。

    ステージID は各 stage_*.json のトップレベルと project_info.json の
    stage_list[i].stage_id に書き戻される。

    `output_df[event_name]["stage"]` / `"live"` は非活性化を含む全行を持つため、
    本関数は何度呼ばれても非活性化ステージ行が CSV から消失することはない。

    `target_event_names` を指定すると当該イベントのみ書き出す
    (auto-trigger からの呼び出し用)。None の場合は全イベントを対象とする。

    `only_event_types` を指定すると、stage_id 書き戻し対象の種別を限定する
    (採番済み種別のみを確定し、未採番種別の stage_*.json へは書き戻さない)。
    未採番種別へ書き戻すと `id_apply_to_json` が該当行0件で例外になるため、
    `output_df` 側も同じ種別群で build した subset を渡すこと。None で全種別。
    """
    event_list = repo.get_event_name_list(project_info_json)
    if target_event_names is not None:
        target_set = set(target_event_names)
        event_list = [ev for ev in event_list if ev in target_set]
    only_types_set = set(only_event_types) if only_event_types is not None else None
    for event_name in event_list:
        if not output_df.get(event_name):
            continue
        output_path = os.path.join(pj_path, event_name)
        output_df[event_name]["stage"].to_csv(os.path.join(output_path, "master_stage.csv"))
        output_df[event_name]["idolname"].to_csv(os.path.join(output_path, "master_idolname.csv"))
        turn_id_data = output_df[event_name]["live"]
        turn_id_data.to_csv(os.path.join(output_path, "turn_id_data.csv"))

        event_no = repo.get_event_no_by_event_name(project_info_json, event_name)
        event_type_list = repo.get_event_type_list(project_info_json, event_no)
        if only_types_set is not None:
            event_type_list = [et for et in event_type_list if et in only_types_set]
        for event_type in event_type_list:
            tgt_event_type_info = repo.get_image_entry_by_dir_name(
                project_info_json, event_no, event_type,
            )
            for stage_no in range(tgt_event_type_info["stage_num"]):
                stage_name = repo.get_stage_name(project_info_json, event_no, event_type, stage_no)
                json_path = os.path.join(output_path, event_type, f"stage_{stage_no}.json")
                if os.path.exists(json_path):
                    with open(json_path, encoding="utf-8") as f:
                        json_data = json.load(f)
                    json_data = timetabledata.id_apply_to_json(
                        json_data, turn_id_data, stage_name,
                        tgt_event_type_info["kind"] == "live_tokutenkai_heiki",
                    )
                    with open(json_path, "w", encoding="utf8") as f:
                        json.dump(json_data, f, indent=4, ensure_ascii=False)

                    # project_info.stage_list[stage_no].stage_id にも同期書き込み
                    top_stage_id = json_data.get("ステージID")
                    if top_stage_id is not None:
                        try:
                            repo.set_stage_id(
                                project_info_json, event_no, event_type, stage_no,
                                int(top_stage_id),
                            )
                        except (KeyError, ValueError):
                            pass
    # project_info.json を永続化(stage_id の書き込みを反映)
    repo.save_project_json(pj_path, project_info_json)


# ---------------------------------------------------------------------------
# ID整合チェック
# ---------------------------------------------------------------------------

def detect_id_anomalies(event_data: dict[str, pd.DataFrame]) -> list[str]:
    """`build_event_output` の戻り値に対し ID の内部整合異常を検出する。

    検出する不変条件 (いずれも違反でブロック対象):
      - 出番ID(=live の index) は単一の (ステージID, ライブ_from) に対応すること。
        異なる時間枠/ステージに跨る出番IDは異常
        (コラボは同一ステージ・同一 ライブ_from のみ正当)。
      - グループID(=idolname の index) と グループ名_採用 が 1:1 であること。
      - ステージID(=stage の index) と ステージ名 が 1:1 であること。

    戻り値: 日本語の異常メッセージ一覧 (空なら正常)。
    """
    messages: list[str] = []

    # --- 出番ID: 単一 (ステージID, ライブ_from) ---
    live = event_data.get("live")
    if live is not None and len(live) > 0:
        key_cols = [c for c in ("ステージID", "ライブ_from") if c in live.columns]
        if key_cols:
            df = live[key_cols].copy()
            df["__turn__"] = live.index
            combos = df.drop_duplicates()
            counts = combos.groupby("__turn__").size()
            for turn_id in counts[counts > 1].index:
                rows = combos[combos["__turn__"] == turn_id]
                detail = " / ".join(
                    "(" + ", ".join(f"{c}={row[c]}" for c in key_cols) + ")"
                    for _, row in rows.iterrows()
                )
                messages.append(
                    f"出番ID={turn_id} が複数の時間枠/ステージに重複しています: {detail}",
                )

    # --- グループID ↔ グループ名_採用 1:1 ---
    messages += _detect_id_name_anomalies(
        event_data.get("idolname"), "グループ名_採用", "グループID",
    )
    # --- ステージID ↔ ステージ名 1:1 ---
    messages += _detect_id_name_anomalies(
        event_data.get("stage"), "ステージ名", "ステージID",
    )
    return messages


def _detect_id_name_anomalies(
    df: pd.DataFrame | None, name_col: str, id_label: str,
) -> list[str]:
    """index(=ID) と `name_col` が 1:1 でない箇所を検出する補助関数。"""
    if df is None or len(df) == 0 or name_col not in df.columns:
        return []
    messages: list[str] = []
    work = pd.DataFrame({"__id__": list(df.index), "__name__": df[name_col].tolist()})
    # 同一IDが複数名に対応
    by_id = work.groupby("__id__")["__name__"].nunique()
    for _id in by_id[by_id > 1].index:
        names = sorted(set(work[work["__id__"] == _id]["__name__"].tolist()))
        messages.append(f"{id_label}={_id} が複数の名前に対応しています: {names}")
    # 同一名が複数IDに対応
    by_name = work.groupby("__name__")["__id__"].nunique()
    for name in by_name[by_name > 1].index:
        ids = sorted(set(work[work["__name__"] == name]["__id__"].tolist()))
        messages.append(f"{name_col}「{name}」が複数の {id_label} に対応しています: {ids}")
    return messages


def _cells_equal(a, b) -> bool:
    """master差分比較用のセル等価判定 (NaN/数値ゆれを吸収)。"""
    a_na, b_na = pd.isna(a), pd.isna(b)
    if a_na and b_na:
        return True
    if a_na or b_na:
        return False
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return str(a) == str(b)


def detect_master_diff(
    event_data: dict[str, pd.DataFrame], output_path: str,
) -> list[str]:
    """再ビルド後 live と永続化済み turn_id_data.csv を比較し、
    両方に存在する出番IDで (グループID, ステージID, ライブ_from) が変化した点を
    非ブロックの「要確認」通知として返す。

    新規追加された出番ID (再ビルド側のみ) は通知しない。
    本関数は採番済み (turn_id_data.csv 存在) 時のみ意味を持つ。
    """
    live = event_data.get("live")
    if live is None or len(live) == 0:
        return []
    csv_path = os.path.join(output_path, "turn_id_data.csv")
    if not os.path.exists(csv_path):
        return []
    try:
        old = pd.read_csv(csv_path, index_col=0)
    except (OSError, ValueError, pd.errors.ParserError):
        return []
    compare_cols = [
        c for c in ("グループID", "ステージID", "ライブ_from")
        if c in live.columns and c in old.columns
    ]
    if not compare_cols:
        return []

    messages: list[str] = []
    common_ids = set(live.index.tolist()) & set(old.index.tolist())
    for turn_id in sorted(common_ids):
        nrow = live.loc[turn_id, compare_cols]
        orow = old.loc[turn_id, compare_cols]
        if isinstance(nrow, pd.DataFrame):
            nrow = nrow.iloc[0]
        if isinstance(orow, pd.DataFrame):
            orow = orow.iloc[0]
        diffs = [c for c in compare_cols if not _cells_equal(nrow[c], orow[c])]
        if diffs:
            messages.append(
                f"出番ID={turn_id}: {'/'.join(diffs)} が前回確定時から変化しています",
            )
    return messages


# ---------------------------------------------------------------------------
# S3保存
# ---------------------------------------------------------------------------

def save_to_s3(pj_name: str) -> None:
    """プロジェクトデータをS3にアップロードする。"""
    s3access.put_project_data(pj_name)


# ---------------------------------------------------------------------------
# Excel出力
# ---------------------------------------------------------------------------

def _join_collab_values(series: pd.Series, col_name: str) -> str:
    """コラボ出番の 1 列分を「,」連結した文字列にする。

    `グループID` は float 表記 (3.0) を避けるため int 文字列に正規化する。
    欠損値はスキップする。
    """
    out: list[str] = []
    for v in series.tolist():
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        if col_name == "グループID":
            try:
                out.append(str(int(v)))
                continue
            except (ValueError, TypeError):
                pass
        out.append(str(v))
    return ",".join(out)


def _merge_collab_live_rows(df_live: pd.DataFrame) -> pd.DataFrame:
    """コラボ出番 (同一 出番ID の複数行) を 1 行に統合する。

    - `グループID` / `グループ名` / `グループ名_raw` はカンマ区切りで連結する
      (例: グループID "3,5")。
    - 単独出番 (1 行) の値はスカラのまま保持する (グループID は数値のまま)。
    - その他の列 (時刻・ステージ・備考・コラボタイトル等) は先頭行の値を採用する。
    - 行順は元データの 出番ID 初出順を維持する。
    """
    if df_live is None or len(df_live) == 0:
        return df_live
    index_name = df_live.index.name
    columns = list(df_live.columns)
    join_cols = [c for c in _LIVE_COLLAB_JOIN_COLUMNS if c in df_live.columns]

    merged_rows: list[dict] = []
    merged_index: list = []
    for turn_id, group in df_live.groupby(level=0, sort=False):
        head = group.iloc[0].to_dict()
        if len(group) > 1:
            for col in join_cols:
                head[col] = _join_collab_values(group[col], col)
        merged_rows.append(head)
        merged_index.append(turn_id)

    merged = pd.DataFrame(merged_rows, index=pd.Index(merged_index, name=index_name))
    return merged[columns]


def _build_metadata_excel_df(
    event_metadata: dict | None,
    project_meta: dict | None,
) -> pd.DataFrame:
    """Stella メタデータ → Excel 出力用の縦持ち (項目, 値) DataFrame。

    参考の `JSON生成シート.xlsm` の【共通】ブロックに相当する。値は
    `build_stella_json` が参照するものと同じ (保存済みの stella_metadata /
    stella_project_meta) を出力する。
    """
    md = event_metadata or {}
    pm = project_meta or {}
    rows = [
        ("ライブ名", pm.get("liveName", "")),
        ("公演日", md.get("date", "")),
        ("liveId", md.get("liveId", "")),
        ("bundleId", md.get("bundleId", "")),
        ("JSONバージョン", md.get("jsonVersion", "")),
        ("開始時", md.get("openTime", "")),
        ("終了時", md.get("closeTime", "")),
        ("お知らせバージョン", md.get("notificationVersion", "")),
        ("お知らせメッセージ", md.get("notification", "")),
    ]
    return pd.DataFrame(
        {"値": ["" if v is None else v for _, v in rows]},
        index=pd.Index([k for k, _ in rows], name="Stellaメタデータ"),
    )


def _build_stage_excel_df(
    df_stage: pd.DataFrame,
    disabled_stage_ids: set,
) -> pd.DataFrame:
    """stage マスタ → Excel 出力用 DataFrame (非活性化除外・表示順ソート・列整形)。"""
    df = df_stage
    if disabled_stage_ids:
        df = df[~df.index.isin(disabled_stage_ids)]
    if "表示順" in df.columns:
        df = df.sort_values("表示順")
    cols = [c for c in _STAGE_EXCEL_COLUMNS if c in df.columns]
    return df[cols]


def _build_live_excel_df(
    df_live: pd.DataFrame,
    disabled_stage_ids: set,
) -> pd.DataFrame:
    """live マスタ → Excel 出力用 DataFrame (非活性化除外・コラボ統合・列整形)。"""
    df = df_live
    if disabled_stage_ids and "ステージID" in df.columns:
        df = df[~df["ステージID"].isin(disabled_stage_ids)]
    df = _merge_collab_live_rows(df)
    drop_cols = [c for c in _LIVE_EXCEL_DROP_COLUMNS if c in df.columns]
    if drop_cols:
        df = df.drop(columns=drop_cols)
    return df.rename(columns={"グループ名_raw": "グループ名(OCR)"})


def export_excel(
    output_df: dict[str, dict[str, pd.DataFrame]],
    pj_path: str,
    event_list: list[str],
    metadata_by_event: dict[str, dict] | None = None,
    project_meta: dict | None = None,
) -> str:
    """Excel形式でデータを出力し、出力パスを返す。

    各イベントを 1 シートとし、(Stellaメタデータ) / stage / idolname / live の
    ブロックを左から順に横並びで配置する。配置開始列は前ブロックの列数から
    動的に算出するため、列数の多寡でブロックが重なって上書きされることはない。

    `metadata_by_event` ({event_name: stella_metadata}) と `project_meta`
    (stella_project_meta) を渡すと、各シート先頭に Stella メタデータブロックを
    出力する。未指定なら従来通りメタデータ列は出力しない。
    """
    output_path = os.path.join(pj_path, "output.xlsx")
    wb = Workbook()
    for event_name in event_list:
        if not output_df.get(event_name):
            continue
        # Excel出力は従来から非活性化ステージとその出番を除外する仕様 (可視のみ)
        df_stage_master = output_df[event_name]["stage"]
        disabled_stage_ids: set = set()
        if "非活性化フラグ" in df_stage_master.columns:
            disabled_stage_ids = set(
                df_stage_master[df_stage_master["非活性化フラグ"]].index.tolist()
            )

        blocks = []
        if metadata_by_event is not None:
            blocks.append(_build_metadata_excel_df(
                metadata_by_event.get(event_name), project_meta,
            ))
        blocks += [
            _build_stage_excel_df(df_stage_master, disabled_stage_ids),
            output_df[event_name]["idolname"].rename(
                columns={"グループ名_採用": "グループ名"}
            ),
            _build_live_excel_df(output_df[event_name]["live"], disabled_stage_ids),
        ]

        col = 1
        for df_to_write in blocks:
            save_dataframe_to_excel(wb, event_name, df_to_write, (col, 1))
            # ブロックは [col, col + 列数] を占有 (先頭 1 列は index)。+1 列の余白を空ける。
            col += len(df_to_write.columns) + 2
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)
    wb.save(output_path)
    return output_path


def save_dataframe_to_excel(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    position: tuple[int, int],
) -> None:
    """DataFrameをExcelワークブックの指定位置に書き込む。"""
    existing_sheets = wb.sheetnames
    if sheet_name not in existing_sheets:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]
    for i, row in enumerate(df.itertuples(), start=position[1]):
        for j, value in enumerate(row, start=position[0]):
            ws.cell(row=i + 1, column=j, value=value)
    for j, header in enumerate(df.columns, start=position[0]):
        ws.cell(row=position[1], column=j + 1, value=header)
    ws.cell(row=position[1], column=position[0], value=df.index.name)


# ---------------------------------------------------------------------------
# グループ名マスタ
# ---------------------------------------------------------------------------

def listup_new_idolname(
    output_df: dict[str, dict[str, pd.DataFrame]],
    event_list: list[str],
) -> pd.DataFrame:
    """新しく出現したグループ名をリストアップする。"""
    idol_name_all = []
    for event_name in event_list:
        if not output_df.get(event_name):
            continue
        idol_name_all.extend(list(output_df[event_name]["idolname"]["グループ名_採用"]))
    new_idol_name = idolname.detect_new_data(list(set(idol_name_all)))
    return pd.DataFrame({
        "追加": [True for _ in range(len(new_idol_name))],
        "グループ名": new_idol_name,
    }).sort_values(by="グループ名").reset_index(drop=True)


def update_master_idolname(
    df_new_idolname: pd.DataFrame,
    data_path: str,
) -> None:
    """新しく出現したグループ名をマスタに追加し、S3にアップロードする。"""
    new_idolname = list(df_new_idolname[df_new_idolname["追加"]]["グループ名"])
    idolname.add_new_data_file(new_idolname)

    json_path = os.path.join(data_path, "master/master_version_s3.json")
    with open(json_path, 'r', encoding='utf-8') as f:
        master_version_s3 = json.load(f)
    jst = ZoneInfo("Asia/Tokyo")
    now_jst = datetime.now(jst)
    updated_at = now_jst.strftime('%Y/%m/%d %H:%M:%S.%f')
    master_version_s3["idolname_embedding_data.csv"] = updated_at
    master_version_s3["idolname_latest.csv"] = updated_at
    with open(json_path, "w", encoding="utf8") as f:
        json.dump(master_version_s3, f, indent=4, ensure_ascii=False)

    s3_prefix = "master"
    s3access.upload_s3_file(s3_prefix, "master_version_s3.json", os.path.join(data_path, "master/master_version_s3.json"))
    s3access.upload_s3_file(s3_prefix, "idolname_embedding_data.csv", os.path.join(data_path, "master/idolname_embedding_data.csv"))
    s3access.upload_s3_file(s3_prefix, "idolname_latest.csv", os.path.join(data_path, "master/idolname_latest.csv"))
