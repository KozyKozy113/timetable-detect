import streamlit as st
from st_aggrid import AgGrid
from st_aggrid.grid_options_builder import GridOptionsBuilder
from streamlit_cropper import st_cropper
# from streamlit_cropperjs import st_cropperjs
# from streamlit_drawable_canvas import st_canvas

import os
import shutil
import tempfile
from operator import itemgetter
# import asyncio

from PIL import Image, ImageDraw
import cv2
from matplotlib import pyplot as plt

import numpy as np
import pandas as pd
from openpyxl import Workbook#, load_workbook
from io import BytesIO
from datetime import datetime
from datetime import time as dttime
from datetime import timedelta
from zoneinfo import ZoneInfo
import json

from backend_functions import gpt_ocr, s3access, timetabledata, idolname
from frontend_functions import timetablepicture, timetable_difference

st.set_page_config(
    page_title="タイムテーブル読み取りアプリ", 
    # page_icon=image, 
    layout="wide", 
    # initial_sidebar_state="auto", 
    menu_items={
        'Get Help': 'https://www.google.com',
        'Report a bug': "https://www.google.com",
        'About': """
        # アイドル対バンタイムテーブル読み取りツール
        ライブ管理アプリへの搭載用のデータを生成できます。
        """
     })

st.title("タイムテーブル読み取りアプリ")
st.markdown(
"""### タイムテーブル画像を構造化データに変換します
- ライブのタイムテーブル画像をアップロードしてください
- ライブの形式を選択して、OCRや生成AIを活用し構造化データに変換します
- 精度は100%ではないので、人の手で修正を行ってください
- 最終的に出来上がった構造化データは、現時点ではStella用に形式を変換してダウンロードすることができます 
""")

DIR_PATH = os.path.dirname(__file__)
DATA_PATH = DIR_PATH +"/../data"

if "pj_name" not in st.session_state:#初期化
    s3access.get_master()
    st.session_state.pj_name = None
    st.session_state.project_master = pd.read_csv(os.path.join(DATA_PATH, "master", "projects_master.csv"), index_col=0)
    st.session_state.project_master_s3 = pd.read_csv(os.path.join(DATA_PATH, "master", "projects_master_s3.csv"), index_col=0)
    # st.session_state.timetable_image_master = pd.read_csv(os.path.join(DATA_PATH, "master", "timetable_image_master.csv"))
pj_name_list = pd.concat((st.session_state.project_master_s3,st.session_state.project_master)).sort_values(by="updated_at",ascending=False).index.to_list() #作成が新しい順に並ぶ
pj_name_list = list(dict.fromkeys(pj_name_list))

def make_project(pj_name=None):
    if pj_name is None:
        pj_name = st.session_state.new_pj_name
    if pj_name in pj_name_list:
        with project_setting:
            st.error("既に存在する名前のプロジェクトです")
    else:
        os.makedirs(os.path.join(DATA_PATH, "projects", pj_name), exist_ok=True)
        created_at = datetime.now().strftime('%Y/%m/%d %H:%M:%S.%f')
        st.session_state.project_master.loc[pj_name] = [created_at,created_at,"フェス",1]
        st.session_state.project_master_s3.loc[pj_name] = [created_at,created_at,"フェス",1]
        st.session_state.project_master.to_csv(os.path.join(DATA_PATH, "master", "projects_master.csv"))
        st.session_state.project_master_s3.to_csv(os.path.join(DATA_PATH, "master", "projects_master_s3.csv"))
        project_info_json = {
            "project_name":pj_name,
            "event_num":1,
            "event_detail":[
                {
                    "event_no":0,
                    "event_name":"event_1",
                    "timetables":{}
                }
                            ]
        }
        json_path = os.path.join(DATA_PATH, "projects", pj_name, "project_info.json")
        with open(json_path,"w",encoding = "utf8") as f:
            json.dump(project_info_json, f, indent = 4, ensure_ascii = False)
        set_project(pj_name)

def set_project(pj_name):
    st.session_state.pj_name = pj_name
    st.session_state.exist_pj_name = pj_name
    st.session_state.pj_path = os.path.join(DATA_PATH, "projects", pj_name)
    
    #S3のproject_masterと比較して更新日時が古いorローカルにそもそもPJが無い場合、データ取得し直す
    s3access.get_master()
    s3access.get_project_data(pj_name)
    st.session_state.project_master = pd.read_csv(os.path.join(DATA_PATH, "master", "projects_master.csv"), index_col=0)
    project_info = st.session_state.project_master.loc[st.session_state.pj_name]

    if not pd.isna(project_info["event_type"]):
        st.session_state.event_type = project_info["event_type"]
    else:
        st.session_state.event_type = "対バン"
    if not pd.isna(project_info["event_num"]):
        st.session_state.event_num = project_info["event_num"]
    else:
        st.session_state.event_num = 1
    for i in range(st.session_state.event_num):
        os.makedirs(os.path.join(st.session_state.pj_path, "event_{}".format(i+1)), exist_ok=True)
    st.session_state.project_info_json = get_project_json()
    #その他以後の変数も初期化等する必要あり
    set_crop_image()
    set_ocr_image()
    # st.session_state.images_eachstage=[]#実際にはロードできるならロードする
    # st.session_state.timeline_eachstage=[]#同上

def get_project_json():
    json_path = os.path.join(st.session_state.pj_path, "project_info.json")
    with open(json_path, 'r', encoding='utf-8') as f:
        json_data = json.load(f)
    return json_data

def update_project_timestamp():
    jst = ZoneInfo("Asia/Tokyo")# 日本時間のタイムゾーンオブジェクトを作成    
    now_jst = datetime.now(jst)# 日本時間で現在時刻を取得
    updated_at = now_jst.strftime('%Y/%m/%d %H:%M:%S.%f')
    st.session_state.project_master.loc[st.session_state.pj_name,"updated_at"] = updated_at
    st.session_state.project_master.to_csv(os.path.join(DATA_PATH, "master", "projects_master.csv"))

def set_project_json(json_data):
    json_path = os.path.join(st.session_state.pj_path, "project_info.json")
    with open(json_path,"w",encoding = "utf8") as f:
        json.dump(json_data, f, indent = 4, ensure_ascii = False)
    update_project_timestamp()

def determine_project_setting():
    st.session_state.project_master.loc[st.session_state.pj_name,"event_type"] = st.session_state.event_type
    st.session_state.project_master.loc[st.session_state.pj_name,"event_num"] = st.session_state.event_num
    updated_at = datetime.now().strftime('%Y/%m/%d %H:%M:%S.%f')
    st.session_state.project_master.loc[st.session_state.pj_name,"updated_at"] = updated_at
    st.session_state.project_master.to_csv(os.path.join(DATA_PATH, "master", "projects_master.csv"))
    st.session_state.project_info_json["event_num"]=st.session_state.event_num
    st.session_state.project_info_json["event_detail"]=st.session_state.project_info_json["event_detail"][:st.session_state.event_num]#本当はイベントフォルダも消した方がいい
    for i in range(st.session_state.event_num):
        os.makedirs(os.path.join(DATA_PATH, "projects", st.session_state.pj_name, "event_"+str(i+1)), exist_ok=True)
        if len(st.session_state.project_info_json["event_detail"])-1<i:
            st.session_state.project_info_json["event_detail"].append({
                "event_no":i,
                "event_name":"event_{}".format(i+1),
                "timetables":{}
            })
        # os.makedirs(os.path.join(DATA_PATH, "projects", st.session_state.pj_name, "event_"+str(i), "ライブ"), exist_ok=True)
        # os.makedirs(os.path.join(DATA_PATH, "projects", st.session_state.pj_name, "event_"+str(i), "特典会"), exist_ok=True)
        # os.makedirs(os.path.join(DATA_PATH, "projects", st.session_state.pj_name, "event_"+str(i), "両方"), exist_ok=True)
    set_project_json(st.session_state.project_info_json)

@st.cache_data
def get_image(img_path):
    return Image.open(img_path)

def determine_timetable_image():
    img_event_no = get_event_no_by_event_name(st.session_state.img_event_name)
    file = st.session_state.uploaded_image
    if st.session_state.img_type in ["ライブ","特典会"]:
        img_type = st.session_state.img_type
        img_path = os.path.join(st.session_state.pj_path, st.session_state.img_event_name, img_type)
        os.makedirs(img_path, exist_ok=True)
        img_path = os.path.join(img_path, "raw.png")
        with open(img_path, 'wb') as f:
            f.write(file.read())
        st.session_state.project_info_json["event_detail"][img_event_no]["timetables"][img_type]={"format":st.session_state.img_format,"stage_num":0,"stage_list":[]}
        with col_file_uploader[1]:
            st.success("画像を登録しました")
        st.session_state.crop_tgt_event = st.session_state.img_event_name
        st.session_state.crop_tgt_img_type = img_type
        st.session_state.ocr_tgt_event = st.session_state.img_event_name
        st.session_state.ocr_tgt_img_type = img_type
    elif st.session_state.img_type == "両方(特典会別添え)":
        file_image = file.read()
        for img_type in ["ライブ","特典会"]:
            img_path = os.path.join(st.session_state.pj_path, st.session_state.img_event_name, img_type)
            os.makedirs(img_path, exist_ok=True)
            img_path = os.path.join(img_path, "raw.png")
            with open(img_path, 'wb') as f:
                f.write(file_image)
            st.session_state.project_info_json["event_detail"][img_event_no]["timetables"][img_type]={"format":st.session_state.img_format,"stage_num":0,"stage_list":[]}
        st.session_state.crop_tgt_event = st.session_state.img_event_name
        st.session_state.crop_tgt_img_type = "ライブ"
        st.session_state.ocr_tgt_event = st.session_state.img_event_name
        st.session_state.ocr_tgt_img_type = "ライブ"
        with col_file_uploader[1]:
            st.success("画像を登録しました")
    elif st.session_state.img_type == "両方(特典会併記)":
        img_type = "ライブ特典会"
        img_path = os.path.join(st.session_state.pj_path, st.session_state.img_event_name, img_type)
        os.makedirs(img_path, exist_ok=True)
        img_path = os.path.join(img_path, "raw.png")
        with open(img_path, 'wb') as f:
            f.write(file.read())
        st.session_state.project_info_json["event_detail"][img_event_no]["timetables"][img_type]={"format":"特典会併記","stage_num":0,"stage_list":[]}
        with col_file_uploader[1]:
            st.success("画像を登録しました")
        st.session_state.crop_tgt_event = st.session_state.img_event_name
        st.session_state.crop_tgt_img_type = img_type
        st.session_state.ocr_tgt_event = st.session_state.img_event_name
        st.session_state.ocr_tgt_img_type = img_type
    elif st.session_state.img_type == "その他":
        img_type = st.session_state.img_type_alternative
        img_path = os.path.join(st.session_state.pj_path, st.session_state.img_event_name, img_type)
        os.makedirs(img_path, exist_ok=True)
        img_path = os.path.join(img_path, "raw.png")
        with open(img_path, 'wb') as f:
            f.write(file.read())
        st.session_state.project_info_json["event_detail"][img_event_no]["timetables"][img_type]={"format":st.session_state.img_format,"stage_num":0,"stage_list":[]}
        with col_file_uploader[1]:
            st.success("画像を登録しました")
        st.session_state.crop_tgt_event = st.session_state.img_event_name
        st.session_state.crop_tgt_img_type = img_type
        st.session_state.ocr_tgt_event = st.session_state.img_event_name
        st.session_state.ocr_tgt_img_type = img_type
    elif st.session_state.img_type == "その他(特典会併記)":
        img_type = st.session_state.img_type_alternative
        img_path = os.path.join(st.session_state.pj_path, st.session_state.img_event_name, img_type)
        os.makedirs(img_path, exist_ok=True)
        img_path = os.path.join(img_path, "raw.png")
        with open(img_path, 'wb') as f:
            f.write(file.read())
        st.session_state.project_info_json["event_detail"][img_event_no]["timetables"][img_type]={"format":"特典会併記","stage_num":0,"stage_list":[]}
        with col_file_uploader[1]:
            st.success("画像を登録しました")
        st.session_state.crop_tgt_event = st.session_state.img_event_name
        st.session_state.crop_tgt_img_type = img_type
        st.session_state.ocr_tgt_event = st.session_state.img_event_name
        st.session_state.ocr_tgt_img_type = img_type
    set_project_json(st.session_state.project_info_json)

def delete_uploaded_image(img_event_no, img_type):
    img_event_name = get_event_name(img_event_no)
    del st.session_state.project_info_json["event_detail"][img_event_no]["timetables"][img_type]
    set_project_json(st.session_state.project_info_json)
    # shutil.rmtree(os.path.join(st.session_state.pj_path, img_event_name, img_type))#フォルダごと削除
    next_img_type = get_event_type_list(img_event_no)
    if len(next_img_type)==0:
        st.session_state.crop_tgt_img_type = None
        st.session_state.ocr_tgt_img_type = None
    else:
        if st.session_state.crop_tgt_img_type == img_type:
            st.session_state.crop_tgt_img_type = next_img_type[0]
        if st.session_state.ocr_tgt_img_type == img_type:
            st.session_state.ocr_tgt_img_type = next_img_type[0]
    with col_file_uploader[1]:
        st.success("画像を削除しました")

def get_event_name(event_no):
    return st.session_state.project_info_json["event_detail"][event_no]["event_name"]

def get_event_name_list():
    return [json_data["event_name"] for json_data in st.session_state.project_info_json["event_detail"]]

def get_event_type_list(event_no):#ライブ、特典会を先頭にしつつイベントごとに存在する画像種別のリストを出力する
    event_type_all = list(st.session_state.project_info_json["event_detail"][event_no]["timetables"].keys())
    event_type_list = []
    for event_type in ["ライブ","特典会","ライブ特典会"]:
        if event_type in event_type_all:
            event_type_list.append(event_type)
    for event_type in event_type_all:
        if event_type not in event_type_list:
            event_type_list.append(event_type)
    return event_type_list

def get_event_no_by_event_name(event_name):
    for event_detail in st.session_state.project_info_json["event_detail"]:
        if event_detail["event_name"]==event_name:
            return event_detail["event_no"]
    else:
        return None

def get_stage_name_list(event_no,img_type):
    return [stage_info["stage_name"] for stage_info in st.session_state.project_info_json["event_detail"][event_no]["timetables"][img_type]["stage_list"]]

def get_stage_name(event_no,img_type,stage_no):
    return st.session_state.project_info_json["event_detail"][event_no]["timetables"][img_type]["stage_list"][stage_no]["stage_name"]

def set_crop_image():
    # st.session_state.crop_tgt_event
    # st.session_state.crop_tgt_img_type
    st.session_state.images_eachstage=[]

def get_x_freq(image, stage_num):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)#グレースケールへの変換   
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)#エッジ検出
    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)#輪郭検出
    #各輪郭を囲む矩形を取得し、抽出
    rectangles = []
    for i, contour in enumerate(contours):
        x, y, w, h = cv2.boundingRect(contour)
        rectangles.append((x, y, w, h))

    df_rectangle = pd.DataFrame(rectangles,columns=["横位置","縦位置","横幅","縦幅"])

    fig, ax = plt.subplots()
    min_width = image.shape[1]/stage_num/3
    x_appear = pd.Series(df_rectangle[df_rectangle["横幅"]>min_width]["横位置"])
    # Streamlitでヒストグラムを表示
    # ax.hist(x_appear, bins=100, edgecolor='black')
    # with timetable_ocr:
    #     st.pyplot(fig)
    return x_appear.value_counts()

def get_xpoint(image, stage_num):
    x_freq = get_x_freq(np.array(image), stage_num)
    response = gpt_ocr.get_xpoint(x_freq, stage_num)

    return json.loads(response.choices[0].message.content)["xpoint"]
    
    with timetable_ocr:
        st.write(json.loads(response.choices[0].message.content)["xpoint"])

def detect_stageline(image):#ステージ領域を特定する縦線を取得
    img_path = os.path.join(st.session_state.pj_path, st.session_state.crop_tgt_event, st.session_state.crop_tgt_img_type, "raw_cropped.png")
    st.session_state.cropped_image.save(img_path)

    bgr_img = cv2.cvtColor(np.array(image), cv2.COLOR_BGR2RGB)
    gray_img = cv2.cvtColor(bgr_img, cv2.COLOR_BGR2GRAY)
    height, width = gray_img.shape
    minlength = height * st.session_state.x_minlength_rate

    im_edges = cv2.Canny(gray_img, st.session_state.x_edge_threshold_1, st.session_state.x_edge_threshold_2, L2gradient=True)#エッジ検出
    lines = []
    lines = cv2.HoughLinesP(im_edges, rho=1, theta=np.pi/360, threshold=st.session_state.x_hough_threshold, minLineLength=minlength, maxLineGap=st.session_state.x_hough_gap)#ハフ変換による直線抽出

    line_list = []
    image_copy = image.copy()
    draw = ImageDraw.Draw(image_copy)
    for line in lines:
        x1, y1, x2, y2 = line[0]
        if x1 == x2 or abs((y1 - y2)/(x1 - x2)) > np.tan(np.pi/180*85):#傾きが85度より大きいのもののみ検出
            line_list.append([x1, y1, x2, y2, abs(y1-y2)])
            draw.line((x1, y1, x2, y2), fill="red", width=20)

    line_list.sort(key=itemgetter(0, 1, 2, 3))
    line_x_list = pd.DataFrame(line_list,columns=["x1","y1","x2","y2","length"]).groupby("x1").sum()[["length"]]
    x_before=0
    new_line_x_list=[]
    for x,row in line_x_list.iterrows():#近接線の統合
        if x-x_before < st.session_state.x_identify_interval:
            if len(new_line_x_list)>0:
                new_line_x_list[-1][0].append(x)
                new_line_x_list[-1][1]+=row["length"]
            else:
                new_line_x_list.append([[0,x],row["length"]])
        else:
            if len(new_line_x_list)>0:
                new_line_x_list[-1][0] = np.mean(new_line_x_list[-1][0])
            new_line_x_list.append([[x],row["length"]])
        x_before=x
    if len(new_line_x_list)>0:
        new_line_x_list[-1][0] = np.mean(new_line_x_list[-1][0])

    st.session_state.stage_line_list = pd.DataFrame(new_line_x_list,columns=["x","length"])
    # x_mintotallength_rate = 0.03
    # st.session_state.stage_line_list = st.session_state.stage_line_list.query(" length > {}".format(int(height * x_mintotallength_rate)))

    left = 0
    num = len(st.session_state.stage_line_list)+1
    st.session_state.images_eachstage=[]
    for i in range(num):
        if i<num-1:
            right = st.session_state.stage_line_list.iat[i,0]
            # draw.line((right, 0, right, height), fill="red", width=10)
        else:
            right =width
        st.session_state.images_eachstage.append(image.crop((left, 0, right, height)))
        left = right

    with edge_result:
        # st.write(st.session_state.stage_line_list)
        st.image(image_copy,caption="縦線抽出結果")
        #ここで抽出に失敗している時のリカバリー（クリックで分割線を追加）をできるようにする

def get_image_eachstage_byocr(image, stage_num):
    if stage_num ==1:
        return [image]
    for _ in range(3):
        try:
            xpoints = get_xpoint(image, stage_num)
            break
        except ValueError as e:
            continue
    else:
        raise ValueError
    print(stage_num)
    print(xpoints)
    images_eachstage = []
    for i in range(stage_num):
        x_from = xpoints[i]
        if i < stage_num - 1:
            x_to = xpoints[i+1]
        else:
            x_to = image.width
        width=x_to-x_from
        if i>0:
            x_from = x_from-width/10
        else:
            x_from = 0
        rect_img = image.crop((x_from, 0, x_to, image.height))
        # img_path = os.path.join(st.session_state.pj_path, st.session_state.crop_tgt_event, st.session_state.crop_tgt_img_type, "stage_{}.png".format(i+1))
        # rect_img.save(img_path)
        images_eachstage.append(rect_img)
    return images_eachstage

def get_image_eachstage_for_linecroppedimage_byocr():
    st.session_state.images_eachstage = []
    for i,line_cropped_image in enumerate(st.session_state.line_cropped_images):
        if st.session_state["stage_num_{}".format(i)]>0:
            st.session_state.images_eachstage += get_image_eachstage_byocr(line_cropped_image, st.session_state["stage_num_{}".format(i)])

def get_image_eachstage_for_linecroppedimage_byevenly(): #使ってない
    st.session_state.images_eachstage = []
    for i,line_cropped_image in enumerate(st.session_state.line_cropped_images):
        stage_num = st.session_state["stage_num_{}".format(i)]
        if stage_num>0:
            width, height = line_cropped_image.size
            segment_width = width // stage_num
            for j in range(stage_num):
                left = max(0, (j - 0.05) * segment_width)
                right = min((j + 1.05) * segment_width,width)
                segment = line_cropped_image.crop((left, 0, right, height))
                st.session_state.images_eachstage.append(segment)

def get_image_eachstage_for_croppedimage_byevenly():
    img_path = os.path.join(st.session_state.pj_path, st.session_state.crop_tgt_event, st.session_state.crop_tgt_img_type, "raw_cropped.png")
    st.session_state.cropped_image.save(img_path)
    st.session_state.images_eachstage = []
    stage_num = st.session_state.devide_stage_num
    if stage_num>0:
        width, height = st.session_state.cropped_image.size
        segment_width = width // stage_num
        for j in range(stage_num):
            left = max(0, (j - 0.05) * segment_width)
            right = min((j + 1.05) * segment_width,width)
            segment = st.session_state.cropped_image.crop((left, 0, right, height))
            st.session_state.images_eachstage.append(segment)

def determine_image_eachstage():
    event_no = get_event_no_by_event_name(st.session_state.crop_tgt_event)
    #前に保存した画像を削除するべきではある
    img_path = os.path.join(st.session_state.pj_path, st.session_state.crop_tgt_event, st.session_state.crop_tgt_img_type, "raw_cropped.png")
    st.session_state.cropped_image.save(img_path)
    for i, image_eachstag in enumerate(st.session_state.images_eachstage):
        img_path = os.path.join(st.session_state.pj_path, st.session_state.crop_tgt_event, st.session_state.crop_tgt_img_type, "stage_{}.png".format(i))
        image_eachstag.save(img_path)
        if len(st.session_state.project_info_json["event_detail"][event_no]["timetables"][st.session_state.crop_tgt_img_type]["stage_list"]) <= i:
            st.session_state.project_info_json["event_detail"][event_no]["timetables"][st.session_state.crop_tgt_img_type]["stage_list"].append({"stage_no":i,"stage_name":"ステージ{}".format(i)})
    
    st.session_state.project_info_json["event_detail"][event_no]["timetables"][st.session_state.crop_tgt_img_type]["stage_num"]=len(st.session_state.images_eachstage)
    set_project_json(st.session_state.project_info_json)
    # st.session_state.timetable_image_master.loc[
    #     (st.session_state.timetable_image_master["project_name"]==st.session_state.pj_name)
    #     &(st.session_state.timetable_image_master["event_no"]==int(st.session_state.crop_tgt_event.split("_")[1]))
    #     &(st.session_state.timetable_image_master["image_type"]==st.session_state.crop_tgt_img_type),"stage_num"]=len(st.session_state.images_eachstage)
    # st.session_state.timetable_image_master.to_csv(os.path.join(DATA_PATH, "master", "timetable_image_master.csv"), index=False)
    # st.session_state.pj_timetable_master = st.session_state.timetable_image_master[st.session_state.timetable_image_master["project_name"]==st.session_state.pj_name]

def determine_image_eachstage_without_nocheck():
    event_no = get_event_no_by_event_name(st.session_state.crop_tgt_event)
    #前に保存した画像を削除するべきではある
    stage_num = 0
    for i, image_eachstag in enumerate(st.session_state.images_eachstage):
        if st.session_state["each_stage_accept_{}".format(i)]:
            img_path = os.path.join(st.session_state.pj_path, st.session_state.crop_tgt_event, st.session_state.crop_tgt_img_type, "stage_{}.png".format(stage_num))
            image_eachstag.save(img_path)
            if len(st.session_state.project_info_json["event_detail"][event_no]["timetables"][st.session_state.crop_tgt_img_type]["stage_list"]) <= stage_num:
                st.session_state.project_info_json["event_detail"][event_no]["timetables"][st.session_state.crop_tgt_img_type]["stage_list"].append({"stage_no":stage_num,"stage_name":"ステージ{}".format(stage_num)})
            stage_num+=1

    st.session_state.project_info_json["event_detail"][event_no]["timetables"][st.session_state.crop_tgt_img_type]["stage_num"]=stage_num
    set_project_json(st.session_state.project_info_json)
    # st.session_state.timetable_image_master.loc[
    #     (st.session_state.timetable_image_master["project_name"]==st.session_state.pj_name)
    #     &(st.session_state.timetable_image_master["event_no"]==int(st.session_state.crop_tgt_event.split("_")[1]))
    #     &(st.session_state.timetable_image_master["image_type"]==st.session_state.crop_tgt_img_type),"stage_num"]=stage_num
    # st.session_state.timetable_image_master.to_csv(os.path.join(DATA_PATH, "master", "timetable_image_master.csv"), index=False)
    # st.session_state.pj_timetable_master = st.session_state.timetable_image_master[st.session_state.timetable_image_master["project_name"]==st.session_state.pj_name]

def save_time_pixel(time_start, top, height, total_duration):
    time_format = "%H:%M"
    time_start = time_start.strftime(time_format)
    st.session_state.project_info_json["event_detail"][get_event_no_by_event_name(st.session_state.crop_tgt_event)]["timetables"][st.session_state.crop_tgt_img_type]["time_pixel"]={
        "time_start" : time_start
        ,"start_pix" : top
        ,"total_pix" : height
        ,"total_duration" : total_duration
    }
    set_project_json(st.session_state.project_info_json)

def set_ocr_image():
    # st.session_state.ocr_tgt_event
    # st.session_state.ocr_tgt_img_type
    st.session_state.time_axis_detect = None
    st.session_state.timeline_eachstage=[]

def pix_to_time(pix):#ピクセル値を時刻に変換する関数
    time_format = "%H:%M"
    try:
        time_pixel = st.session_state.project_info_json["event_detail"][get_event_no_by_event_name(st.session_state.ocr_tgt_event)]["timetables"][st.session_state.ocr_tgt_img_type]["time_pixel"]
        time_start = datetime.strptime(time_pixel["time_start"], time_format).time()
        start_pix = time_pixel["start_pix"]
        total_pix = time_pixel["total_pix"]
        total_duration = time_pixel["total_duration"]
        # min = np.round((pix-st.session_state.start_pix)/(st.session_state.total_pix/st.session_state.total_duration*5))*5
        # return (datetime(2024,1,1,st.session_state.time_start.hour,st.session_state.time_start.minute)+timedelta(minutes=min)).time()
        min = np.round((pix-start_pix)/(total_pix/total_duration*5))*5
        return (datetime(2024,1,1,time_start.hour,time_start.minute)+timedelta(minutes=min)).time()
    except KeyError:
        st.warning("基準時間を設定してください")

def time_to_pix(tgt_time):#時刻をピクセル値に変換する関数
    time_format = "%H:%M"
    try:
        time_pixel = st.session_state.project_info_json["event_detail"][get_event_no_by_event_name(st.session_state.ocr_tgt_event)]["timetables"][st.session_state.ocr_tgt_img_type]["time_pixel"]
        time_start = datetime.strptime(time_pixel["time_start"], time_format).time()
        start_pix = time_pixel["start_pix"]
        total_pix = time_pixel["total_pix"]
        total_duration = time_pixel["total_duration"]
        # min = (datetime.combine(datetime.today(), tgt_time) - datetime.combine(datetime.today(), st.session_state.time_start)).total_seconds() / 60
        # return st.session_state.start_pix + int(min*st.session_state.total_pix/st.session_state.total_duration)
        min = (datetime.combine(datetime.today(), tgt_time) - datetime.combine(datetime.today(), time_start)).total_seconds() / 60
        return start_pix + int(min*total_pix/total_duration)
    except KeyError:
        return None
        # st.warning("基準時間を設定してください")

def time_length_to_pix(minutes, int_flag=True):#時間の長さをピクセル値に変換する関数
    # time_format = "%H:%M"
    try:
        time_pixel = st.session_state.project_info_json["event_detail"][get_event_no_by_event_name(st.session_state.ocr_tgt_event)]["timetables"][st.session_state.ocr_tgt_img_type]["time_pixel"]
        # time_start = datetime.strptime(time_pixel["time_start"], time_format).time()
        # start_pix = time_pixel["start_pix"]
        total_pix = time_pixel["total_pix"]
        total_duration = time_pixel["total_duration"]
        if int_flag:
            return int(minutes*total_pix/total_duration)
        else:
            return minutes*total_pix/total_duration
    except KeyError:
        return None
        # st.warning("基準時間を設定してください")

def get_timetabledata_onlyonestage(stage_no,user_prompt):
    img_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.png".format(stage_no))
    user_prompt = "この画像のタイムテーブルをJSONデータとして出力して。" + user_prompt
    return_json = gpt_ocr.getocr_fes_timetable_functioncalling(img_path,user_prompt)
    return_json["ステージ名"] = st.session_state.project_info_json["event_detail"][get_event_no_by_event_name(st.session_state.ocr_tgt_event)]["timetables"][st.session_state.ocr_tgt_img_type]["stage_list"][stage_no]["stage_name"]
    # if "ステージ名" not in return_json or return_json["ステージ名"] == "不明":
    #     return_json["ステージ名"] = "ステージ{}".format(stage_no)#st.session_state.stage_names[stage_no]
    # else:
    #     if str(return_json["ステージ名"]).isdigit():
    #         stage_name_prefix = st.session_state.ocr_tgt_img_type
    #         return_json["ステージ名"] = stage_name_prefix+str(return_json["ステージ名"])
    #     set_stage_name(stage_no, return_json["ステージ名"])
    json_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.json".format(stage_no))
    with open(json_path,"w",encoding = "utf8") as f:
        json.dump(return_json, f, indent = 4, ensure_ascii = False)
    output_timetable_picture_onlyonestage(stage_no)
    update_project_timestamp()

def get_timetabledata_eachstage(user_prompt):
    for i in range(st.session_state.ocr_tgt_stage_num):
        get_timetabledata_onlyonestage(i,user_prompt)

def get_timetabledata_withtokutenkai_onlyonestage(stage_no,user_prompt):
    img_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.png".format(stage_no))
    user_prompt = "この画像のタイムテーブルをJSONデータとして出力して。" + user_prompt
    return_json = gpt_ocr.getocr_fes_withtokutenkai_timetable_functioncalling(img_path,user_prompt)
    return_json["ステージ名"] = st.session_state.project_info_json["event_detail"][get_event_no_by_event_name(st.session_state.ocr_tgt_event)]["timetables"][st.session_state.ocr_tgt_img_type]["stage_list"][stage_no]["stage_name"]
    json_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.json".format(stage_no))
    with open(json_path,"w",encoding = "utf8") as f:
        json.dump(return_json, f, indent = 4, ensure_ascii = False)
    output_timetable_picture_onlyonestage(stage_no)
    update_project_timestamp()

def get_timetabledata_withtokutenkai_eachstage(user_prompt):
    for i in range(st.session_state.ocr_tgt_stage_num):
        get_timetabledata_withtokutenkai_onlyonestage(i,user_prompt)

def detect_timeline_onlyonestage(stage_no):
    if len(st.session_state.timeline_eachstage)!=st.session_state.ocr_tgt_stage_num:
        st.session_state.timeline_eachstage = [None for _ in range(st.session_state.ocr_tgt_stage_num)]

    img_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.png".format(stage_no))
    if os.path.exists(img_path):
        image = Image.open(img_path)
    else:
        raise ValueError

    bgr_img = cv2.cvtColor(np.array(image), cv2.COLOR_BGR2RGB)
    gray_img = cv2.cvtColor(bgr_img, cv2.COLOR_BGR2GRAY)
    height, width = gray_img.shape
    minlength = height * st.session_state.y_minlength_rate

    # if st.session_state.y_binary_threshold>0:
    #     _, gray_img = cv2.threshold(gray_img, st.session_state.y_binary_threshold, 255, cv2.THRESH_BINARY_INV)
    im_edges = cv2.Canny(gray_img, st.session_state.y_edge_threshold_1, st.session_state.y_edge_threshold_2, L2gradient=True)#エッジ検出
    lines = []
    lines = cv2.HoughLinesP(im_edges, rho=1, theta=np.pi/360, threshold=st.session_state.y_hough_threshold, minLineLength=minlength, maxLineGap=st.session_state.y_hough_gap)#ハフ変換による直線抽出
    if lines is None or len(lines)==0:
        return

    line_list = []
    image_copy = image.copy()
    draw = ImageDraw.Draw(image_copy)
    for line in lines:
        x1, y1, x2, y2 = line[0]
        if x1 != x2 and abs((y1 - y2)/(x1 - x2)) < np.tan(np.pi/180*5):#傾きが5度より小さいもののみ検出
            line_list.append([x1, y1, x2, y2, abs(x1-x2)])
            draw.line((x1, y1, x2, y2), fill="white", width=10)

    line_list.sort(key=itemgetter(1, 0, 2, 3))
    line_y_list = pd.DataFrame(line_list,columns=["x1","y1","x2","y2","length"]).groupby("y1").sum()[["length"]]
    y_before=0
    new_line_y_list=[]
    for y,row in line_y_list.iterrows():#近接線の統合
        if y-y_before < st.session_state.y_identify_interval:
            if len(new_line_y_list)>0:
                new_line_y_list[-1][0].append(y)
                new_line_y_list[-1][1]+=row["length"]
            else:
                new_line_y_list.append([[0,y],row["length"]])
        else:
            if len(new_line_y_list)>0:
                new_line_y_list[-1][0] = np.mean(new_line_y_list[-1][0])
            new_line_y_list.append([[y],row["length"]])
        y_before=y
    if len(new_line_y_list)>0:
        new_line_y_list[-1][0] = np.mean(new_line_y_list[-1][0])
    st.session_state.timeline_eachstage[stage_no-1] = pd.DataFrame(new_line_y_list,columns=["y","length"])
    # y_mintotallength_rate = 0.03
    # st.session_state.timeline_eachstage[stage_no-1] = st.session_state.timeline_eachstage[stage_no-1].query(" length > {}".format(int(height * y_mintotallength_rate)))
    st.session_state.timeline_eachstage[stage_no-1]["time"] = st.session_state.timeline_eachstage[stage_no-1]["y"].map(pix_to_time)
    # st.session_state.timeline_eachstage[stage_no-1]["time"].to_csv(img_path.replace(".png","_timeline.csv"),index=False)
    
    # font_size = int(st.session_state.total_pix/st.session_state.total_duration*4)
    font_size = time_length_to_pix(4)
    face = cv2.FONT_HERSHEY_PLAIN
    thickness = 1#パラメータ化する？
    scale = cv2.getFontScaleFromHeight(face, font_size, thickness)
    font_pixel, baseline = cv2.getTextSize("00:00-00:00", face, scale, thickness)
    if font_pixel[0]<width:
        # font_size = int(st.session_state.total_pix/st.session_state.total_duration*4*width/font_pixel[0])
        font_size = time_length_to_pix(4*width/font_pixel[0])
        face = cv2.FONT_HERSHEY_PLAIN
        thickness = 1#パラメータ化する？
        scale = cv2.getFontScaleFromHeight(face, font_size, thickness)
        font_pixel, baseline = cv2.getTextSize("00:00-00:00", face, scale, thickness)

    #時刻情報を画像に追記
    # img_time = bgr_img.copy()
    # for _ ,row in st.session_state.timeline_eachstage[stage_no-1].iterrows():
    #     # cv2.putText(img_time, row["time"].strftime('%H:%M'), (int(width*0.95-font_pixel[0]), int(row["y"])-int(font_size*0.2)), cv2.FONT_HERSHEY_PLAIN, scale, (0,0,0), thickness)
    #     cv2.putText(img_time, row["time"].strftime('%H:%M'), (int(width*0.05), int(row["y"])-int(font_size*0.2)), cv2.FONT_HERSHEY_PLAIN, scale, (0,0,0), thickness)
    #     cv2.putText(img_time, row["time"].strftime('%H:%M'), (int(width*0.05), int(row["y"])+int(font_size*1.2)), cv2.FONT_HERSHEY_PLAIN, scale, (0,0,0), thickness)
    
    #時刻情報を画像右側に拡張して追記
    extension_width = int(font_pixel[0]*1.2)#右側に拡張する幅
    extension_image = np.full((height, width + extension_width, 3), (255, 255, 255), dtype=np.uint8)#拡張した領域を持つ新しい画像
    extension_image[:, :width] = bgr_img#元の画像を新しい画像の左側に貼り付ける
    for j in range(len(st.session_state.timeline_eachstage[stage_no-1])-1):
        if st.session_state.y_ignoretime_threshold<(datetime.combine(datetime.today(),st.session_state.timeline_eachstage[stage_no-1].loc[j+1,"time"])-datetime.combine(datetime.today(),st.session_state.timeline_eachstage[stage_no-1].loc[j,"time"])).seconds/60:
            time_pix = (st.session_state.timeline_eachstage[stage_no-1].loc[j,"y"]+st.session_state.timeline_eachstage[stage_no-1].loc[j+1,"y"])/2
            time_stamp = st.session_state.timeline_eachstage[stage_no-1].loc[j,"time"].strftime('%H:%M') + "-" + st.session_state.timeline_eachstage[stage_no-1].loc[j+1,"time"].strftime('%H:%M')
            cv2.putText(extension_image, time_stamp, (width+int(width*0.05), int(time_pix)+int(font_size*0.6)), cv2.FONT_HERSHEY_PLAIN, scale, (0,0,0), thickness)
    for _ ,row in st.session_state.timeline_eachstage[stage_no-1].iterrows():
        cv2.line(extension_image, (width, int(row["y"])), (width + extension_width, int(row["y"])), (0,0,0), 1)

    with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as tmpfile:
        temp_path = tmpfile.name
        cv2.imwrite(temp_path, extension_image)
    shutil.move(temp_path, img_path.replace(".png","_addtime.png"))

    # with tmp_timeline:
    #     img_time = cv2.cvtColor(img_time, cv2.COLOR_BGR2RGB)
    #     st.image(img_time)

    # left = 0
    # num = len(st.session_state.stage_line_list)+1
    # st.session_state.images_eachstage=[]
    # for i in range(num):
    #     if i<num-1:
    #         right = st.session_state.stage_line_list.iat[i,0]
    #         draw.line((right, 0, right, height), fill="red", width=10)
    #     else:
    #         right =width
    #     st.session_state.images_eachstage.append(image.crop((left, 0, right, height)))
    #     left = right

    # with tmp_timeline:
    #     st.write(st.session_state.timeline_eachstage[stage_no-1])
    #     st.image(image_copy)
        #ここで抽出に失敗している時のリカバリー（クリックで分割線を追加）
    update_project_timestamp()

def detect_timeline_eachstage():
    if len(st.session_state.timeline_eachstage)!=st.session_state.ocr_tgt_stage_num:
        st.session_state.timeline_eachstage = [None for _ in range(st.session_state.ocr_tgt_stage_num)]
    for i in range(st.session_state.ocr_tgt_stage_num):
        detect_timeline_onlyonestage(i)

def get_timetabledata_onlyonestage_notime(stage_no,user_prompt):
    img_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}_addtime.png".format(stage_no))
    user_prompt = "この画像のタイムテーブルをJSONデータとして出力して。" + user_prompt
    if not os.path.exists(img_path):
        detect_timeline_onlyonestage(stage_no)#時刻の読み取り
    if st.session_state.ocr_tgt_img_type == "ライブ":
        return_json = gpt_ocr.getocr_fes_timetable_notime_functioncalling(img_path,user_prompt)
    elif st.session_state.ocr_tgt_img_type == "特典会":
        return_json = gpt_ocr.getocr_fes_timetable_notime_functioncalling(img_path,user_prompt,live=False)
    return_json["ステージ名"] = st.session_state.project_info_json["event_detail"][get_event_no_by_event_name(st.session_state.ocr_tgt_event)]["timetables"][st.session_state.ocr_tgt_img_type]["stage_list"][stage_no]["stage_name"]
    json_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.json".format(stage_no))
    with open(json_path,"w",encoding = "utf8") as f:
        json.dump(return_json, f, indent = 4, ensure_ascii = False)
    output_timetable_picture_onlyonestage(stage_no)
    update_project_timestamp()

# async def get_timetabledata_eachstage_notime_async(user_prompt):
#     tasks = [get_timetabledata_onlyonestage_notime(i,user_prompt) for i in range(st.session_state.ocr_tgt_stage_num)]
#     await asyncio.gather(*tasks)

def get_timetabledata_eachstage_notime(user_prompt):
    for i in range(st.session_state.ocr_tgt_stage_num):
        get_timetabledata_onlyonestage_notime(i,user_prompt)
    # asyncio.run(get_timetabledata_eachstage_notime_async(user_prompt))

def idolname_correct_onlyonestage(stage_no):#, idolname_confirmed_list=None
    if st.session_state.correct_idolname_in_confirmed_list:
    #     if idolname_confirmed_list is None:
        idolname_confirmed_list = get_idolname_confirmed_list()
    #     if len(idolname_confirmed_list)==0:
    #         st.session_state.correct_idolname_in_confirmed_list=False

    json_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.json".format(stage_no))
    if os.path.exists(json_path):
        with open(json_path, encoding="utf-8") as f:
            timetable_json = json.load(f)
        for item in timetable_json["タイムテーブル"]:
            if st.session_state.correct_idolname_in_confirmed_list:
                item['グループ名_採用'] = idolname.get_name_by_inlist(item["グループ名"], idolname_confirmed_list)
            else:
                item['グループ名_採用'] = idolname.get_name_by_levenshtein_and_vector(item["グループ名"])
            # group_name_correct = idolname.get_name_list_by_vector(item["グループ名"])
            # if not group_name_correct[0]:
            #     item['グループ名_採用'] = group_name_correct[1]
            # else:
            #     item['グループ名_採用'] = item["グループ名"]
        with open(json_path,"w",encoding = "utf8") as f:
            json.dump(timetable_json, f, indent = 4, ensure_ascii = False)
        output_timetable_picture_onlyonestage(stage_no)
    update_project_timestamp()

def idolname_correct_eachstage():#アイドル名の修正
    if st.session_state.correct_idolname_in_confirmed_list:
        idolname_confirmed_list = get_idolname_confirmed_list()
        if len(idolname_confirmed_list)==0:
            st.session_state.correct_idolname_in_confirmed_list = False
    for i in range(st.session_state.ocr_tgt_stage_num):
        idolname_correct_onlyonestage(i)

def get_idolname_confirmed_list():#確定したアイドル名一覧の取得
    idolname_confirmed_list=[]
    event_no = get_event_no_by_event_name(st.session_state.ocr_tgt_event)
    event_type_list = get_event_type_list(event_no)
    for event_type in event_type_list:
        for stage_info in st.session_state.project_info_json["event_detail"][event_no]["timetables"][event_type]["stage_list"]:
            stage_no = stage_info["stage_no"]
            json_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, event_type, "stage_{}.json".format(stage_no))
            if os.path.exists(json_path):
                with open(json_path, encoding="utf-8") as f:
                    timetable_json = json.load(f)
            for group_stage in timetable_json["タイムテーブル"]:
                if "グループ名_採用" in group_stage:
                    if type(group_stage["グループ名_採用"])==str and len(group_stage["グループ名_採用"])>0:
                        idolname_confirmed_list.append(group_stage["グループ名_採用"])
    return list(set(idolname_confirmed_list))

def get_stagelist(user_prompt):#OCRによるステージ名一覧の読み取り
    img_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "raw_cropped.png")
    try:
        stage_list, rule = gpt_ocr.getocr_fes_stagelist(img_path, st.session_state.ocr_tgt_stage_num, user_prompt)
        if rule in ["数字", "アルファベット"]:
            prefix_flag=True
        else:
            prefix_flag=False
        for i in range(st.session_state.ocr_tgt_stage_num):
            if prefix_flag:
                stage_name=st.session_state.ocr_tgt_img_type+str(stage_list[i])
            else:
                stage_name=str(stage_list[i])
            st.session_state.project_info_json["event_detail"][get_event_no_by_event_name(st.session_state.ocr_tgt_event)]["timetables"][st.session_state.ocr_tgt_img_type]["stage_list"][i]["stage_name"]=stage_name
        set_project_json(st.session_state.project_info_json)
    except:
        print("ステージ名がうまく取得できませんでした")

def set_stage_name(stage_no, stage_name):#ステージ名の修正
    st.session_state.project_info_json["event_detail"][get_event_no_by_event_name(st.session_state.ocr_tgt_event)]["timetables"][st.session_state.ocr_tgt_img_type]["stage_list"][stage_no]["stage_name"]=stage_name
    set_project_json(st.session_state.project_info_json)

    json_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.json".format(stage_no))
    if os.path.exists(json_path):
        with open(json_path, encoding="utf-8") as f:
            json_old = json.load(f)
    else:
        json_old={}
    json_old["ステージ名"] = stage_name
    with open(json_path,"w",encoding = "utf8") as f:
        json.dump(json_old, f, indent = 4, ensure_ascii = False)

def booth_name_add_prefix_onlyonestage(stage_no):#特典会併記タイテにおいてブース名に会場名を接頭辞として付加する
    ocr_tgt_event_no = get_event_no_by_event_name(st.session_state.ocr_tgt_event)
    stage_name = get_stage_name(ocr_tgt_event_no,st.session_state.ocr_tgt_img_type,stage_no)
    st.session_state.df_timetables[stage_no]["ブース"] = st.session_state.df_timetables[stage_no]["ブース"].apply(lambda x: x if x.startswith(stage_name) else stage_name + x)
    save_timetable_data_onlyonestage(stage_no)

def booth_name_add_prefix_eachstage():
    for i in range(st.session_state.ocr_tgt_stage_num):
        booth_name_add_prefix_onlyonestage(i)

def get_timetabledata_together():
    event_list = get_event_name_list()
    for i,event_name in enumerate(event_list):
        event_type_list = get_event_type_list(i)
        st.session_state.ocr_tgt_event = event_name
        if st.session_state.correct_idolname_in_confirmed_list_toghther:
            idolname_confirmed_list = get_idolname_confirmed_list()
            if len(idolname_confirmed_list)==0:
                st.session_state.correct_idolname_in_confirmed_list = False
            else:
                st.session_state.correct_idolname_in_confirmed_list = True
        else:
            st.session_state.correct_idolname_in_confirmed_list = False
        for event_type in event_type_list:
            if st.session_state["together_"+event_list[i]+"/"+event_type]:
                st.session_state.ocr_tgt_img_type = event_type
                st.session_state.ocr_tgt_image_info = st.session_state.project_info_json["event_detail"][i]["timetables"][event_type]
                st.session_state.ocr_tgt_stage_num = st.session_state.ocr_tgt_image_info["stage_num"]
                if st.session_state.together_ocr_stage:
                    get_stagelist(st.session_state.ocr_stage_user_prompt_together)
                if st.session_state.together_ocr_timetable:
                    if st.session_state.ocr_tgt_image_info["format"]=="ライムライト式":
                        get_timetabledata_eachstage_notime(st.session_state.ocr_user_prompt_together)
                    elif st.session_state.ocr_tgt_image_info["format"]=="特典会併記":
                        get_timetabledata_withtokutenkai_eachstage(st.session_state.ocr_user_prompt_together)
                    else:
                        get_timetabledata_eachstage(st.session_state.ocr_user_prompt_together)
                if st.session_state.toghther_correct:
                    idolname_correct_eachstage()

def save_timetable_data_onlyonestage(stage_no):
    df_timetable = st.session_state.df_timetables[stage_no]
    json_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.json".format(stage_no))
    json_timetable = timetabledata.df_to_json(df_timetable)
    if os.path.exists(json_path):
        with open(json_path, encoding="utf-8") as f:
            json_old = json.load(f)
    else:
        json_old={}
    json_old["タイムテーブル"]=json_timetable
    json_old["ステージ名"]=st.session_state["stage_name_stage{}".format(stage_no)]
    set_stage_name(stage_no,json_old["ステージ名"])
    with open(json_path,"w",encoding = "utf8") as f:
        json.dump(json_old, f, indent = 4, ensure_ascii = False)
    output_timetable_picture_onlyonestage(stage_no)
    update_project_timestamp()

def save_timetable_data_eachstage():
    for i in range(st.session_state.ocr_tgt_stage_num):
        save_timetable_data_onlyonestage(i)

def output_timetable_picture_onlyonestage(stage_no):#読み取り結果から作られる構造化データを逆に画像化
    json_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.json".format(stage_no))
    output_path = json_path.replace(".json","_timetable.png")
    with open(json_path, encoding="utf-8") as f:
        json_data = json.load(f)
    if st.session_state.ocr_output_picture_time_match:
        #基準時刻のpixと時間幅pixを計算
        time_format = "%H:%M"
        start_time = min(datetime.strptime(live["ライブステージ"]["from"], time_format) 
                     for live in json_data["タイムテーブル"]).time()
        start_time = start_time.replace(minute=0)
        start_margin = time_to_pix(start_time)
        time_line_spacing = time_length_to_pix(30,False)
        timetable_image = timetablepicture.create_timetable_image(json_data, start_margin, time_line_spacing)
    else:
        timetable_image = timetablepicture.create_timetable_image(json_data)
    timetable_image.save(output_path)
    update_project_timestamp()

def output_timetable_picture_eachstage():
    for i in range(st.session_state.ocr_tgt_stage_num):
        output_timetable_picture_onlyonestage(i)

def output_difference_image(new_image):
    old_image_path = os.path.join(st.session_state.pj_path, st.session_state.diff_tgt_event, st.session_state.diff_tgt_img_type, "raw.png")
    old_image = Image.open(old_image_path)
    difference_image = timetable_difference.output_difference(Image.open(new_image), old_image)
    with timetable_compare_col[1]:
        st.image(difference_image)

# def get_all_stage_info():#全ステージ情報の出力 #暫定
#     all_stage_df = pd.concat(st.session_state.df_timetables).reset_index(drop=True)
#     with all_stage_info:
#         st.dataframe(all_stage_df,hide_index=True)

def determine_id_master():#ステージマスタやグループマスタ、出番マスタのIDが変動しないよう確定させる
    event_list = get_event_name_list()
    for event_name in event_list:
        output_path =  os.path.join(st.session_state.pj_path, event_name)
        st.session_state.output_df[event_name]["stage"].to_csv(os.path.join(output_path, "master_stage.csv"))
        st.session_state.output_df[event_name]["idolname"].to_csv(os.path.join(output_path, "master_idolname.csv"))
        turn_id_data = st.session_state.output_df[event_name]["live"]
        turn_id_data.to_csv(os.path.join(output_path, "turn_id_data.csv"))
        event_no = get_event_no_by_event_name(event_name)
        event_type_list = get_event_type_list(event_no)
        for event_type in event_type_list:
            tgt_event_type_info = st.session_state.project_info_json["event_detail"][event_no]["timetables"][event_type]
            # stage_name_list = get_stage_name_list(edit_tgt_event_no,event_type)
            for stage_no in range(tgt_event_type_info["stage_num"]):
                stage_name = get_stage_name(event_no,event_type,stage_no)
                json_path = os.path.join(output_path, event_type, "stage_{}.json".format(stage_no))
                if os.path.exists(json_path):
                    with open(json_path, encoding="utf-8") as f:
                        json_data = json.load(f)
                    json_data = timetabledata.id_apply_to_json(json_data, turn_id_data, stage_name, tgt_event_type_info["format"]=="特典会併記")
                    with open(json_path,"w",encoding = "utf8") as f:
                        json.dump(json_data, f, indent = 4, ensure_ascii = False)
    update_project_timestamp()        

def save_to_s3():#プロジェクトのデータをS3に上書き保存する
    s3access.put_project_data(st.session_state.pj_name)

def output_data_for_stella():#Excel形式でデータを出力する
    output_path =  os.path.join(st.session_state.pj_path, "output.xlsx")
    wb = Workbook()
    event_list = get_event_name_list()
    for event_name in event_list:
        for df_type, potision in zip(["stage","idolname","live"], [(1,1),(5,1),(8,1)]):
            save_dataframe_to_excel(wb, event_name, st.session_state.output_df[event_name][df_type], potision)
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)
    wb.save(output_path)

def save_dataframe_to_excel(wb, sheet_name, df, potision):#あるpandas.DataFrameをExcelワークブックの指定したシートの指定した位置に保存する
    #potisionは(2,5)=(B5)
    existing_sheets = wb.sheetnames
    if sheet_name not in existing_sheets:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]
    for i, row in enumerate(df.itertuples(), start=potision[1]):
        for j, value in enumerate(row, start=potision[0]):
            ws.cell(row=i+1, column=j, value=value)
    for j, header in enumerate(df.columns, start=potision[0]):
        ws.cell(row=potision[1], column=j+1, value=header)
    ws.cell(row=potision[1], column=potision[0], value=df.index.name)


project_setting = st.container()#プロジェクト設定
with project_setting:
    st.markdown("""#### ①プロジェクトの設定""")
    col_project_setting = st.columns(3)
    project_setting_determine = st.container()
with col_project_setting[0]:
    st.info(
"""既存のプロジェクトを呼び出すか、新しくプロジェクトを作成してください。  
一つのプロジェクトは一つのイベントに紐付きます。  
""")
    col_makepj = st.columns((5,1))
    with col_makepj[0]:
        st.text_input(label="新しいプロジェクト名", placeholder="入力してください", key="new_pj_name")
    with col_makepj[1]:
        st.button(label="作成", on_click=make_project)
    col_setpj = st.columns((5,1))
    with col_setpj[0]:
        st.selectbox("既存のプロジェクト一覧"
                                , pj_name_list
                                , placeholder = "プロジェクトを選択または作成してください"
                                , key="exist_pj_name")
    with col_setpj[1]:
        st.button(label="呼出", on_click=set_project, args=(st.session_state.exist_pj_name,))
    if st.session_state.pj_name is None:
        st.stop()
    st.text("選択中のプロジェクト："+st.session_state.pj_name)
with col_project_setting[1]:
    st.info(
"""イベント形式を選択してください。  
・フェス：複数のステージが同時進行するイベント  
・対バン：ステージが一つのみのイベント  
※Stella向けなので、一旦フェスのみに対応して作っています  
""")
    st.radio("イベント形式", ("対バン", "フェス"), key="event_type", horizontal=True)
with col_project_setting[2]:
    st.info(
"""イベント数を入力してください。  
イベント数とは、チケットの販売単位に相当する概念です。  
複数日にわたって開かれるフェスや、昼夜で別イベントとして開かれる対バンの場合、その数を記入してください。  
""")
    st.number_input("イベント数", min_value=1, step=1, key="event_num")
with project_setting_determine:
    if st.session_state.event_type is not None and st.session_state.event_num is not None:
        st.button(label="プロジェクト設定を反映する", on_click=determine_project_setting, type="primary")
st.divider()

file_uploader = st.container()#タイテ画像の登録
with file_uploader:
    st.markdown("""#### ②タイムテーブル画像の登録""")
    all_files_raw =  st.container(height=200)
    col_file_uploader = st.columns((3,1))
with all_files_raw:
    st.markdown("""###### 登録済みタイムテーブル画像一覧""")
    image_num = 0
    event_list = get_event_name_list()
    for i in range(len(event_list)):
        image_num += len(st.session_state.project_info_json["event_detail"][i]["timetables"])
        # image_num += len(os.listdir(os.path.join(st.session_state.pj_path, "event_{}".format(i))))
    st.markdown("- 画像数：{}".format(str(image_num)))
    if image_num <1:
        image_num=1
    col_all_files = st.columns(image_num)
    image_idx = 0
    for i, event_name in enumerate(event_list):
        event_type_list = get_event_type_list(i)
        for img_type in event_type_list:
            img_path = os.path.join(st.session_state.pj_path, event_name, img_type, "raw.png")
            if os.path.exists(img_path):
                with col_all_files[image_idx]:
                    col_uploaded_image = st.columns(2)
                    with col_uploaded_image[0]:
                        st.markdown("- {}/{}".format(event_name, img_type))
                    with col_uploaded_image[1]:
                        st.button("削除",key="delete_uploaded_image_{}".format(image_idx),on_click=delete_uploaded_image,args=(i, img_type))
                    image = get_image(img_path)
                    st.image(image)
            image_idx += 1
with col_file_uploader[0]:
    st.file_uploader("読み取りたいタイムテーブル画像をアップロードしてください。"
                            , type=["jpg", "jpeg", "png", "jfif"]
                            , key="uploaded_image")
    if st.session_state.uploaded_image is not None:
        st.image(
            st.session_state.uploaded_image,
            use_container_width=True
        )
with col_file_uploader[1]:
    if st.session_state.uploaded_image is not None:
        st.info(
    """画像が何の情報を示しているか入力してください。  
    ・イベント：複数イベントがある場合（2days開催など）、どのイベントの情報であるか  
    ・種別：画像に載っている情報の種類  
    ・形式：タイムテーブルの形式（通常：時間が各枠に記載 / ライムライト式：時間軸が枠外に記載）
    """)
        event_list = get_event_name_list()
        st.selectbox("イベント", event_list,index=0, key="img_event_name")
        # st.number_input("イベントNo.", min_value=1, max_value=st.session_state.project_info_json["event_num"], step=1, key="img_event_no")
        st.radio("種別", ("ライブ", "特典会", "両方(特典会別添え)", "両方(特典会併記)","その他", "その他(特典会併記)"), key="img_type", horizontal=True)
        st.text_input("その他の種別", key="img_type_alternative")
        if st.session_state.img_type != "両方(特典会併記)" and st.session_state.img_type != "その他(特典会併記)":
            st.radio("形式", ("通常", "ライムライト式"), key="img_format", horizontal=True)
        else:
            st.radio("形式", ("通常", "ライムライト式"), key="img_format", horizontal=True, disabled=True)
        if st.session_state.img_type in ["その他", "その他(特典会併記)"] and (st.session_state.img_type_alternative == "" or st.session_state.img_type_alternative is None):
            st.button(label="画像を登録する", on_click=determine_timetable_image, type="primary", disabled=True)
        else:
            st.button(label="画像を登録する", on_click=determine_timetable_image, type="primary")

if image_idx == 0:#画像登録確認
    st.warning("画像を登録してください")
    st.stop()
st.divider()

timetable_crop = st.container()#タイテ画像の切り取り
with timetable_crop:
    st.markdown("""#### ③タイムテーブル画像の切り取り""")
    st.info(
"""元のタイムテーブル画像に加工などを施して、読み取りが行えるための準備をします。  
大きく分けて3つのステップで準備します。  
- （ⅰ）まず必要最小限の領域に画像を切り出します。  
- （ⅱ）次に複数あるであろうステージごとに画像を分割します。  
- （ⅲ）最後に基準となる時間軸の位置を指定します（推奨だがオプション）。  
""")

    event_list = get_event_name_list()
    st.selectbox("イベント", event_list,index=0,key="crop_tgt_event",on_change=set_crop_image)
    crop_tgt_event_no = get_event_no_by_event_name(st.session_state.crop_tgt_event)
    event_type_list = get_event_type_list(crop_tgt_event_no)
    if len(event_type_list) == 0:
        st.warning("画像を登録するか他のイベントを選択してください")
        st.stop()
    st.selectbox("種別", event_type_list,index=0,key="crop_tgt_img_type",on_change=set_crop_image)
    img_path = os.path.join(st.session_state.pj_path, st.session_state.crop_tgt_event, st.session_state.crop_tgt_img_type, "raw.png")
    if os.path.exists(img_path):
        image = Image.open(img_path)
        image_info = st.session_state.project_info_json["event_detail"][crop_tgt_event_no]["timetables"][st.session_state.crop_tgt_img_type]
        # image_info = st.session_state.pj_timetable_master[
        #     (st.session_state.pj_timetable_master["event_no"]==int(st.session_state.crop_tgt_event.split("_")[1]))
        #     & (st.session_state.pj_timetable_master["image_type"]==st.session_state.crop_tgt_img_type)
        # ].iloc[0]

        with st.container():# タイムテーブルに関係する領域を切り出す
            st.markdown("""###### ③（ⅰ）タイムテーブルに関係する領域を切り出す""")
            st.info(
"""まず、タイムテーブルに関係する領域の切り出しを行います。  
必要十分な領域を指定してください。  
- 上下については、出演枠だけでなくステージ名の部分まで含めることを推奨します（ステージ名自動読み取りが可能になります）。  
- 左右については、時間軸の部分まで含めることを推奨します（時間軸の設定で使います）。  
    - 特にライムライト形式のタイテでは出演枠から時間が分からないため、時間軸の部分が必須になります。  
- また、このあとのステップでステージの均等割を行う場合には、それに適した領域に切り出してください。  
""")
            col_cropimage_first = st.columns([2,1])
            with col_cropimage_first[0]:
                st.markdown("""###### 切り出し設定""")
                col_cropimage_first_setting = st.columns(2)
                with col_cropimage_first_setting[0]:
                    box_color = st.color_picker(label="Box Color", value='#0000FF', key="crop_box_coler")
                with col_cropimage_first_setting[1]:
                    stroke_width = st.number_input(label="Box Thickness", value=1, step=1, key="crop_stroke_width")
                st.session_state.cropped_image = st_cropper(image,
                                box_color=box_color,
                                stroke_width=stroke_width)
                # st.session_state.cropped_image = st_cropper(image)
            with col_cropimage_first[1]:
                st.markdown("""###### 切り出し結果""")
                st.image(st.session_state.cropped_image,use_container_width=True)

        with st.container():# ステージごとにタイムテーブル領域を分割する
            st.markdown("""###### ③（ⅱ）ステージごとにタイムテーブル領域を分割する""")
            st.info(
"""次に、切り出した領域を複数あるステージごとに分割します。  
分割の方法は2種類あり、「縦線検出による分割」と「均等幅での分割」です。  
各ステージの幅が概ね均等であり、また間に時間軸なども等しく入っているor入っていない場合は、「均等幅での分割」を推奨します。  
しかしそうでない場合などには、「縦線を機械的に検出してそれに基づいた分割」を行います。  
手間はかかりますが、縦線検出法ではステージに関係ある情報の部分のみを取得できるというメリットがあります。  
画像を見て、どちらのパターンがふさわしいかを判断して作業を行ってください。  
いずれかの方法で分割を行い、採用不採用を決めたら確定ボタンを押してステージごとの画像を確定してください。
""")
            split_button = st.columns(2)
            with split_button[0]:
                st.info(
"""###### 縦線検出による分割
- エッジ検出とハフ変換という手法によって、縦に長いラインを画像から発見し、そこで画像を分割します。  
- 分割の精度は100%ではありませんが、アルゴリズムのパラメータ変更である程度対応できるようになります。  
- パラメータ変更でも対応できなかったときの場合に、手動で分割併合が出来る機能を実装予定です。  
- 分割された領域の中には、ステージに該当しない余白領域などが存在する可能性がありますので、採用不採用をチェックしてください。  
""")
                st.button("縦ラインの自動抽出による分割",on_click=detect_stageline,args=(st.session_state.cropped_image,),type="primary",help="縦に長いラインを機械的に画像から発見し、そこで画像を分割します。うまく行かない場合はパラメータを変更してください。")
                with st.expander("縦ライン抽出のパラメータ"):
                    st.info(
"""- 「エッジ抽出の閾値」は、特定の色と色の間の線が検出できていない場合に下げると良いです。劇的に下げてもOKです。
- 「ハフ変換の閾値」は、全体的に検出できていない場合に下げると良いです。
- 「抽出したエッジを伸ばす際の閾値」は、検出された線が途切れ途切れになったり短かったりする場合に下げると良いです。「エッジ抽出の閾値」以下に設定してください。
- 「ハフ変換で許容する線分の飛び」は、どの程度の破線を許容するかのパラメータです。連続した文字などの上に不要な線分が検出されてしまっている場合に上げると良いです。
- 「抽出線分の長さ（元画像の縦に対する比率）」は、短い線分を除くためのパラメータです。不要な短い線分が検出されている場合に上げると良いです。
- 「同一視する線分の許容誤差幅」は、線分の位置がほぼ同じ場合にそれらをまとめ上げる範囲についてのパラメータです。
""")
                    st.slider('エッジ抽出の閾値', value=285, min_value=1, max_value=500, step=1, key="x_edge_threshold_2", on_change=detect_stageline,args=(st.session_state.cropped_image,))
                    st.slider('抽出したエッジを伸ばす際の閾値', value=130, min_value=1, max_value=500, step=1, key="x_edge_threshold_1", on_change=detect_stageline,args=(st.session_state.cropped_image,))
                    st.slider('ハフ変換の閾値', value=100, min_value=1, max_value=500, step=1, key="x_hough_threshold", on_change=detect_stageline,args=(st.session_state.cropped_image,))
                    st.slider('ハフ変換で許容する線分の飛び', value=1, min_value=0, max_value=100, step=1, key="x_hough_gap", on_change=detect_stageline,args=(st.session_state.cropped_image,))
                    st.slider('抽出線分の長さ（元画像の縦に対する比率）', value=0.01, min_value=0.0, max_value=1.0, step=0.001, key="x_minlength_rate", on_change=detect_stageline,args=(st.session_state.cropped_image,))
                    st.slider('同一視する線分の許容誤差幅', value=5, min_value=1, max_value=30, step=1, key="x_identify_interval", on_change=detect_stageline,args=(st.session_state.cropped_image,))
            with split_button[1]:
                st.info(
"""###### 均等幅での分割  
- ステージ数を入力し、横幅が均等になるように画像をステージの数だけ分割します。  
- 均等にステージが表現されている場合はこちらの手法が安定します。
- 分割位置がきれいでない場合、一つ前の工程での領域抽出を修正すると良いかもしれません。
- 前後5%ぶん余裕をとっているので、隣り合う画像に重複する部分が残ります。
""")
                st.button("画像を均等な横幅で指定数に分割",on_click=get_image_eachstage_for_croppedimage_byevenly,type="primary")
                st.number_input("ステージ数",1,step=1,key="devide_stage_num")
            edge_result = st.container()
            determine_image_eachstage_button_area_1 = st.container()
            if "images_eachstage" in st.session_state:
                stage_num = len(st.session_state.images_eachstage)
                if stage_num >0:
                    each_stage_area = st.container(height=500)
                    determine_image_eachstage_button_area_2 = st.container()
                    with determine_image_eachstage_button_area_1:
                        st.button("ステージごとの画像を確定",on_click=determine_image_eachstage_without_nocheck,type="primary",key="determine_image_eachstage_button_1")
                    with each_stage_area:
                        col_eachstage = st.columns(stage_num)
                        for i in range(stage_num):
                            with col_eachstage[i]:
                                #画像の幅（最頻値近辺）で採用不採用のデフォルト値を切り替えても良いかも
                                st.checkbox("採用",key="each_stage_accept_{}".format(i),value=True)
                                st.image(st.session_state.images_eachstage[i])
                    with determine_image_eachstage_button_area_2:
                        st.button("ステージごとの画像を確定",on_click=determine_image_eachstage_without_nocheck,key="determine_image_eachstage_button_2")


            # if len(st.session_state.images_eachstage)>0:
            #     col_cropimage_eachstage = st.columns([img.size[0] for img in st.session_state.images_eachstage])
            #     for i,eachstage_image in enumerate(st.session_state.images_eachstage):
            #         with col_cropimage_eachstage[i]:
            #             st.image(eachstage_image)

            #     st.button("ステージごとの画像分割を確定",on_click=determine_image_eachstage, type="primary")

        with st.container():# 基準時間を設定する
            st.markdown("""###### ③（ⅲ）基準時間を設定する（オプション）""")
            st.info(
"""最後に、画像のどのラインが何時に相当するかを指定します。  
画像内の枠線をドラッグして、上端と下端が基準となる時刻になるよう調整してください。  
ライムライト式のタイテでは、この情報を元に時刻の読み取りを行います。  
それ以外の形式のタイテでも、読み取り結果と元画像の比較において時間軸を合わせるのに使います。  
ただし画像によっては時間軸が線形でない（同じ時間の長さが同じ縦幅ではない）場合があるので、その場合は指定しないでください。  
- 横幅は特に関係ないので、上端と下端だけを調整してください。  
- 上下幅は出演枠が存在している範囲でセットしてください。余白部分は微妙に間隔が違ったりします。  
- 上端と下端に対応する時刻を入力して、画像上のどの位置がどの時刻かを対応づけてください。  
- 最後に「時間軸の設定を確定する」ボタンを押してください。  
""")
            cropped_img_path = os.path.join(st.session_state.pj_path, st.session_state.crop_tgt_event, st.session_state.crop_tgt_img_type, "raw_cropped.png")
            if os.path.exists(cropped_img_path):
                cropped_image = Image.open(cropped_img_path)
                col_timeaxis = st.columns([2,1])
                with col_timeaxis[0]:
                    col_timeaxis_setting = st.columns(2)
                    with col_timeaxis_setting[0]:
                        box_color = st.color_picker(label="Box Color", value='#0000FF', key="timeaxis_box_coler")
                    with col_timeaxis_setting[1]:
                        stroke_width = st.number_input(label="Box Thickness", value=1, step=1, key="timeaxs_stroke_width")
                    rect = st_cropper(cropped_image,
                                    box_color=box_color,
                                    stroke_width=stroke_width,
                                    return_type="box")
                    left, top, width, height = tuple(map(int, rect.values()))
                with col_timeaxis[1]:
                    st.image(cropped_image.crop((left, top, left+width, top+height)),use_container_width=True)
                    st.slider('上端時間', value=dttime(10), key="time_start", step=timedelta(minutes=5))
                    st.slider('下端時間', value=dttime(20), key="time_finish", step=timedelta(minutes=5))
                    total_duration = (datetime(2024,1,1,st.session_state.time_finish.hour,st.session_state.time_finish.minute)-datetime(2024,1,1,st.session_state.time_start.hour,st.session_state.time_start.minute)).seconds/60
                    if st.session_state.time_finish.hour < st.session_state.time_start.hour:
                        st.warning("上端時間より下端時間が早くなっています。深夜イベントなどで日を跨ぐ場合はそのまま実行可能ですが、そうでない場合は修正してください。")
                    if "time_pixel" not in image_info:
                        st.warning("基準時間を確定してください")
                        st.button("基準時間を確定する",on_click=save_time_pixel, args=(st.session_state.time_start, top, height, total_duration), type="primary")
                    else:
                        st.button("基準時間を更新する",on_click=save_time_pixel, args=(st.session_state.time_start, top, height, total_duration))

# timetable_crop = st.container()#タイテ画像の切り取り
# with timetable_crop:
#     st.markdown("""#### タイムテーブル画像の切り取り""")
#     st.info(
# """複数ステージが一つの画像に存在している場合、各ステージごとに画像を切り分けます。  
# この時、各ステージの幅が概ね均等であり、また間に時間軸なども等しく入っているor入っていない場合は、  
# 画像を均等に分割することによりステージごとの画像を取得できます。  
# しかし、そうでない場合には「均等割」できるいくつかの小さい画像に分割する必要があります。  
# 画像を見て、どちらのパターンかを判断して作業を行ってください。  
# """)
#     event_list = os.listdir(st.session_state.pj_path)
#     st.selectbox("イベント", event_list,index=0,key="crop_tgt_event")
#     st.selectbox("種別", ["ライブ","特典会"],index=0,key="crop_tgt_img_type")
#     img_path = os.path.join(st.session_state.pj_path, st.session_state.crop_tgt_event, st.session_state.crop_tgt_img_type, "raw.png")
#     if os.path.exists(img_path):
#         image = Image.open(img_path)
#         image_info = st.session_state.pj_timetable_master[
#             (st.session_state.pj_timetable_master["event_no"]==int(st.session_state.crop_tgt_event.split("_")[1]))
#             & (st.session_state.pj_timetable_master["image_type"]==st.session_state.crop_tgt_img_type)
#         ].iloc[0]

#         with st.container():# タイムテーブルに関係する領域を切り出す
#             st.markdown("""###### タイムテーブルに関係する領域を切り出す""")
#             st.info(
# """まず、タイムテーブルに関係する領域の切り出しを行います。  
# 必要十分な領域を指定してください。  
# ・ライムライト式の場合、時間軸の情報は含める必要があります（後で使います）  
# ・上下にステージ名などの情報は含まれていても構いません  
# ・そのまま一枚の画像として均等割を行う場合にはそれに適した領域に切り出してください  
# """)
#             col_cropimage_first = st.columns(2)
#             with col_cropimage_first[0]:
#                 st.session_state.cropped_image = st_cropper(image)
#             with col_cropimage_first[1]:
#                 st.image(st.session_state.cropped_image,use_container_width=True)

#         with st.container():# 「均等割」できるようにタイムテーブル領域を分割する
#             st.markdown("""###### 「均等割」できるようにタイムテーブル領域を分割する""")
#             st.info(
# """次に、タイムテーブルが均等でない場合、各々が均等な領域になるよう分割します。  
# 分割が不要な場合、この工程を無視して進んでください。  
# 分割が必要な場合、「分割を行う」をONにして分割を行ってください。  
  
# ドラッグで縦線を入力し、それに従って分割を行います。  
# ・実際には、縦線ではなく縦線の始点の横位置で分割します。線はイメージです  
# ・なので、斜めな線でも、画像の上から下まで線を引かなくとも、問題はありません  
# ・右側に表示される分割結果を参照して、適宜線を修正してください  
# ・なお画像の表示比率がおかしかったり欠けていたりするかもしれませんが、一旦気にせず進めてください  
  
# 分割が終了したら、各領域に何個のステージが含まれてているかを入力します。  
# ・タイムテーブル以外の情報（時間軸領域など）しか含まない領域を無視したい場合、ステージ数を0にすることで実現できます  
# ・なお次のステップにおいて、均等割ではなく矩形抽出による自動調整分割もできるので、必ずしも時間軸領域を無視せずとも動きますが、精度は100%ではないので、無視できるならしたほうがよいです  
# ・このステップで完全人力でステージ1つずつへの分割を行うこともできます。時間はかかりますがそれが一番確実です  
# """)
#             line_cropping = st.toggle("分割を行う")
#             if line_cropping:
#                 col_cropimage_line = st.columns(2)
#                 with col_cropimage_line[0]:
#                     canvas_width = 800
#                     canvas_rate = canvas_width/st.session_state.cropped_image.size[0]
#                     canvas_height = int(st.session_state.cropped_image.size[1]*canvas_rate)
#                     resized_image = st.session_state.cropped_image.resize((canvas_height, canvas_width))
#                     canvas_result = st_canvas(
#                         fill_color="rgba(255, 165, 0, 0.3)",  # Fixed fill color with some opacity
#                         stroke_width=3,
#                         # stroke_color=stroke_color,
#                         # background_color=bg_color,
#                         background_image=resized_image,
#                         update_streamlit=True,
#                         width=canvas_width,
#                         height=canvas_height,
#                         drawing_mode="line",
#                         # point_display_radius=1,
#                         key="crop_line",
#                     )
#                 with col_cropimage_line[1]:
#                     if canvas_result.json_data is not None:
#                         objects = pd.json_normalize(canvas_result.json_data["objects"])
#                         # for col in objects.select_dtypes(include=["object"]).columns:
#                         #     objects[col] = objects[col].astype("str")
#                         # st.dataframe(objects)

#                         st.session_state.line_cropped_images = []
#                         line_x_before = 0
#                         for i,row in objects.iterrows():
#                             line_x = row["left"]+row["width"]/2
#                             st.session_state.line_cropped_images.append(st.session_state.cropped_image.crop((line_x_before/canvas_rate, 0, line_x/canvas_rate, st.session_state.cropped_image.size[1])))
#                             line_x_before = line_x
#                         st.session_state.line_cropped_images.append(st.session_state.cropped_image.crop((line_x_before/canvas_rate, 0, st.session_state.cropped_image.size[0], st.session_state.cropped_image.size[1])))

#                         col_cropimage_line_after = st.columns([max(img.size[0],st.session_state.cropped_image.size[0]/len(st.session_state.line_cropped_images)/2) for img in st.session_state.line_cropped_images])
#                         for i,line_cropped_image in enumerate(st.session_state.line_cropped_images):
#                             with col_cropimage_line_after[i]:
#                                 st.image(line_cropped_image)
#                                 st.number_input("ステージ数",0,value=1,step=1,key="stage_num_{}".format(i))
#             else:
#                 st.session_state.line_cropped_images = [st.session_state.cropped_image]
#                 st.number_input("ステージ数",1,step=1,key="stage_num_0")

#         with st.container():# 各ステージに1枚の画像が対応するよう分割する
#             st.markdown("""###### 各ステージに1枚の画像が対応するよう分割する""")
#             st.info(
# """最後に、ステージ一つひとつに一枚の画像が対応するよう、画像を分割します。  
# 画像の分割は、均等でないタイムテーブル領域を分割した場合にはそれぞれに対して、  
# 分割していない場合にはタイムテーブルに関係する領域として切り出した1枚に対して、  
# 指定されたステージ数への分割が実施されます。  
# 分割の方法は下記の通りです。どちらかを選んでボタンを押してください。  
# 1. OCRにより矩形を抽出し、ステージの幅を推定して分割する  
# 1. 画像を均等な横幅になるよう分割する  

# なお先述の通りステージ数が1の場合はそのまま一つのステージの画像として採用されます。  
# """)
#             # st.button("ステージごとの画像を抽出",on_click=get_image_eachstage_byocr,args=(cropped_image, st.session_state.stage_num))
#             st.button("OCR自動分割",on_click=get_image_eachstage_for_linecroppedimage_byocr)
#             st.button("横幅均等分割",on_click=get_image_eachstage_for_linecroppedimage_byevenly)

#             if len(st.session_state.images_eachstage)>0:
#                 col_cropimage_eachstage = st.columns([img.size[0] for img in st.session_state.images_eachstage])
#                 for i,eachstage_image in enumerate(st.session_state.images_eachstage):
#                     with col_cropimage_eachstage[i]:
#                         st.image(eachstage_image)

#                 st.button("ステージごとの画像分割を確定",on_click=determine_image_eachstage, type="primary")
st.divider()

timetable_ocr = st.container()#タイテ画像の読み取り
with timetable_ocr:
    st.markdown("""#### ④タイムテーブル画像の読み取り""")
    event_list = get_event_name_list()
    with st.expander("まとめて読み取りを実施"):
        st.info("""ライムライト式の画像の時刻推定も同時にまとめて行えますが、画像によって時刻の基準位置が違う場合が多いので、できるだけ先に時刻の読み取りだけ別途それぞれの画像で行うことを強く推奨します。""")
        col_toghether_ocr = st.columns([1,1,3])
        with col_toghether_ocr[0]:
            for i,event_name in enumerate(event_list):
                event_type_list = get_event_type_list(i)
                for event_type in event_type_list:
                    image_info = st.session_state.project_info_json["event_detail"][i]["timetables"][event_type]
                    stage_num = image_info["stage_num"]
                    if stage_num>0:
                        st.checkbox(event_list[i]+"/"+event_type,key="together_"+event_list[i]+"/"+event_type,value=True)
        with col_toghether_ocr[1]:
            st.checkbox("ステージ名の読み取り",key="together_ocr_stage",value=True)
            st.checkbox("タイムテーブルの読み取り",key="together_ocr_timetable",value=True)
            st.checkbox("グループ名の修正（マスタ参照）",key="toghther_correct",value=True)
            st.checkbox("既に確定したタイテ種別で採用したグループ名一覧の中からグループ名を選ぶ",key="correct_idolname_in_confirmed_list_toghther",help="""例えばライブのタイムテーブルを先に作成し、後から特典会のタイムテーブルを作成する際に、ライブのタイムテーブルデータで「グループ名_採用」に入力したグループ名の一覧を候補として、特典会のタイムテーブルデータでも「グループ名_採用」への修正を行うことが出来ます。  
この処理はイベントごとに切り分けて行われるため、day1はday1の中で候補を用意してグループ名を修正し、day2はday2でまた別になります。  
ライブと特典会、あるいは他の種別についてはどのような順番でもよく、「全種別を通じて既に『グループ名_修正』に入力されているグループ一覧」が候補になります。  
どの種別においても一つもグループ名を確定していない場合は、通常通り全グループリストから出力されます。""")
        with col_toghether_ocr[2]:
            st.text_input("ステージ名読み取りの追加指示",key="ocr_stage_user_prompt_together")
            st.text_input("タイムテーブル読み取りの追加指示",key="ocr_user_prompt_together")
            st.button("まとめて実行",on_click=get_timetabledata_together)

    st.selectbox("イベント", event_list,index=0,key="ocr_tgt_event",on_change=set_ocr_image)#変わった時にst.session_state.timeline_eachstageなどをリセット
    ocr_tgt_event_no = get_event_no_by_event_name(st.session_state.ocr_tgt_event)
    event_type_list = get_event_type_list(ocr_tgt_event_no)
    if len(event_type_list) == 0:
        st.warning("画像を登録するか他のイベントを選択してください")
        st.stop()
    st.selectbox("種別", event_type_list,index=0,key="ocr_tgt_img_type",on_change=set_ocr_image)#同上
    img_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "raw.png")
    if os.path.exists(img_path):
        image = Image.open(img_path)
        st.session_state.ocr_tgt_image_info = st.session_state.project_info_json["event_detail"][ocr_tgt_event_no]["timetables"][st.session_state.ocr_tgt_img_type]
        st.session_state.ocr_tgt_stage_num = st.session_state.ocr_tgt_image_info["stage_num"]
        if st.session_state.ocr_tgt_stage_num<=0:
            st.warning("各ステージの画像を確定してください")
            st.stop()
#         if st.session_state.ocr_tgt_image_info["format"]=="ライムライト式":# 時間軸の設定（ライムライト式の場合のみ実施）
#             cropped_img_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "raw_cropped.png")
#             if os.path.exists(cropped_img_path):# 時間軸の設定
#                 cropped_image = Image.open(cropped_img_path)
#                 with st.container():
#                     st.markdown("""###### 時間軸の設定""")#（ライムライト式の場合のみ実施）→読み取り後の画像生成のため全体で実施
#                     st.info(
# """時間軸の基準となる位置を指定してください。  
# ・横幅は特に関係ないので、適当で大丈夫です  
# ・縦位置は出演枠が存在している領域でセットしてください。余白部分は微妙に間隔が違ったりします  
# ・ドラッグした領域の上端と下端に対応する時刻を入力して、画像上のどの位置がどの時刻か対応づけます  
# ・ライムライト式の画像では、この情報を元に時刻の読み取りを行います  
# ・全ての画像形式で、読み取り結果の画像化においてこの情報を元に時間軸を合わせます  
# """)
#                     col_timeaxis = st.columns([2,1])
#                     with col_timeaxis[0]:
#                         col_timeaxis_setting = st.columns(2)
#                         with col_timeaxis_setting[0]:
#                             box_color = st.color_picker(label="Box Color", value='#0000FF', key="timeaxis_box_coler")
#                         with col_timeaxis_setting[1]:
#                             stroke_width = st.number_input(label="Box Thickness", value=1, step=1, key="timeaxs_stroke_width")
#                         rect = st_cropper(cropped_image,
#                                         box_color=box_color,
#                                         stroke_width=stroke_width,
#                                         return_type="box")
#                         left, top, width, height = tuple(map(int, rect.values()))
#                     with col_timeaxis[1]:
#                         # st.time_input("開始時間", value=dttime(10), key=None, step=300)
#                         # st.time_input("終了時間", value=dttime(20), key=None, step=300)
#                         st.slider('開始時間', value=dttime(10), key="time_start", step=timedelta(minutes=5))
#                         st.slider('終了時間', value=dttime(20), key="time_finish", step=timedelta(minutes=5))
#                         total_duration = (datetime(2024,1,1,st.session_state.time_finish.hour,st.session_state.time_finish.minute)-datetime(2024,1,1,st.session_state.time_start.hour,st.session_state.time_start.minute)).seconds/60
#                         if st.session_state.time_finish.hour < st.session_state.time_start.hour:
#                             st.warning("終了時間が開始時間よりも早くなっています。深夜イベントなどで日を跨ぐ場合はそのまま実行可能ですが、そうでない場合は修正してください。")
#                         st.session_state.start_pix = top
#                         st.session_state.total_pix = height
#                         st.session_state.total_duration = total_duration 

#                     with col_timeaxis[0]:
#                         # canvas_height = st.number_input("表示する縦幅",min_value=100,step=100,value=800)
#                         canvas_height = 800
#                         canvas_rate = canvas_height/cropped_image.size[1]
#                         canvas_width = int(cropped_image.size[0]*canvas_rate)
#                         resized_image = cropped_image.resize((canvas_width, canvas_height))
#                         canvas_result = st_canvas(
#                             fill_color="rgba(255, 165, 0, 0.3)",  # Fixed fill color with some opacity
#                             stroke_width=1,
#                             # stroke_color=stroke_color,
#                             # background_color=bg_color,
#                             background_image=resized_image,
#                             update_streamlit=True,
#                             width=canvas_width,
#                             height=canvas_height,
#                             drawing_mode="rect",
#                             # point_display_radius=1,
#                             key="time_axis_detect"
#                         )
#                     with col_timeaxis[1]:
#                         # st.time_input("開始時間", value=dttime(10), key=None, step=300)
#                         # st.time_input("終了時間", value=dttime(20), key=None, step=300)
#                         st.slider('開始時間', value=dttime(10), key="time_start", step=timedelta(minutes=5))
#                         st.slider('終了時間', value=dttime(20), key="time_finish", step=timedelta(minutes=5))
#                         total_duration = (datetime(2024,1,1,st.session_state.time_finish.hour,st.session_state.time_finish.minute)-datetime(2024,1,1,st.session_state.time_start.hour,st.session_state.time_start.minute)).seconds/60
#                         if st.session_state.time_finish.hour < st.session_state.time_start.hour:
#                             st.warning("終了時間が開始時間よりも早くなっています。深夜イベントなどで日を跨ぐ場合はそのまま実行可能ですが、そうでない場合は修正してください。")
#                         if canvas_result.json_data is not None:
#                             objects = pd.json_normalize(canvas_result.json_data["objects"])
#                             # for col in objects.select_dtypes(include=["object"]).columns:
#                             #     objects[col] = objects[col].astype("str")
#                             # st.dataframe(objects)
#                             try:
#                                 resized_start_pix = objects.iloc[0]["top"]
#                                 resized_total_pix = objects.iloc[0]["height"]
#                                 st.session_state.start_pix = resized_start_pix/canvas_rate
#                                 st.session_state.total_pix = resized_total_pix/canvas_rate
#                                 st.session_state.total_duration = total_duration 

#                                 # from PIL import ImageDraw
#                                 # lines = [
#                                 #     ((0, resized_start_pix/canvas_rate), (cropped_image.size[0], resized_start_pix/canvas_rate)),  # (x1, y1), (x2, y2)
#                                 #     ((0, (resized_start_pix+resized_total_pix)/canvas_rate), (cropped_image.size[0], (resized_start_pix+resized_total_pix)/canvas_rate))
#                                 # ]

#                                 # # 直線を描画する関数
#                                 # def draw_lines_on_image(image, lines, color="red", width=3):
#                                 #     draw = ImageDraw.Draw(image)
#                                 #     for line in lines:
#                                 #         draw.line(line, fill=color, width=width)
#                                 #     return image

#                                 # # 画像のコピーを作成し、直線を描画
#                                 # image_with_lines = cropped_image.copy()
#                                 # image_with_lines = draw_lines_on_image(image_with_lines, lines)

#                                 # # 直線を描画した画像を表示
#                                 # st.image(image_with_lines,use_container_width=True)

#                             except IndexError:
#                                 st.warning("画像上で時間範囲をドラッグしてください")

        with st.container():# 各ステージ情報の読み取り
            st.markdown("""###### 各ステージ情報の読み取り""")
            stage_name_list = get_stage_name_list(ocr_tgt_event_no,st.session_state.ocr_tgt_img_type)
            if st.session_state.ocr_tgt_image_info["format"]=="ライムライト式":
                st.info(
"""アーティストごとに時間が書いていないタイムテーブルの場合、横線を検出して時間を推定します。  
その後推定した時刻を画像に書き込み、そこからタイムテーブル情報を生成します。  
画像の色合いによって横線の検出の難易度が変わるので、パラメータをいじってください。  
- 「エッジ抽出の閾値」は、特定の色と色の間の線が検出できていない場合に下げると良いです。劇的に下げてもOKです。
- 「ハフ変換の閾値」は、全体的に検出できていない場合に下げると良いです。
- 「抽出したエッジを伸ばす際の閾値」は、検出された線が途切れ途切れになったり短かったりする場合に下げると良いです。
- 「ハフ変換で許容する線分の飛び」は、どの程度の破線を許容するかのパラメータです。連続した文字などの上に不要な線分が検出されてしまっている場合に上げると良いです。
- 「抽出線分の長さ（元画像の縦に対する比率）」は、短い線分を除くためのパラメータです。不要な短い線分が検出されている場合に上げると良いです。
- 「同一視する線分の許容誤差幅」は、線分の位置がほぼ同じ場合にそれらをまとめ上げる範囲についてのパラメータです。
- 「無視する時間幅」パラメータは、その値（分）以下の長さしかない横線と横線の間には時刻を書き込まないようにするものです  
  
※「全ステージの画像を読み取りを実施」をいきなり押すと、先に横線の時刻の読み取りを裏で実行してからタイムテーブル情報の読み取りを行います  
※「全ステージの横線の時刻の読み取りを実施」を押して、正しく時刻が取得できたことを画像で確認してから、「全ステージの画像を読み取りを実施」を押すことを推奨します
""")
                st.text_input("ステージ名読み取りの追加指示",key="ocr_stage_user_prompt")
                st.button("ステージ名の読み取りを実施",on_click=get_stagelist,args=(st.session_state.ocr_stage_user_prompt,),type="primary")
                st.button("全ステージの横線の時刻の読み取りを実施",on_click=detect_timeline_eachstage,type="primary")
                st.text_input("タイムテーブル読み取りの追加指示",key="ocr_user_prompt")
                st.button("全ステージのタイムテーブルを読み取りを実施",on_click=get_timetabledata_eachstage_notime,args=(st.session_state.ocr_user_prompt,),type="primary")
                # st.number_input("ステージ番号",0,st.session_state.ocr_tgt_stage_num-1,key="ocr_tgt_stage_no")
                # st.selectbox("ステージ",stage_name_list,index=0,key="ocr_tgt_stage_no")
                # st.button("あるステージのみ横線の時刻の読み取りを実施",on_click=detect_timeline_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,))
                # st.button("あるステージのみタイムテーブルの読み取りを実施",on_click=get_timetabledata_onlyonestage_notime,args=(st.session_state.ocr_tgt_stage_no,st.session_state.ocr_user_prompt))
                with st.expander("時間ライン抽出のパラメータ"):
                    # st.slider('白黒二値化の閾値（0はグレースケールのまま実施）', value=0, min_value=0, max_value=255, step=1, key="y_binary_threshold")#, on_change=detect_timeline_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,))
                    st.slider('無視する時間幅（分）（以下）', value=5, min_value=0, max_value=60, step=5, key="y_ignoretime_threshold")#, on_change=detect_timeline_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,))
                    st.slider('エッジ抽出の閾値', value=150, min_value=1, max_value=500, step=1, key="y_edge_threshold_2")#, on_change=detect_timeline_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,))
                    st.slider('抽出したエッジを伸ばす際の閾値（エッジ抽出の閾値以下にする）', value=80, min_value=1, max_value=500, step=1, key="y_edge_threshold_1")#, on_change=detect_timeline_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,))
                    st.slider('ハフ変換の閾値', value=60, min_value=1, max_value=500, step=1, key="y_hough_threshold")#, on_change=detect_timeline_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,))
                    st.slider('ハフ変換で許容する線分の飛び', value=1, min_value=0, max_value=100, step=1, key="y_hough_gap")#, on_change=detect_timeline_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,))
                    st.slider('抽出線分の長さ（元画像の縦に対する比率）', value=0.05, min_value=0.0, max_value=1.0, step=0.01, key="y_minlength_rate")#, on_change=detect_timeline_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,))
                    st.slider('同一視する線分の許容誤差幅', value=5, min_value=1, max_value=30, step=1, key="y_identify_interval")#, on_change=detect_timeline_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,))
                tmp_timeline = st.container()#暫定
            elif st.session_state.ocr_tgt_image_info["format"]=="特典会併記":
                st.text_input("ステージ名読み取りの追加指示",key="ocr_stage_user_prompt")
                st.button("ステージ名の読み取りを実施",on_click=get_stagelist,args=(st.session_state.ocr_stage_user_prompt,),type="primary")
                st.text_input("タイムテーブル読み取りの追加指示",key="ocr_user_prompt")
                st.button("全ステージのタイムテーブルをそれぞれ読み取り実施",on_click=get_timetabledata_withtokutenkai_eachstage,args=(st.session_state.ocr_user_prompt,),type="primary")
                # st.number_input("ステージ番号",0,st.session_state.ocr_tgt_stage_num-1,key="ocr_tgt_stage_no")
                # st.button("あるステージのみタイムテーブルの読み取りを実施",on_click=get_timetabledata_withtokutenkai_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,st.session_state.ocr_user_prompt))
                st.button("全ステージの特典会ブース名にステージ名を接頭辞として付与する",on_click=booth_name_add_prefix_eachstage)
            else:
                st.text_input("ステージ名読み取りの追加指示",key="ocr_stage_user_prompt")
                st.button("ステージ名の読み取りを実施",on_click=get_stagelist,args=(st.session_state.ocr_stage_user_prompt,),type="primary")
                st.text_input("タイムテーブル読み取りの追加指示",key="ocr_user_prompt")
                st.button("全ステージのタイムテーブルをそれぞれ読み取り実施",on_click=get_timetabledata_eachstage,args=(st.session_state.ocr_user_prompt,),type="primary")
                # st.number_input("ステージ番号",0,st.session_state.ocr_tgt_stage_num-1,key="ocr_tgt_stage_no")
                # st.button("あるステージのみタイムテーブルの読み取りを実施",on_click=get_timetabledata_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,st.session_state.ocr_user_prompt))

            #個別ステージ
            with st.container():#表示設定
                if st.session_state.ocr_tgt_image_info["format"]=="ライムライト式":
                    img_path_tmp = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}_addtime.png".format(0))
                    if not os.path.exists(img_path_tmp):
                        img_path_tmp = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.png".format(0))
                else:
                    img_path_tmp = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.png".format(0))
                if os.path.exists(img_path_tmp):
                    image_tmp = Image.open(img_path_tmp)
                img_path_tmp_output = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}_timetable.png".format(0))
                if os.path.exists(img_path_tmp_output):
                    image_tmp_output = Image.open(img_path_tmp_output)

                
                ocr_eachimage_width_default = 40#int((image_tmp.size[0])*100/(image_tmp.size[0]+600))
                # output_eachimage_width_default = int(ocr_eachimage_width_default*100/(100-ocr_eachimage_width_default))
                st.slider('タイテ元画像の表示幅（%）', value=ocr_eachimage_width_default, min_value=1, max_value=99, step=1, key="ocr_eachimage_width")#, on_change=detect_timeline_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,))
                # ocr_slider_col = st.columns([st.session_state.ocr_eachimage_width,100-st.session_state.ocr_eachimage_width])
                # with ocr_slider_col[1]:
                #     st.slider('タイテ出力画像の表示幅（%）', value=st.session_state.ocr_eachimage_width, min_value=0, max_value=100-st.session_state.ocr_eachimage_width, step=1, key="output_eachimage_width")#, on_change=detect_timeline_onlyonestage,args=(st.session_state.ocr_tgt_stage_no,))
                ocr_show_setting_col = st.columns([1,1,1,1,1])
                with ocr_show_setting_col[0]:
                    st.checkbox("画像の縦スクロール表示", value=True, key="ocr_eachimage_scroll")
                with ocr_show_setting_col[1]:
                    st.checkbox("読み取り結果の画像の時間軸を元画像に合わせる", value=True, key="ocr_output_picture_time_match",on_change=output_timetable_picture_eachstage , help="""「時間軸の設定」で指定したラインに合わせて画像を生成します。  
このチェックボックスのオンオフまたはステージの編集結果の保存で画像が更新されます。""")

            stage_tabs = st.tabs(stage_name_list)
            st.session_state.df_timetables = []
            for i in range(st.session_state.ocr_tgt_stage_num):
                with stage_tabs[i]:
                    if st.session_state.ocr_tgt_image_info["format"]=="ライムライト式":
                        img_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}_addtime.png".format(i))
                        if not os.path.exists(img_path):
                            img_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.png".format(i))
                    else:
                        img_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.png".format(i))
                    if os.path.exists(img_path):
                        image = Image.open(img_path)
                    img_path_output = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}_timetable.png".format(i))
                    if os.path.exists(img_path_output):
                        image_output = Image.open(img_path_output)
                    ocr_col = st.columns([st.session_state.ocr_eachimage_width,100-st.session_state.ocr_eachimage_width])
                    with ocr_col[0]:
                        if st.session_state.ocr_output_picture_time_match:
                            if os.path.exists(img_path_output):
                                st.markdown("""###### タイテ画像+読み取り結果画像""")
                                new_width = image.width + image_output.width
                                new_height = max(image.height, image_output.height)
                                new_image = Image.new("RGB", (new_width, new_height), "white")
                                new_image.paste(image, (0, 0))
                                new_image.paste(image_output, (image.width, 0))
                                if st.session_state.ocr_eachimage_scroll:                        
                                    with st.container(height=800):
                                        st.image(new_image,use_container_width=True)
                                else:
                                    with st.container():
                                        st.image(new_image,use_container_width=True)
                            else:
                                ocr_image_col = st.columns(2)
                                with ocr_image_col[0]:
                                    st.markdown("""###### タイテ画像""")
                                    if st.session_state.ocr_eachimage_scroll:                        
                                        with st.container(height=700):
                                            if os.path.exists(img_path):
                                                st.image(image,use_container_width=True)
                                    else:
                                        with st.container():
                                            if os.path.exists(img_path):
                                                st.image(image,use_container_width=True)
                                with ocr_image_col[1]:
                                    st.markdown("""###### 読み取り結果画像""")
                                    st.warning("各ステージの読み取りを行うとタイムテーブル画像が表示されます")
                        else:
                            ocr_image_col = st.columns(2)
                            with ocr_image_col[0]:
                                st.markdown("""###### タイテ画像""")
                                if st.session_state.ocr_eachimage_scroll:                        
                                    with st.container(height=700):
                                        if os.path.exists(img_path):
                                            st.image(image,use_container_width=True)
                                else:
                                    with st.container():
                                        if os.path.exists(img_path):
                                            st.image(image,use_container_width=True)
                            with ocr_image_col[1]:
                                st.markdown("""###### 読み取り結果画像""")
                                if os.path.exists(img_path_output):
                                    if st.session_state.ocr_eachimage_scroll:                        
                                        with st.container(height=700):
                                            st.image(image_output,use_container_width=True)
                                    else:
                                        with st.container():
                                            st.image(image_output,use_container_width=True)
                                else:
                                    st.warning("各ステージの読み取りを行うとタイムテーブル画像が表示されます")

                    with ocr_col[1]:
                        st.markdown("""###### 読み取り結果""")
                        stage_name = get_stage_name(ocr_tgt_event_no,st.session_state.ocr_tgt_img_type,i)
                        st.text_input("ステージ名",value=stage_name,key="stage_name_stage{}".format(i))
                        if st.session_state.ocr_tgt_image_info["format"]=="ライムライト式":
                            st.button("このステージの横線の時刻の読み取りを実施",on_click=detect_timeline_onlyonestage,args=(i,),key="button_ocr_timeline_stage{}".format(i))
                            st.text_input("タイムテーブル読み取りの追加指示",key="ocr_user_prompt_stage{}".format(i))
                            st.button("このステージのタイムテーブルの読み取りを実施",on_click=get_timetabledata_onlyonestage_notime,args=(i,st.session_state["ocr_user_prompt_stage{}".format(i)]),key="button_ocr_stage{}".format(i))
                        elif st.session_state.ocr_tgt_image_info["format"]=="特典会併記":
                            st.text_input("タイムテーブル読み取りの追加指示",key="ocr_user_prompt_stage{}".format(i))
                            st.button("このステージのタイムテーブルの読み取りを実施",on_click=get_timetabledata_withtokutenkai_onlyonestage,args=(i,st.session_state["ocr_user_prompt_stage{}".format(i)]),key="button_ocr_stage{}".format(i))
                            st.button("特典会ブース名にステージ名を接頭辞として付与する",on_click=booth_name_add_prefix_onlyonestage,args=(i,),key="booth_name_add_prefix_stage{}".format(i))
                        else:
                            st.text_input("タイムテーブル読み取りの追加指示",key="ocr_user_prompt_stage{}".format(i))
                            st.button("このステージのタイムテーブルの読み取りを実施",on_click=get_timetabledata_onlyonestage,args=(i,st.session_state["ocr_user_prompt_stage{}".format(i)]),key="button_ocr_stage{}".format(i))
                        json_path = os.path.join(st.session_state.pj_path, st.session_state.ocr_tgt_event, st.session_state.ocr_tgt_img_type, "stage_{}.json".format(i))
                        if os.path.exists(json_path):
                            with open(json_path, encoding="utf-8") as f:
                                return_json = json.load(f)
                            if st.session_state.ocr_tgt_image_info["format"]=="特典会併記":
                                return_json_df = timetabledata.json_to_df(return_json)
                            else:
                                return_json_df = timetabledata.json_to_df(return_json,tokutenkai=False)
                            # st.dataframe(return_json_df)
                            edited_df = st.data_editor(return_json_df, key="timetabledata_stage{}".format(i), num_rows="dynamic")
                            
                            # edited_df["ステージID"]=i
                            edited_df["ステージ名"]=stage_name
                            st.session_state.df_timetables.append(edited_df)

                            # st.button("このステージのグループ名を修正（マスタ参照）",on_click=idolname_correct_onlyonestage,args=(i,),key="button_correct_idolname_stage{}".format(i))
                            if st.button("このステージのグループ名を修正（マスタ参照）",key="button_correct_idolname_stage{}_confirm".format(i)):
                                st.warning('「グループ名_採用」が上書きされます。本当に処理を実行しますか？')
                                st.button("OK",on_click=idolname_correct_onlyonestage,args=(i,),key="button_correct_idolname_stage{}".format(i))
                            st.button("このステージの編集結果を保存",on_click=save_timetable_data_onlyonestage,args=(i,),key="button_save_timetable_stage{}".format(i))

            # st.button("全ステージのグループ名を修正（マスタ参照）",on_click=idolname_correct_eachstage,key="button_correct_idolname")
            st.checkbox("既に確定したタイテ種別で採用したグループ名一覧の中からグループ名を選ぶ",key="correct_idolname_in_confirmed_list",help="""例えばライブのタイムテーブルを先に作成し、後から特典会のタイムテーブルを作成する際に、ライブのタイムテーブルデータで「グループ名_採用」に入力したグループ名の一覧を候補として、特典会のタイムテーブルデータでも「グループ名_採用」への修正を行うことが出来ます。  
この処理はイベントごとに切り分けて行われるため、day1はday1の中で候補を用意してグループ名を修正し、day2はday2でまた別になります。  
ライブと特典会、あるいは他の種別についてはどのような順番でもよく、「全種別を通じて既に『グループ名_修正』に入力されているグループ一覧」が候補になります。  
どの種別においても一つもグループ名を確定していない場合は、通常通り全グループリストから出力されます。""")
            if st.button("全ステージのグループ名を修正（マスタ参照）",key="button_correct_idolname_confirm"):
                st.warning('「グループ名_採用」が上書きされます。本当に処理を実行しますか？')
                st.button('OK', on_click=idolname_correct_eachstage,key="button_correct_idolname")
            st.button("全ステージの編集結果を保存",on_click=save_timetable_data_eachstage,key="button_save_timetable")

            # st.button("全ステージのデータを表形式で出力",on_click=get_all_stage_info,type="primary")
            # all_stage_info = st.container()
st.divider()

timetable_change = st.container()#タイテ画像の追加・変更
with timetable_change:
    st.markdown("""#### ⑤タイムテーブル画像の追加・変更
- 読み取りを行った後にタイムテーブルが変更になった場合に、画像の変更点のみを読み取って修正してくれる機能をいつか実装します""")
    timetable_compare_setting_col = st.columns(2)
    timetable_compare_col = st.columns(2)
    with timetable_compare_setting_col[0]:
        st.file_uploader("読み取りたいタイムテーブル画像をアップロードしてください。"
                                , type=["jpg", "jpeg", "png", "jfif"]
                                , key="uploaded_image_updated")
    with timetable_compare_col[0]:
        if st.session_state.uploaded_image_updated is not None:
            st.image(
                st.session_state.uploaded_image_updated,
                use_container_width=True
            )
    with timetable_compare_setting_col[1]:
        st.selectbox("イベント", event_list,index=0,key="diff_tgt_event")
        diff_tgt_event_no = get_event_no_by_event_name(st.session_state.diff_tgt_event)
        event_type_list = get_event_type_list(diff_tgt_event_no)
        if len(event_type_list) == 0:
            st.warning("画像を登録するか他のイベントを選択してください")
            st.stop()
        st.selectbox("種別", event_type_list,index=0,key="diff_tgt_img_type")#同上
        st.button("差分画像を出力する",on_click=output_difference_image,args=(st.session_state.uploaded_image_updated,))

st.divider()

timetable_output = st.container()#タイテ情報の出力
st.session_state.output_df = {}
with timetable_output:
    st.markdown("""#### ⑥タイムテーブル情報の出力""")

    event_list = get_event_name_list()
    event_tabs = st.tabs(event_list)
    for i, event_tab in enumerate(event_tabs):
        st.session_state.output_df[event_list[i]]={}
        with event_tab:#イベントごとに出力を作る
            edit_tgt_event_no = get_event_no_by_event_name(event_list[i])
            event_type_list = get_event_type_list(edit_tgt_event_no)
            output_path =  os.path.join(st.session_state.pj_path, event_list[i])
            if os.path.exists(os.path.join(output_path, "master_stage.csv")):
                stage_master_df = pd.read_csv(os.path.join(output_path, "master_stage.csv"), index_col=0)
                stage_master = json.loads(stage_master_df.T.to_json())
                tokutenkai_timetable = []
                stage_id = int(max(stage_master_df.index))+1
            else:
                stage_master = {}
                tokutenkai_timetable = []
                stage_id = 0
            if os.path.exists(os.path.join(output_path, "master_idolname.csv")):
                idolname_master_df = pd.read_csv(os.path.join(output_path, "master_idolname.csv"), index_col=0).rename(columns={"グループ名":"グループ名_採用"})
                artist_id = int(max(idolname_master_df.index))+1
            else:
                idolname_master_df = pd.DataFrame(columns=["グループID","グループ名_採用"]).set_index("グループID")
                artist_id = 0
            # if os.path.exists(os.path.join(output_path, "turn_id_data.csv")):
            #     live_master_df = pd.read_csv(os.path.join(output_path, "turn_id_data.csv"), index_col=0)

            stage_master_tokutenkai = {}
            event_timetable_all = []
            for event_type in event_type_list:#全種別をまとめる
                tgt_event_type_info = st.session_state.project_info_json["event_detail"][edit_tgt_event_no]["timetables"][event_type]
                stage_name_list = get_stage_name_list(edit_tgt_event_no,event_type)
                tokutenkai_flg = event_type=="特典会"
                for j in range(tgt_event_type_info["stage_num"]):
                    json_path = os.path.join(st.session_state.pj_path, event_list[i], event_type, "stage_{}.json".format(j))
                    if os.path.exists(json_path):
                        with open(json_path, encoding="utf-8") as f:
                            edit_tgt_json = json.load(f)
                        if tgt_event_type_info["format"]=="特典会併記":#特典会併記タイテは分離してライブのみをまず扱う
                            df_edit_tgt = timetabledata.json_to_df(edit_tgt_json, tokutenkai=True)
                            df_edit_live, df_edit_tokutenkai = timetabledata.devide_df_live_tokutenkai(df_edit_tgt)
                            # st.dataframe(df_edit_live)
                            # st.dataframe(df_edit_tokutenkai)
                            tokutenkai_timetable.append(df_edit_tokutenkai)
                        else:
                            df_edit_live = timetabledata.json_to_df(edit_tgt_json, tokutenkai=False)
                        df_edit_live = df_edit_live.copy()
                        for k,v in stage_master.items():#既にID確定済のステージの場合はそれを採用
                            if v["ステージ名"]==stage_name_list[j]:
                                this_stage_id = k
                                break
                        else:
                            this_stage_id = stage_id
                            stage_master[this_stage_id]={"ステージ名":stage_name_list[j],"特典会フラグ":tokutenkai_flg}
                            stage_id += 1
                        if "ステージID" not in df_edit_live.columns:
                            df_edit_live["ステージID"]=None
                            df_edit_live["ステージ名"]=None
                        df_edit_live.loc[:,"ステージID"]=this_stage_id
                        df_edit_live.loc[:,"ステージ名"]=stage_name_list[j]
                        event_timetable_all.append(df_edit_live)
            if len(tokutenkai_timetable)>0:#特典会併記タイテの場合の特典会情報の処理
                df_tokutenkai = pd.concat((tokutenkai_timetable)).reset_index(drop=True)
                if "ステージID" in df_tokutenkai.columns:
                    df_tokutenkai = df_tokutenkai.drop(columns=["ステージID"])
                booth_name_list = df_tokutenkai["ステージ名"].drop_duplicates().tolist()
                for j, booth_name in enumerate(booth_name_list):
                    for k,v in stage_master.items():
                        if v["ステージ名"]==booth_name:
                            this_stage_id = k
                            break
                    else:
                        this_stage_id = stage_id
                        stage_master[this_stage_id]={"ステージ名":booth_name,"特典会フラグ":True}
                        stage_id += 1
                    stage_master_tokutenkai[this_stage_id]={"ステージ名":booth_name,"特典会フラグ":True}
            if len(event_timetable_all)==0:
                st.stop()
            df_stage = pd.DataFrame.from_dict(stage_master, orient='index')
            df_stage.index.name = "ステージID"
            df_stage_tokutenkai = pd.DataFrame.from_dict(stage_master_tokutenkai, orient='index')
            df_stage_tokutenkai.index.name = "ステージID"
            df_live = pd.concat((event_timetable_all)).reset_index(drop=True)
            if len(tokutenkai_timetable)>0:
                df_tokutenkai = pd.merge(df_tokutenkai,df_stage_tokutenkai.reset_index().drop("特典会フラグ",axis=1),on="ステージ名",how="left")
                df_live = pd.concat((df_live, df_tokutenkai)).reset_index(drop=True)
            ## ここまでがステージマスタ作成
            ## ここからアーティストマスタ作成
            df_idolname = pd.DataFrame(df_live["グループ名_採用"].drop_duplicates().sort_values().reset_index(drop=True))
            df_idolname = df_idolname[~df_idolname["グループ名_採用"].isin(idolname_master_df["グループ名_採用"])].reset_index(drop=True)
            df_idolname.index = df_idolname.index + artist_id
            df_idolname.index.name = 'グループID'
            df_idolname = pd.concat((idolname_master_df,df_idolname))
            ## ここまでがアーティストマスタ作成
            ## ここから出番データ作成
            if "グループID" in df_live.columns:
                df_live = df_live.drop(columns=["グループID"])
            df_live = pd.merge(df_live,df_idolname.reset_index(),on="グループ名_採用",how="left").rename(columns={"グループ名":"グループ名_raw","グループ名_採用":"グループ名"})
            if "出番ID" in df_live:
                turn_id = int(df_live["出番ID"].max())+1
                for row_id, row in df_live.iterrows():
                    try:
                        df_live.loc[row_id,"出番ID"]=int(row["出番ID"])
                    except ValueError:
                        df_live.loc[row_id,"出番ID"]=turn_id
                        turn_id+=1
                df_live["出番ID"]=df_live["出番ID"].astype(int)
                df_live.set_index("出番ID",inplace=True)
            else:
                df_live.reset_index(drop=True,inplace=True)
                df_live.index.name = "出番ID"
            output_cols = st.columns([1,1,3])
            with output_cols[0]:
                st.dataframe(df_stage)
            with output_cols[1]:
                st.dataframe(df_idolname)
            with output_cols[2]:
                st.dataframe(df_live[["ライブ_from","ライブ_長さ(分)","グループID","ステージID","グループ名_raw","グループ名","ステージ名","備考"]])
            st.session_state.output_df[event_list[i]]["stage"]=df_stage
            st.session_state.output_df[event_list[i]]["idolname"]=df_idolname
            st.session_state.output_df[event_list[i]]["live"]=df_live[["ライブ_from","ライブ_長さ(分)","グループID","ステージID","グループ名_raw","グループ名","ステージ名","備考"]]
            # st.dataframe(df_live[["グループID","グループ名","グループ名_採用","ライブ_from","ライブ_to","ライブ_長さ(分)","ステージ名","ステージID","備考"]])

    st.button("IDマスタを確定",on_click=determine_id_master)
    st.button("プロジェクトデータをクラウドにアップロード ※通信料・保存料が発生するので留意",on_click=save_to_s3)
    if st.button("Excelデータを出力",on_click=output_data_for_stella):#全イベントのタイテをシートに分けてExcelで出力
        file_path =  os.path.join(st.session_state.pj_path, "output.xlsx")
        with open(file_path, "rb") as file:
            excel_data = file.read()
        st.download_button("ファイルをダウンロード",data=excel_data, file_name="{}.xlsx".format(st.session_state.pj_name), mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

